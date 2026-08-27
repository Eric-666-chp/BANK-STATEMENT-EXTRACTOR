import re
import csv
import sys
from difflib import SequenceMatcher
from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font

# ================= 公共判定/清洗 =================

DATE_MMDDSHORT = re.compile(
    r'^\s*(?:1[0-2]|0[1-9])/(?:3[01]|[12]\d|0[1-9])\s*$'
)

DATE_MMDDYY = re.compile(
    r'^\s*(?:1[0-2]|0[1-9])/(?:3[01]|[12]\d|0[1-9])/\d{2}\s*$'
)

DATE_MMDDYYYY = re.compile(
    r'^\s*(?:1[0-2]|0[1-9])/(?:3[01]|[12]\d|0[1-9])/\d{4}\s*$'
)

DEFAULT_DATE_ANY_IN_TEXT_PATTERN = (
    r'(?:1[0-2]|0[1-9])/'
    r'(?:3[01]|[12]\d|0[1-9])'
    r'(?:/(?:\d{2}|\d{4}))?'
)

DATE_ANY_IN_TEXT_PATTERN = DEFAULT_DATE_ANY_IN_TEXT_PATTERN

AMOUNT_RE = re.compile(
    r'^\s*(?:'
    r'-\s*\$?\s*[\d,]+\.\d{2}'
    r'|'
    r'\$?\s*-\s*[\d,]+\.\d{2}'
    r'|'
    r'\(?\s*\$?\s*[\d,]+\.\d{2}\s*\)?'
    r')\s*$'
)

PHONE_RE = re.compile(r'^\s*\+?\d[\d\s\-]{6,}\s*$')
STATE_RE = re.compile(r'^\s*[A-Z]{2}\s*$')
DOT_RE = re.compile(r'^\s*\.\s*$')

MERCHANT_CUT_TOKENS = (" DES:", " ID:", " INDN:", " CO ID:")

DATE_IN_TEXT_RE = re.compile(DATE_ANY_IN_TEXT_PATTERN, re.IGNORECASE)

DATE_FULL_RE = re.compile(
    rf'^\s*(?:{DATE_ANY_IN_TEXT_PATTERN})\s*$', re.IGNORECASE
)

DATE_MERCHANT_LINE_RE = re.compile(
    rf'^\s*({DATE_ANY_IN_TEXT_PATTERN})\s+(.*\S)\s*$', re.IGNORECASE
)


def convert_custom_date_format_to_regex(date_format: str) -> str:
    """Convert one or more user date formats into a regex pattern.

    Unified notation used by the UI:
      M       month: numeric or English month name
      MM      numeric month only
      DD      day of month
      YY      two-digit year
      YYYY    four-digit year

    Examples:
      M-DD       -> 4-22
      M DD       -> May 17
      DD M       -> 17 May
      M-DD-YYYY  -> 4-22-2025

    Multiple formats may be separated by commas, Chinese commas, semicolons,
    or vertical bars. Old tokens D, MON, MMM and MONTH remain accepted for
    backward compatibility, but the UI only presents the unified notation.
    """
    raw = str(date_format or "").strip()
    if not raw:
        return DEFAULT_DATE_ANY_IN_TEXT_PATTERN

    formats = [
        part.strip()
        for part in re.split(r'[,，;；|]+', raw)
        if part.strip()
    ]
    if not formats:
        return DEFAULT_DATE_ANY_IN_TEXT_PATTERN

    month_name_pattern = (
        r'(?:JAN(?:UARY)?|FEB(?:RUARY)?|MAR(?:CH)?|APR(?:IL)?|MAY|'
        r'JUN(?:E)?|JUL(?:Y)?|AUG(?:UST)?|SEP(?:T(?:EMBER)?)?|'
        r'OCT(?:OBER)?|NOV(?:EMBER)?|DEC(?:EMBER)?)'
    )
    numeric_month_pattern = r'(?:1[0-2]|0?[1-9])'
    day_pattern = r'(?:3[01]|[12]\d|0?[1-9])'

    # M means Month and can therefore recognize either a numeric month
    # (4) or an English month name (May/September).
    any_month_pattern = rf'(?:{numeric_month_pattern}|{month_name_pattern})'

    token_patterns = {
        "YYYY": r"\d{4}",
        "MONTH": any_month_pattern,
        "MMM": any_month_pattern,
        "MON": any_month_pattern,
        "YY": r"\d{2}",
        "MM": numeric_month_pattern,
        "DD": day_pattern,
        "M": any_month_pattern,
        "D": day_pattern,  # legacy alias
    }
    token_order = ("YYYY", "MONTH", "MMM", "MON", "YY", "MM", "DD", "M", "D")

    def one_format_to_regex(fmt_text: str) -> str:
        fmt = fmt_text.strip().upper()
        result = []
        i = 0

        while i < len(fmt):
            matched = False
            for token in token_order:
                if fmt.startswith(token, i):
                    result.append(token_patterns[token])
                    i += len(token)
                    matched = True
                    break

            if matched:
                continue

            current_char = fmt[i]
            if current_char.isalpha():
                raise ValueError(
                    "无法识别的日期格式。请使用 M、MM、DD、YY、YYYY，"
                    "例如 M-DD、M DD、DD M、M-DD-YYYY。"
                )

            if current_char.isspace():
                while i < len(fmt) and fmt[i].isspace():
                    i += 1
                result.append(r"\s+")
                continue

            result.append(re.escape(current_char))
            i += 1

        return "".join(result)

    patterns = [one_format_to_regex(fmt) for fmt in formats]
    return patterns[0] if len(patterns) == 1 else r"(?:" + "|".join(patterns) + r")"


def infer_date_format_from_sample(sample_text: str) -> str:
    """Infer a unified date format from one user-entered date example.

    Examples:
      4-22       -> M-DD
      04/22/25   -> M/DD/YY
      2025-4-22  -> YYYY-M-DD
      May 17     -> M DD
      17 May     -> DD M
    """
    sample = " ".join(str(sample_text or "").strip().split())
    if not sample:
        raise ValueError("日期示例不能为空。")

    month_names = (
        r"JAN(?:UARY)?|FEB(?:RUARY)?|MAR(?:CH)?|APR(?:IL)?|MAY|"
        r"JUN(?:E)?|JUL(?:Y)?|AUG(?:UST)?|SEP(?:T(?:EMBER)?)?|"
        r"OCT(?:OBER)?|NOV(?:EMBER)?|DEC(?:EMBER)?"
    )

    # English month first: May 17, May 17 2025, May-17-25
    m = re.fullmatch(
        rf"(?i)({month_names})([\s./-]+)(\d{{1,2}})(?:([\s./-]+)(\d{{2}}|\d{{4}}))?",
        sample,
    )
    if m:
        sep1 = m.group(2)
        year = m.group(5)
        fmt = f"M{sep1}DD"
        if year:
            fmt += f"{m.group(4)}{'YYYY' if len(year) == 4 else 'YY'}"
        return fmt

    # English month second: 17 May, 17 May 2025, 17-May-25
    m = re.fullmatch(
        rf"(?i)(\d{{1,2}})([\s./-]+)({month_names})(?:([\s./-]+)(\d{{2}}|\d{{4}}))?",
        sample,
    )
    if m:
        sep1 = m.group(2)
        year = m.group(5)
        fmt = f"DD{sep1}M"
        if year:
            fmt += f"{m.group(4)}{'YYYY' if len(year) == 4 else 'YY'}"
        return fmt

    # Numeric formats. A four-digit first field is treated as the year.
    m = re.fullmatch(r"(\d{1,4})([./-])(\d{1,2})(?:([./-])(\d{2}|\d{4}))?", sample)
    if m:
        first, sep1, second, sep2, third = m.groups()
        if len(first) == 4:
            if third is None:
                raise ValueError("年份在前的日期需要包含年、月、日，例如 2025-4-22。")
            return f"YYYY{sep1}M{sep2}DD"

        fmt = f"M{sep1}DD"
        if third:
            fmt += f"{sep2}{'YYYY' if len(third) == 4 else 'YY'}"
        return fmt

    raise ValueError(
        "无法从日期示例中判断格式。请输入类似 4-22、04/22/2025、"
        "2025-4-22、May 17 或 17 May。"
    )


def resolve_date_input_to_format(date_input: str) -> str:
    """Accept either format tokens or real date examples and return formats.

    Multiple entries may be separated by commas, semicolons, or vertical bars.
    """
    raw = str(date_input or "").strip()
    if not raw:
        return ""

    parts = [p.strip() for p in re.split(r"[,，;；|]+", raw) if p.strip()]
    resolved = []
    format_token_re = re.compile(r"(?i)^(?:YYYY|YY|MM|M|DD|D|MMM|MON|MONTH|[\s./-])+$")

    for part in parts:
        if format_token_re.fullmatch(part):
            resolved.append(part.upper())
        else:
            resolved.append(infer_date_format_from_sample(part))

    return ", ".join(resolved)


def configure_date_format(custom_date_format: str):
    """Apply the custom date format, or restore defaults when blank."""
    global DATE_ANY_IN_TEXT_PATTERN
    global DATE_IN_TEXT_RE
    global DATE_FULL_RE
    global DATE_MERCHANT_LINE_RE

    custom_date_format = resolve_date_input_to_format(custom_date_format)
    DATE_ANY_IN_TEXT_PATTERN = (
        convert_custom_date_format_to_regex(custom_date_format)
        if custom_date_format
        else DEFAULT_DATE_ANY_IN_TEXT_PATTERN
    )

    DATE_IN_TEXT_RE = re.compile(DATE_ANY_IN_TEXT_PATTERN, re.IGNORECASE)
    DATE_FULL_RE = re.compile(rf'^\s*(?:{DATE_ANY_IN_TEXT_PATTERN})\s*$', re.IGNORECASE)
    DATE_MERCHANT_LINE_RE = re.compile(
        rf'^\s*({DATE_ANY_IN_TEXT_PATTERN})\s+(.*\S)\s*$', re.IGNORECASE
    )


EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', re.IGNORECASE)
URLISH_RE = re.compile(r'(https?://|www\.|\.com\b|squareup\.com\b)', re.IGNORECASE)

# ================= 常量 =================

MONTH_HEADERS = ["Jan", "Feb", "March", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

UI_TO_SHEET_MONTH = {
    "JAN": "Jan",
    "FEB": "Feb",
    "MAR": "March",
    "APR": "Apr",
    "MAY": "May",
    "JUN": "Jun",
    "JUL": "Jul",
    "AUG": "Aug",
    "SEP": "Sep",
    "OCT": "Oct",
    "NOV": "Nov",
    "DEC": "Dec",
}

CREDIT_DEFAULT_HEADER = "CASH / CHECK DEPOSIT"
CREDIT_TOTAL_HEADER = "TOTAL CREDIT"
CREDIT_DEBIT_HEADER = "DEBIT"
CREDIT_ENDING_HEADER = "ENDING BALANCE"
CREDIT_BEGIN_HEADER = "BEGIN BALANCE"
DEFAULT_BANK_NAME = ""

# ================= Excel position constants / runtime cache =================
# Credit type rule requires Month labels in Column A, therefore the fixed
# Beginning Balance column is the column immediately to the right: Column B.
CREDIT_MONTH_COL = 1
CREDIT_BEGIN_COL = 2

# Debit template rules are fixed and do NOT depend on header keywords.
# Column A is always Merchant. Month/date headers begin from Column B.
DEBIT_MERCHANT_COL = 1
DEBIT_FIRST_MONTH_COL = 2
DEBIT_MIN_PERIOD_HEADERS = 2
DEBIT_SCAN_MAX_ROWS = 50
DEBIT_SCAN_MAX_COLS = 300

# Debit columns do not move when transactions are added.  Detect each Debit
# sheet once per workbook path during the current program session, then reuse
# the numeric coordinates directly instead of rescanning the worksheet.
DEBIT_LAYOUT_CACHE: Dict[Tuple[str, str], Tuple[int, int, Tuple[int, ...], int, Optional[int]]] = {}

# Credit/Debit sheet-type classification is also cached per workbook path +
# sheet name, so switching sheets in the UI does not repeatedly rescan Column A
# of every worksheet in the workbook.
SHEET_TYPE_CACHE: Dict[Tuple[str, str], str] = {}


def _debit_cache_key(path: Path, sheet_name: str) -> Tuple[str, str]:
    return (str(Path(path).expanduser().resolve()).casefold(), str(sheet_name))


def clear_debit_layout_cache(path: Optional[Path] = None):
    if path is None:
        DEBIT_LAYOUT_CACHE.clear()
        SHEET_TYPE_CACHE.clear()
        return
    target = str(Path(path).expanduser().resolve()).casefold()
    for key in list(DEBIT_LAYOUT_CACHE):
        if key[0] == target:
            DEBIT_LAYOUT_CACHE.pop(key, None)
    for key in list(SHEET_TYPE_CACHE):
        if key[0] == target:
            SHEET_TYPE_CACHE.pop(key, None)


def cache_debit_layout(path: Path, sheet_name: str, layout):
    if layout is None:
        return None
    header_row, merchant_col, month_cols, total_col, category_col = layout
    frozen = (header_row, merchant_col, tuple(month_cols), total_col, category_col)
    DEBIT_LAYOUT_CACHE[_debit_cache_key(path, sheet_name)] = frozen
    return frozen


def get_cached_debit_layout(path: Path, sheet_name: str):
    cached = DEBIT_LAYOUT_CACHE.get(_debit_cache_key(path, sheet_name))
    if cached is None:
        return None
    header_row, merchant_col, month_cols, total_col, category_col = cached
    return header_row, merchant_col, list(month_cols), total_col, category_col


def cache_sheet_type(path: Path, sheet_name: str, sheet_type: str) -> str:
    SHEET_TYPE_CACHE[_debit_cache_key(path, sheet_name)] = sheet_type
    return sheet_type


def get_cached_sheet_type(path: Path, sheet_name: str) -> Optional[str]:
    return SHEET_TYPE_CACHE.get(_debit_cache_key(path, sheet_name))


CREDIT_MONTH_ROWS = {
    "JAN": 3,
    "FEB": 4,
    "MAR": 5,
    "APR": 6,
    "MAY": 7,
    "JUN": 8,
    "JUL": 9,
    "AUG": 10,
    "SEP": 11,
    "OCT": 12,
    "NOV": 13,
    "DEC": 14,
}

CREDIT_MONTHS = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]

DEBIT_SHEET_NAME = "debit summary"
DEBIT_MONTH_HEADERS = ["Jan", "Feb", "March", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

DEFAULT_UI_MONTH_LABELS = list(DEBIT_MONTH_HEADERS)

# Credit structure is fixed at 12 months + 1 TOTAL row (~15 rows). Limiting the
# Column-A scan range to this size prevents stray date-like text left over
# elsewhere in a large/old worksheet (far below the real data) from being
# miscounted as Credit period rows and misclassifying a Debit sheet as Credit.
CREDIT_COLUMN_A_SCAN_MAX_ROWS = 60


def normalize_month_labels(labels) -> List[str]:
    """Return exactly 12 non-empty display labels, preserving Excel order."""
    cleaned = []
    for value in list(labels or [])[:12]:
        text = str(value).strip() if value is not None else ""
        cleaned.append(text)
    while len(cleaned) < 12:
        cleaned.append(DEFAULT_UI_MONTH_LABELS[len(cleaned)])
    return [label or DEFAULT_UI_MONTH_LABELS[i] for i, label in enumerate(cleaned)]


def read_month_labels_from_wb(wb) -> List[str]:
    """Read the 12 display labels already stored in the Monthly workbook."""
    if DEBIT_SHEET_NAME in wb.sheetnames:
        ws = wb[DEBIT_SHEET_NAME]
        labels = [ws.cell(row=1, column=col).value for col in range(2, 14)]
        if any(v is not None and str(v).strip() for v in labels):
            return normalize_month_labels(labels)
    if "Credit Summary" in wb.sheetnames:
        ws = wb["Credit Summary"]
        labels = [ws.cell(row=row, column=1).value for row in range(3, 15)]
        if any(v is not None and str(v).strip() for v in labels):
            return normalize_month_labels(labels)
    return list(DEFAULT_UI_MONTH_LABELS)


def read_month_labels_from_path(path: Path) -> List[str]:
    if not path.exists():
        return list(DEFAULT_UI_MONTH_LABELS)
    try:
        wb = load_workbook(path, data_only=False, read_only=True)
        labels = read_month_labels_from_wb(wb)
        wb.close()
        return labels
    except Exception:
        return list(DEFAULT_UI_MONTH_LABELS)


# ================= 基础工具 =================

def is_date_short(s: str) -> bool:
    return bool(DATE_MMDDSHORT.match(s))


def is_date_yy(s: str) -> bool:
    return bool(DATE_MMDDYY.match(s))


def is_date_yyyy(s: str) -> bool:
    return bool(DATE_MMDDYYYY.match(s))


def is_date_any(s: str) -> bool:
    return bool(DATE_FULL_RE.match(str(s)))


def is_amount(s: str) -> bool:
    return bool(AMOUNT_RE.match(s))


def is_phone(s: str) -> bool:
    return bool(PHONE_RE.match(s))


def is_state(s: str) -> bool:
    return bool(STATE_RE.match(s))


def is_dot(s: str) -> bool:
    return bool(DOT_RE.match(s))


def is_email(s: str) -> bool:
    return bool(EMAIL_RE.match(s.strip()))


def looks_like_url_or_site(s: str) -> bool:
    return bool(URLISH_RE.search(s.strip()))


def clean_amount(s: str) -> str:
    s = s.strip()
    neg = False

    if s.startswith("(") and s.endswith(")"):
        neg = True
    if "-" in s:
        neg = True

    s = (
        s.replace("$", "")
            .replace(",", "")
            .replace("(", "")
            .replace(")", "")
            .replace("-", "")
            .strip()
    )

    if neg and s:
        return f"-{s}"
    return s


def clean_merchant(line: str) -> str:
    raw = ' '.join(line.split())
    for t in MERCHANT_CUT_TOKENS:
        idx = raw.find(t)
        if idx != -1:
            return raw[:idx].rstrip()
    return raw


def extract_date_and_merchant(line: str):
    m = DATE_MERCHANT_LINE_RE.match(line.strip())
    if not m:
        return None, None
    return m.group(1), m.group(2).strip()


def normalize_amount_string(amount) -> str:
    value = float(str(amount).replace(",", "").strip())
    return f"{value:.2f}"


def safe_float(x):
    try:
        if x is None:
            return 0.0
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip()
        if not s:
            return 0.0
        if s.startswith("="):
            nums = re.findall(r'[-+]?\d+(?:\.\d+)?', s)
            total = 0.0
            for n in nums:
                total += float(n)
            return total
        return float(s.replace(",", ""))
    except Exception:
        return 0.0


def build_plus_formula(parts: List[str]) -> str:
    clean_parts = []
    for p in parts:
        p = str(p).strip()
        if not p:
            continue
        if p.startswith("+"):
            p = p[1:]
        clean_parts.append(p)

    if not clean_parts:
        return ""

    expr = clean_parts[0]
    for p in clean_parts[1:]:
        if p.startswith("-"):
            expr += p
        else:
            expr += f"+{p}"
    return f"={expr}"


def split_formula_parts(formula_or_value) -> List[str]:
    if formula_or_value is None:
        return []

    if isinstance(formula_or_value, (int, float)):
        return [f"{float(formula_or_value):.2f}"]

    s = str(formula_or_value).strip()
    if not s:
        return []

    if s.startswith("="):
        s = s[1:].strip()

    if not s:
        return []

    nums = re.findall(r'[-+]?\d+(?:\.\d+)?', s)
    out = []
    for n in nums:
        if n.startswith("+"):
            n = n[1:]
        out.append(f"{float(n):.2f}")
    return out


def script_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


# ============================================================================
# 原始数据预处理引擎（内置）
# ============================================================================
#
# 负责把顺序混乱的原始账单整理成本程序能直接解析的格式，每次 Start 都会
# 先跑一遍。它能处理的杂乱形态包括：
#
#   商户在前 / 日期在后      AMAZON.COM 04/22 125.00
#   金额穿插在中间           125.00 AMAZON.COM 04/22
#   一笔拆成多行             AMAZON.COM \n 04/22 \n 125.00
#   多笔挤成一整行           04/22 A 9.99 04/23 B 19.99 04/24 C 29.99
#   行尾还带余额列           04/22 AMAZON 125.00 4,875.00
#   夹杂表头/小计/页码行
#
# 核心思路是不假设日期、商户、金额的先后顺序，而是按特征强弱逐层剥离：
# 先找出金额并遮蔽（金额特征最强：两位小数、千分位、$ 符号），再在剩余
# 文本里找日期，最后剩下的文字就是商户。先遮蔽金额这一步很关键，否则
# "5.12" 这类金额会被误判成 5 月 12 日。
#
# 整理结果统一输出成零填充的 MM/DD，正好是本程序默认日期规则识别的格式。


# ==================== 月份名称 ====================

MONTH_WORD_PATTERN = (
    r"(?:JAN(?:UARY)?|FEB(?:RUARY)?|MAR(?:CH)?|APR(?:IL)?|MAY|"
    r"JUN(?:E)?|JUL(?:Y)?|AUG(?:UST)?|SEP(?:T(?:EMBER)?)?|"
    r"OCT(?:OBER)?|NOV(?:EMBER)?|DEC(?:EMBER)?)"
)

MONTH_NAME_TO_NUMBER = {
    "JAN": 1, "JANUARY": 1,
    "FEB": 2, "FEBRUARY": 2,
    "MAR": 3, "MARCH": 3,
    "APR": 4, "APRIL": 4,
    "MAY": 5,
    "JUN": 6, "JUNE": 6,
    "JUL": 7, "JULY": 7,
    "AUG": 8, "AUGUST": 8,
    "SEP": 9, "SEPT": 9, "SEPTEMBER": 9,
    "OCT": 10, "OCTOBER": 10,
    "NOV": 11, "NOVEMBER": 11,
    "DEC": 12, "DECEMBER": 12,
}


def month_name_to_number(name: str) -> Optional[int]:
    return MONTH_NAME_TO_NUMBER.get(str(name or "").strip().upper().rstrip("."))


# ==================== 金额识别 ====================

# 金额分两种写法分别匹配，然后合并去重：
#
#   A 型：带 $ 符号。有 $ 就足以确认是金额，因此小数部分可以省略（$125）。
#   B 型：不带 $ 符号。必须有两位小数（125.00）或千分位逗号（5,000），
#         否则一个孤立的 "125" 无法与参考号/单据号区分，宁可不认。
#
# 两种都支持负号在前（-125.00）、负号在后（125.00-）、括号表示负数
# （(125.00)）以及 CR / DR 后缀，这些都是真实账单里常见的写法。

_AMOUNT_WITH_DOLLAR_RE = re.compile(
    r"(?P<open>\()?\s*"
    r"(?P<sign1>[-+])?\s*"
    r"\$\s*"
    r"(?P<num>\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?|\d+(?:\.\d{1,2})?)"
    r"\s*(?P<sign2>-)?"
    r"\s*(?P<close>\))?"
    r"(?:\s*(?P<crdr>CR|DR)\b)?",
    re.IGNORECASE,
)

_AMOUNT_PLAIN_RE = re.compile(
    r"(?<![\w.,])"
    r"(?P<open>\()?\s*"
    r"(?P<sign1>[-+])?\s*"
    r"(?P<num>\d{1,3}(?:,\d{3})+(?:\.\d{1,2})?|\d+\.\d{2})"
    r"\s*(?P<sign2>-)?"
    r"\s*(?P<close>\))?"
    r"(?:\s*(?P<crdr>CR|DR)\b)?"
    r"(?![\d.])",
    re.IGNORECASE,
)


@dataclass
class AmountHit:
    start: int
    end: int
    raw: str
    value: float


def _amount_value_from_match(m) -> Optional[float]:
    num_text = (m.group("num") or "").replace(",", "").strip()
    if not num_text:
        return None
    try:
        value = float(num_text)
    except ValueError:
        return None

    negative = False
    if m.group("open") and m.group("close"):
        negative = True
    if m.group("sign1") == "-":
        negative = True
    if m.group("sign2") == "-":
        negative = True
    crdr = (m.group("crdr") or "").upper()
    if crdr == "DR":
        # DR = Debit，账单上通常表示扣款。CR = Credit 保持正数。
        negative = True

    return -value if negative else value


def find_amounts(text: str) -> List[AmountHit]:
    """找出一段文字里所有的金额，按出现位置排序，互不重叠。"""
    hits: List[AmountHit] = []
    taken: List[Tuple[int, int]] = []

    def overlaps(start: int, end: int) -> bool:
        return any(not (end <= s or start >= e) for s, e in taken)

    # 带 $ 的优先，避免 "$5,000" 被 B 型规则拆成别的结果。
    for regex in (_AMOUNT_WITH_DOLLAR_RE, _AMOUNT_PLAIN_RE):
        for m in regex.finditer(text):
            value = _amount_value_from_match(m)
            if value is None:
                continue
            start, end = m.start(), m.end()
            if overlaps(start, end):
                continue
            taken.append((start, end))
            hits.append(AmountHit(start, end, m.group(0).strip(), value))

    hits.sort(key=lambda h: h.start)
    return hits


# ==================== 日期识别 ====================

@dataclass
class DateHit:
    start: int
    end: int
    raw: str
    month: int
    day: int
    year: Optional[int] = None


def _normalize_two_digit_year(year_text: str) -> Optional[int]:
    text = str(year_text or "").strip()
    if not text.isdigit():
        return None
    if len(text) == 4:
        return int(text)
    if len(text) == 2:
        # 两位年份统一按 20xx 解释；账单基本不会出现 19xx。
        return 2000 + int(text)
    return None


def _valid_month_day(month: int, day: int) -> bool:
    return 1 <= month <= 12 and 1 <= day <= 31


def _resolve_numeric_month_day(
        first: int, second: int, day_first_mode: str
) -> Optional[Tuple[int, int]]:
    """把两个数字解释成 (月, 日)。

    day_first_mode:
      "auto"       先按月/日理解；月份不可能大于 12，因此当第一个数字 > 12
                   而第二个 <= 12 时自动改判为 日/月。
      "month_first" 永远按 月/日 理解（美国账单常见）。
      "day_first"   永远按 日/月 理解（欧洲/部分国际账单）。
    """
    if day_first_mode == "day_first":
        month, day = second, first
        return (month, day) if _valid_month_day(month, day) else None

    if day_first_mode == "month_first":
        month, day = first, second
        return (month, day) if _valid_month_day(month, day) else None

    # auto
    if _valid_month_day(first, second):
        return first, second
    if _valid_month_day(second, first):
        return second, first
    return None


def build_date_patterns(allow_dot_separator: bool,
                        allow_compact_mmdd: bool = False) -> List[Tuple[str, re.Pattern]]:
    """按优先级返回日期正则。

    三段式（含年份）必须排在两段式前面，否则 "04/22/25" 会被先切成 "04/22"。

    是否允许用点号作为日期分隔符由调用方决定：账单里 "4.22" 既可能是日期
    也可能是金额，默认关闭以免误判，需要时可在界面上打开。

    allow_compact_mmdd 控制是否识别没有分隔符的四位日期，例如美国银行
    "CHECKCARD 0422 AMAZON.COM" 这种写法。这个规则默认关闭：任何一个四位
    数字（单据号、门店号）都可能碰巧长得像 MMDD，开启后误判风险明显更高，
    因此交给用户按自己账单的实际格式决定。开启时也要求严格零填充
    （0422 可以，422 不行），把误判面压到最小。
    """
    sep = r"[-/.]" if allow_dot_separator else r"[-/]"

    patterns = [
        # 2025-04-22 / 2025/4/22
        ("iso", re.compile(rf"(?<!\d)(\d{{4}}){sep}(\d{{1,2}}){sep}(\d{{1,2}})(?!\d)")),
        # 04/22/25 / 4-22-2025
        ("mdy", re.compile(rf"(?<!\d)(\d{{1,2}}){sep}(\d{{1,2}}){sep}(\d{{2,4}})(?!\d)")),
        # Apr 22 2025 / April 22, 2025 / Apr-22-25
        ("word_day", re.compile(
            rf"\b({MONTH_WORD_PATTERN})\.?[\s,\-/]+(\d{{1,2}})"
            rf"(?:[\s,\-/]+(\d{{2,4}})(?!\d))?",
            re.IGNORECASE)),
        # 22 Apr 2025 / 22-Apr-25
        ("day_word", re.compile(
            rf"(?<!\d)(\d{{1,2}})[\s,\-/]*({MONTH_WORD_PATTERN})\.?"
            rf"(?:[\s,\-/]+(\d{{2,4}})(?!\d))?",
            re.IGNORECASE)),
        # 04/22 / 4-22
        ("md", re.compile(rf"(?<!\d)(\d{{1,2}}){sep}(\d{{1,2}})(?!\d)")),
    ]

    if allow_compact_mmdd:
        # 0422 -> 4 月 22 日。必须严格零填充的四位，且左右不能再有数字。
        patterns.append(
            ("mmdd", re.compile(r"(?<!\d)(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])(?!\d)"))
        )

    return patterns


def find_dates(text: str, day_first_mode: str = "auto",
               allow_dot_separator: bool = False,
               allow_compact_mmdd: bool = False) -> List[DateHit]:
    """找出一段文字里所有的日期，按位置排序，互不重叠。

    调用前建议先用 mask_spans() 把金额位置遮蔽掉，否则 "5.12" 这类
    金额可能被当成 5 月 12 日。
    """
    hits: List[DateHit] = []
    taken: List[Tuple[int, int]] = []

    def overlaps(start: int, end: int) -> bool:
        return any(not (end <= s or start >= e) for s, e in taken)

    for kind, regex in build_date_patterns(allow_dot_separator, allow_compact_mmdd):
        for m in regex.finditer(text):
            start, end = m.start(), m.end()
            if overlaps(start, end):
                continue

            month = day = None
            year = None

            if kind == "iso":
                year = int(m.group(1))
                month, day = int(m.group(2)), int(m.group(3))
                if not _valid_month_day(month, day):
                    continue

            elif kind == "mdy":
                resolved = _resolve_numeric_month_day(
                    int(m.group(1)), int(m.group(2)), day_first_mode
                )
                if resolved is None:
                    continue
                month, day = resolved
                year = _normalize_two_digit_year(m.group(3))

            elif kind == "word_day":
                month = month_name_to_number(m.group(1))
                day = int(m.group(2))
                if month is None or not _valid_month_day(month, day):
                    continue
                if m.group(3):
                    year = _normalize_two_digit_year(m.group(3))

            elif kind == "day_word":
                day = int(m.group(1))
                month = month_name_to_number(m.group(2))
                if month is None or not _valid_month_day(month, day):
                    continue
                if m.group(3):
                    year = _normalize_two_digit_year(m.group(3))

            elif kind == "md":
                resolved = _resolve_numeric_month_day(
                    int(m.group(1)), int(m.group(2)), day_first_mode
                )
                if resolved is None:
                    continue
                month, day = resolved

            elif kind == "mmdd":
                # 无分隔符写法只按 MMDD 解释，不做月/日互换猜测。
                month, day = int(m.group(1)), int(m.group(2))
                if not _valid_month_day(month, day):
                    continue

            if month is None or day is None:
                continue

            taken.append((start, end))
            hits.append(DateHit(start, end, m.group(0).strip(), month, day, year))

    hits.sort(key=lambda h: h.start)
    return hits


def mask_spans(text: str, spans: List[Tuple[int, int]]) -> str:
    """把指定区间替换成等长空格，保持其余字符的下标不变。"""
    if not spans:
        return text
    chars = list(text)
    for start, end in spans:
        for i in range(max(0, start), min(len(chars), end)):
            chars[i] = " "
    return "".join(chars)


# ==================== 商户名清洗 ====================

# 常见噪音：参考号、卡号尾号、授权码、网址、电话等。
NOISE_PATTERNS = [
    re.compile(r"\bREF\s*#?\s*[:#]?\s*[\w-]+", re.IGNORECASE),
    re.compile(r"\bTRACE\s*#?\s*[:#]?\s*[\w-]+", re.IGNORECASE),
    re.compile(r"\bCONF(?:IRMATION)?\s*#?\s*[:#]?\s*[\w-]+", re.IGNORECASE),
    re.compile(r"\bAUTH\s*#?\s*[:#]?\s*[\w-]+", re.IGNORECASE),
    re.compile(r"\bCARD\s*#?\s*\d{2,}", re.IGNORECASE),
    re.compile(r"\bX{2,}\d{2,}\b", re.IGNORECASE),
    re.compile(r"\*{2,}\d{2,}\b"),
    re.compile(r"https?://\S+", re.IGNORECASE),
    re.compile(r"\bwww\.\S+", re.IGNORECASE),
    re.compile(r"\b\d{3}[-.\s]\d{3}[-.\s]\d{4}\b"),          # 电话
    re.compile(r"\b[\w.+-]+@[\w-]+\.[\w.]+\b"),               # 邮箱
]


def clean_merchant_text(text: str, remove_noise: bool = False,
                        remove_items: Optional[List[str]] = None) -> str:
    """整理商户名：去掉标记后缀、可选噪音、用户指定关键词，并规范空白。"""
    result = " ".join(str(text or "").split())
    if not result:
        return ""

    # 用户自定义要删除的关键词（与主程序"预处理删除内容"用途一致）。
    for item in (remove_items or []):
        item = str(item).strip()
        if item:
            result = re.compile(re.escape(item), re.IGNORECASE).sub(" ", result)

    if remove_noise:
        for pattern in NOISE_PATTERNS:
            result = pattern.sub(" ", result)

    result = " ".join(result.split())

    # 与主程序一致：遇到 DES: / ID: / INDN: / CO ID: 就截断。
    upper = result.upper()
    for token in MERCHANT_CUT_TOKENS:
        idx = upper.find(token)
        if idx != -1:
            result = result[:idx]
            upper = result.upper()

    # 去掉首尾多余的标点和分隔符号。
    result = result.strip(" \t-–—|,;:*#.")
    return " ".join(result.split())


# ==================== 需要跳过的非交易行 ====================

DEFAULT_SKIP_KEYWORDS = [
    "beginning balance",
    "ending balance",
    "previous balance",
    "new balance",
    "statement balance",
    "balance forward",
    "total deposits",
    "total withdrawals",
    "total credits",
    "total debits",
    "daily balance",
    "minimum payment",
    "payment due",
    "statement period",
    "account summary",
    "page ",
    "continued",
    "date description",
    "transaction detail",
    "posting date",
]


def looks_like_skippable_line(line: str, skip_keywords: List[str],
                              day_first_mode: str = "auto",
                              allow_dot_separator: bool = False,
                              allow_compact_mmdd: bool = False) -> bool:
    """判断一行是不是表头 / 小计 / 页码之类的非交易行。

    这一步很重要：像 "Beginning Balance   5,000.00" 这种行含有金额但没有
    交易含义，如果不剔除，它的金额可能被错误地和后面某一行的日期拼成一条
    假交易。
    """
    text = " ".join(str(line or "").split()).lower()
    if not text:
        return True

    for keyword in skip_keywords:
        keyword = str(keyword).strip().lower()
        if keyword and keyword in text:
            return True

    # 纯页码行，例如 "1 / 5"。
    #
    # 注意：这条规则必须先排除掉真正的日期行和金额行。原始数据被逐行拆开时
    # （PDF 复制常见），日期会单独占一行，例如只有 "04/22" 五个字符 —— 它同样
    # 由数字和斜杠组成、长度也很短，如果不做保护就会被当成页码整行丢掉，
    # 那条交易就再也拼不回来了。
    if re.fullmatch(r"[\d\s/of\-]+", text) and len(text) <= 12:
        raw = str(line or "")
        amount_hits = find_amounts(raw)
        masked = mask_spans(raw, [(h.start, h.end) for h in amount_hits])
        date_hits = find_dates(
            masked,
            day_first_mode=day_first_mode,
            allow_dot_separator=allow_dot_separator,
            allow_compact_mmdd=allow_compact_mmdd,
        )
        if not date_hits and not amount_hits:
            return True

    return False


# ==================== 数据结构 ====================

@dataclass
class Item:
    """一行文本被拆解出来的最小单元：日期 / 金额 / 文字。

    line_no 是【分组键】，不一定等于原始行号：当一行里连着挤了好几笔交易时
    （见 split_compact_line_items），这一行会被切成多段，每段拿到各自独立的
    分组键，这样后续组装逻辑就能像处理普通多行数据一样处理它们。
    origin_line 始终保存真实的原始行号，只用于错误提示。
    """
    kind: str          # 'date' | 'amount' | 'text'
    line_no: int       # 分组键（同一段落内的单元共享）
    raw: str
    date: Optional[DateHit] = None
    amount: Optional[float] = None
    origin_line: int = 0


@dataclass
class Record:
    date: Optional[DateHit] = None
    amount: Optional[float] = None
    merchant_parts: List[str] = field(default_factory=list)
    line_nos: List[int] = field(default_factory=list)     # 原始行号，用于提示
    group_keys: List[int] = field(default_factory=list)   # 分组键，用于同段判断
    last_group_key: Optional[int] = None

    @property
    def merchant_raw(self) -> str:
        return " ".join(p for p in self.merchant_parts if p.strip())

    def is_empty(self) -> bool:
        return (self.date is None and self.amount is None
                and not self.merchant_raw.strip())

    def is_complete(self) -> bool:
        return (self.date is not None and self.amount is not None
                and bool(self.merchant_raw.strip()))


@dataclass
class Transaction:
    date: DateHit
    merchant: str
    amount: float
    source_lines: List[int]


@dataclass
class Issue:
    line_nos: List[int]
    reason: str
    content: str


@dataclass
class ProcessOptions:
    # 日期解释
    day_first_mode: str = "auto"          # auto | month_first | day_first
    allow_dot_separator: bool = False     # 是否把 4.22 也当作日期
    allow_compact_mmdd: bool = False      # 是否把无分隔符的 0422 也当作日期
    default_year: str = ""                # 原始数据没有年份时补充的年份

    # 日期输出格式（默认 MM/DD，主程序默认设置即可识别）
    date_output_format: str = "MM/DD"     # KEEP | MM/DD | MM/DD/YY | MM/DD/YYYY

    # 金额处理
    balance_column_mode: str = "auto"     # auto | none | last_is_balance
    amount_sign_mode: str = "keep"        # keep | abs | negative

    # 商户处理
    remove_noise: bool = False
    remove_items: List[str] = field(default_factory=list)

    # 行过滤
    skip_keywords: List[str] = field(default_factory=lambda: list(DEFAULT_SKIP_KEYWORDS))

    # 输出格式
    output_style: str = "two_line"        # two_line | one_line

    # 一条记录最多允许跨几行
    max_record_span: int = 4


@dataclass
class ProcessResult:
    transactions: List[Transaction] = field(default_factory=list)
    issues: List[Issue] = field(default_factory=list)
    skipped_lines: List[Tuple[int, str]] = field(default_factory=list)
    balance_column_detected: bool = False
    hints: List[str] = field(default_factory=list)
    output_text: str = ""
    total_amount: float = 0.0


# ==================== 行拆解 ====================

def tokenize_line(line: str, line_no: int, options: ProcessOptions) -> List[Item]:
    """把一行文本按出现顺序拆成 日期 / 金额 / 文字 单元。

    顺序很关键：先识别金额并遮蔽，再识别日期，最后剩下的才是文字。
    正因为最终是按【出现位置】排序输出，所以商户在前、日期在后这类
    顺序颠倒的数据也能被后续组装逻辑正确处理。
    """
    text = line.replace("\t", "  ")
    if not text.strip():
        return []

    amount_hits = find_amounts(text)
    masked = mask_spans(text, [(h.start, h.end) for h in amount_hits])
    date_hits = find_dates(
        masked,
        day_first_mode=options.day_first_mode,
        allow_dot_separator=options.allow_dot_separator,
        allow_compact_mmdd=options.allow_compact_mmdd,
    )

    spans: List[Tuple[int, int, str, object]] = []
    for h in amount_hits:
        spans.append((h.start, h.end, "amount", h))
    for d in date_hits:
        spans.append((d.start, d.end, "date", d))
    spans.sort(key=lambda s: s[0])

    items: List[Item] = []
    cursor = 0
    for start, end, kind, payload in spans:
        gap = text[cursor:start]
        if gap.strip():
            items.append(Item("text", line_no, gap.strip(), origin_line=line_no))
        if kind == "amount":
            items.append(Item("amount", line_no, payload.raw,
                              amount=payload.value, origin_line=line_no))
        else:
            items.append(Item("date", line_no, payload.raw,
                              date=payload, origin_line=line_no))
        cursor = end

    tail = text[cursor:]
    if tail.strip():
        items.append(Item("text", line_no, tail.strip(), origin_line=line_no))

    return items


# ==================== 余额列处理 ====================

def _balance_pairs_match(pairs: List[Tuple[float, float]]) -> bool:
    """判断一组 (交易金额, 疑似余额) 是否符合余额列的数学特征。

    余额列的可靠特征是：本条余额 = 上一条余额 ± 本条交易金额。只要多数
    条目满足这个递推关系，就可以确定第二个数字是余额而不是交易金额。
    这比"看表头有没有 Balance 字样"可靠得多，因为粘贴出来的数据经常
    根本没有表头。
    """
    if len(pairs) < 3:
        return False

    matches = 0
    checks = 0
    for i in range(1, len(pairs)):
        prev_balance = pairs[i - 1][1]
        txn_amount, balance = pairs[i]
        checks += 1
        if (abs(balance - (prev_balance + txn_amount)) < 0.01
                or abs(balance - (prev_balance - txn_amount)) < 0.01):
            matches += 1

    return checks > 0 and (matches / checks) >= 0.6


def detect_balance_column(line_items: List[List[Item]]) -> bool:
    """自动判断【跨行】数据的行尾是否为余额列。"""
    pairs = []
    for items in line_items:
        amounts = [i.amount for i in items if i.kind == "amount"]
        if len(amounts) >= 2:
            pairs.append((amounts[-2], amounts[-1]))
    return _balance_pairs_match(pairs)


def demote_extra_dates_to_text(items: List[Item]) -> List[Item]:
    """把一行里"多余的日期"降级成普通文字。

    一行里的交易笔数由金额个数决定。如果日期个数比金额个数还多，多出来的
    日期几乎都是描述文本的一部分，而不是另一笔交易的日期，例如：

        04/22 PAYMENT FOR INVOICE DATED 03/15 125.00

    这里只有一笔交易（一个金额），03/15 属于摘要内容。如果不做这一步，
    第二个日期会触发"日期槽已占用"的收尾规则，把这一行硬拆成两条残缺
    记录，两条都进不了结果。

    处理方式是保留最靠前的那个日期作为交易日期，其余日期原样转成文字
    留在商户描述里，信息不会丢失。

    这个规则刻意只在【这一行恰好只有一个金额】时才生效，也就是确定这行
    只有一笔交易的时候。一旦出现多个金额，就说明这是多笔交易挤在一行
    （见 split_compact_line_items），此时哪个日期属于哪一笔本身就是有
    歧义的，硬猜反而容易把正确的交易日期改错，不如交给切分逻辑处理，
    实在拿不准的会进入"待确认"由人工判断。

    跨行拆开的数据（每行只有日期或只有金额）也完全不受影响。
    """
    date_positions = [i for i, it in enumerate(items) if it.kind == "date"]
    amount_count = sum(1 for it in items if it.kind == "amount")

    if amount_count != 1 or len(date_positions) <= 1:
        return items

    keep = set(date_positions[:1])
    for index in date_positions:
        if index in keep:
            continue
        old_item = items[index]
        items[index] = Item(
            "text", old_item.line_no, old_item.raw,
            origin_line=old_item.origin_line,
        )
    return items


def split_compact_line_items(items: List[Item], balance_mode: str = "auto"
                             ) -> Tuple[List[List[Item]], bool]:
    """把"好几笔交易挤在同一行"的数据切成一笔一段。

    原始数据本该一行一笔，但复制粘贴时经常变成一整条长串：

        11/22/2025 AMAZON 9.99 11/23/2025 WALMART 19.99 11/24/2025 SHELL 29.99

    这里先判断这一行是不是这种"连行"数据，是的话就切开，让后面的组装
    逻辑可以像处理正常多行数据一样处理它们。

    判定条件刻意收得比较紧：必须同时出现至少 2 个日期，且金额个数等于
    日期个数（每笔一个金额）或等于日期个数的两倍（每笔还跟一个余额）。
    只有 1 个日期的普通行、或者日期和金额个数对不上的行，一律原样返回，
    避免把正常单笔交易误切开。

    返回 (切分后的段落列表, 是否识别出并剔除了行内余额)。
    """
    date_positions = [i for i, it in enumerate(items) if it.kind == "date"]
    amount_positions = [i for i, it in enumerate(items) if it.kind == "amount"]

    if len(date_positions) < 2:
        return [items], False
    if len(amount_positions) not in (len(date_positions), 2 * len(date_positions)):
        return [items], False

    balance_detected = False

    # 每笔交易配了两个金额时，第二个通常是余额。用同样的递推关系确认，
    # 确认成立才剔除；否则宁可保留，让它进入"待确认"提示人工判断。
    if len(amount_positions) == 2 * len(date_positions):
        pairs = [
            (items[amount_positions[i]].amount, items[amount_positions[i + 1]].amount)
            for i in range(0, len(amount_positions), 2)
        ]
        should_drop = (
            balance_mode == "last_is_balance"
            or (balance_mode == "auto" and _balance_pairs_match(pairs))
        )
        if should_drop:
            drop_positions = set(amount_positions[1::2])
            items = [it for i, it in enumerate(items) if i not in drop_positions]
            balance_detected = True
            date_positions = [i for i, it in enumerate(items) if it.kind == "date"]

    # 判断这一行的版式：日期在商户前面，还是商户在日期前面。
    # 两种版式的切分点不同，必须先认出来：
    #   日期在前 -> 每遇到一个新日期就是新的一笔
    #   商户在前 -> 当前这笔已经凑齐日期和金额后，再遇到文字就是下一笔的商户
    first_date = date_positions[0] if date_positions else None
    first_text = next((i for i, it in enumerate(items) if it.kind == "text"), None)
    date_leads = (first_text is None) or (first_date is not None and first_date < first_text)

    segments: List[List[Item]] = []
    current: List[Item] = []
    has_date = False
    has_amount = False

    for item in items:
        start_new = False
        if item.kind == "date" and has_date:
            start_new = True
        elif item.kind == "amount" and has_amount:
            start_new = True
        elif item.kind == "text" and (not date_leads) and has_date and has_amount:
            start_new = True

        if start_new and current:
            segments.append(current)
            current = []
            has_date = False
            has_amount = False

        current.append(item)
        if item.kind == "date":
            has_date = True
        elif item.kind == "amount":
            has_amount = True

    if current:
        segments.append(current)

    return segments, balance_detected


def apply_balance_column(line_items: List[List[Item]], mode: str) -> Tuple[List[List[Item]], bool]:
    """按设置去掉每行末尾的余额金额。"""
    detected = False
    if mode == "auto":
        detected = detect_balance_column(line_items)
        drop = detected
    elif mode == "last_is_balance":
        drop = True
    else:
        drop = False

    if not drop:
        return line_items, detected

    cleaned: List[List[Item]] = []
    for items in line_items:
        amount_indexes = [idx for idx, i in enumerate(items) if i.kind == "amount"]
        if len(amount_indexes) >= 2:
            last = amount_indexes[-1]
            cleaned.append([i for idx, i in enumerate(items) if idx != last])
        else:
            cleaned.append(items)
    return cleaned, (detected or mode == "last_is_balance")


# ==================== 记录组装 ====================

def assemble_records(all_items: List[Item], options: ProcessOptions
                     ) -> Tuple[List[Record], List[Record]]:
    """把散落的单元组装成一条条完整记录。

    组装规则（与日期/商户/金额的先后顺序无关）：

      * 日期单元 -> 填入当前记录的日期槽；若日期槽已被占用，说明上一条
        记录结束了，先收尾再开新记录。
      * 金额单元 -> 同理填入金额槽。
      * 文字单元 -> 追加到商户名。

    另外有两条防串行的保护规则：

      1. 一条记录一旦已经拿到金额，就只接受【同一行】的后续单元。
         新的一行必定属于新记录。这可以防止某行残留的孤立金额
         （例如小计行）跟下一行的日期粘成一条假交易。
      2. 一条记录最多跨 max_record_span 行，超过就强制收尾。
    """
    completed: List[Record] = []
    unresolved: List[Record] = []
    current = Record()

    def flush():
        nonlocal current
        if not current.is_empty():
            if current.is_complete():
                completed.append(current)
            else:
                unresolved.append(current)
        current = Record()

    for item in all_items:
        same_line = (current.last_group_key == item.line_no)

        # 保护规则 1：已经有金额的记录，不再接受新行（新段落）的内容。
        if (not current.is_empty()) and current.amount is not None and not same_line:
            flush()
            same_line = False

        # 保护规则 2：跨行数超限就收尾。
        if (current.group_keys
                and (item.line_no - min(current.group_keys)) >= options.max_record_span):
            flush()
            same_line = False

        if item.kind == "date":
            if current.date is not None:
                flush()
            current.date = item.date

        elif item.kind == "amount":
            if current.amount is not None:
                flush()
            current.amount = item.amount

        else:  # text
            if current.is_empty():
                current.merchant_parts.append(item.raw)
            elif not current.merchant_raw.strip():
                # 记录还没有商户名，这段文字就是它的描述（可能在下一行）。
                current.merchant_parts.append(item.raw)
            elif same_line:
                # 同一行内的补充说明，直接接在商户名后面。
                current.merchant_parts.append(item.raw)
            else:
                # 记录已经有商户名了，新一行的文字视为新记录的开头。
                flush()
                current.merchant_parts.append(item.raw)

        current.line_nos.append(item.origin_line or item.line_no)
        current.group_keys.append(item.line_no)
        current.last_group_key = item.line_no

    flush()
    return completed, unresolved


# ==================== 输出格式化 ====================

def format_date(date_hit: DateHit, options: ProcessOptions) -> str:
    """按设置输出日期。

    默认 MM/DD 且零填充，因为 BSDP 主程序的默认日期规则要求两位月份和
    两位日期（04/22 可以，4/22 不行）。这样整理完的数据不需要在主程序里
    额外设置日期格式就能直接使用。
    """
    style = (options.date_output_format or "MM/DD").upper()

    if style == "KEEP":
        return date_hit.raw

    year = date_hit.year
    if year is None:
        default_year = str(options.default_year or "").strip()
        if default_year.isdigit():
            year = _normalize_two_digit_year(default_year)

    if style == "MM/DD/YYYY" and year:
        return f"{date_hit.month:02d}/{date_hit.day:02d}/{year:04d}"
    if style == "MM/DD/YY" and year:
        return f"{date_hit.month:02d}/{date_hit.day:02d}/{year % 100:02d}"

    # 没有年份可用时统一退回 MM/DD，主程序同样支持。
    return f"{date_hit.month:02d}/{date_hit.day:02d}"


def format_amount(value: float, options: ProcessOptions) -> str:
    """输出金额。结果保证能被主程序的金额规则识别。"""
    mode = (options.amount_sign_mode or "keep").lower()
    if mode == "abs":
        value = abs(value)
    elif mode == "negative":
        value = -abs(value)
    return f"{value:.2f}"


def build_output_text(transactions: List[Transaction], options: ProcessOptions) -> str:
    """生成可直接粘贴回 BSDP 主程序的文本。

    two_line（默认，兼容性最好）：
        04/22 AMAZON.COM
        125.00

    one_line：
        04/22 AMAZON.COM 125.00
    """
    lines: List[str] = []
    for txn in transactions:
        date_text = format_date(txn.date, options)
        amount_text = format_amount(txn.amount, options)
        if options.output_style == "one_line":
            lines.append(f"{date_text} {txn.merchant} {amount_text}")
        else:
            lines.append(f"{date_text} {txn.merchant}")
            lines.append(amount_text)
    return "\n".join(lines)


# ==================== 主处理入口 ====================

def process_text(raw_text: str, options: Optional[ProcessOptions] = None) -> ProcessResult:
    """把杂乱的原始账单文本整理成标准格式。"""
    options = options or ProcessOptions()
    result = ProcessResult()

    text = str(raw_text or "").replace("\r\n", "\n").replace("\r", "\n")
    raw_lines = text.split("\n")

    line_items: List[List[Item]] = []
    group_key = 0
    inline_balance_detected = False
    for index, line in enumerate(raw_lines, start=1):
        if not line.strip():
            continue
        if looks_like_skippable_line(
                line, options.skip_keywords,
                day_first_mode=options.day_first_mode,
                allow_dot_separator=options.allow_dot_separator,
                allow_compact_mmdd=options.allow_compact_mmdd):
            result.skipped_lines.append((index, line.strip()))
            continue
        items = tokenize_line(line, index, options)
        if not items:
            continue

        # 先把"比金额还多出来的日期"还原成描述文字，避免摘要里的日期被
        # 误当成另一笔交易的开始。
        items = demote_extra_dates_to_text(items)

        # 一行里挤了多笔交易时，先切成一笔一段。切开后每段拿到独立的分组键，
        # 于是后续组装逻辑对待它们的方式和对待正常的一行一笔完全一致。
        segments, inline_balance = split_compact_line_items(
            items, options.balance_column_mode
        )
        if inline_balance:
            inline_balance_detected = True

        for segment in segments:
            group_key += 1
            for item in segment:
                item.line_no = group_key
            line_items.append(segment)

    line_items, balance_detected = apply_balance_column(
        line_items, options.balance_column_mode
    )
    result.balance_column_detected = balance_detected or inline_balance_detected

    all_items: List[Item] = [i for items in line_items for i in items]
    completed, unresolved = assemble_records(all_items, options)

    for record in completed:
        merchant = clean_merchant_text(
            record.merchant_raw,
            remove_noise=options.remove_noise,
            remove_items=options.remove_items,
        )
        if not merchant:
            # 清洗之后商户名空了（整行都是噪音），交给人工确认，
            # 不要生成一条没有商户名的记录 —— 主程序也无法识别。
            result.issues.append(Issue(
                sorted(set(record.line_nos)),
                "清理后商户名为空",
                record.merchant_raw or "(无文字内容)",
            ))
            continue
        result.transactions.append(Transaction(
            date=record.date,
            merchant=merchant,
            amount=record.amount,
            source_lines=sorted(set(record.line_nos)),
        ))

    for record in unresolved:
        missing = []
        if record.date is None:
            missing.append("日期")
        if record.amount is None:
            missing.append("金额")
        if not record.merchant_raw.strip():
            missing.append("商户")
        content_parts = []
        if record.date is not None:
            content_parts.append(record.date.raw)
        if record.merchant_raw.strip():
            content_parts.append(record.merchant_raw.strip())
        if record.amount is not None:
            content_parts.append(f"{record.amount:.2f}")
        result.issues.append(Issue(
            sorted(set(record.line_nos)),
            "缺少" + "、".join(missing),
            " | ".join(content_parts) or "(空)",
        ))

    # 智能提示：帮用户判断该调整哪个选项。
    #
    # 余额列的自动检测依赖"本行余额 = 上一行余额 ± 本行金额"这个连续关系，
    # 因此当账单里只有部分行带余额（中间夹着没有余额的行）时，连续性被打断，
    # 自动检测会保守地判定"没有余额列"。这时那些余额数字会变成一堆
    # 只有金额、没有日期也没有商户的待确认记录 —— 这个特征非常明显，
    # 据此主动提醒用户把选项改成"最后一个是余额"，比让用户自己猜要好。
    lonely_amounts = sum(
        1 for issue in result.issues
        if "日期" in issue.reason and "商户" in issue.reason
    )
    if lonely_amounts >= 2 and options.balance_column_mode != "last_is_balance":
        result.hints.append(
            f"检测到 {lonely_amounts} 条只有金额、没有日期和商户的内容，"
            "这通常是行尾的余额列，未被写入。"
        )

    if result.balance_column_detected:
        result.hints.append("已自动识别并排除行尾的余额列。")

    unresolved_dates = sum(1 for issue in result.issues if "缺少日期" in issue.reason)
    if unresolved_dates >= 2 and not options.allow_compact_mmdd:
        result.hints.append(
            "有多条内容缺少可识别的日期，未被写入。"
        )

    result.output_text = build_output_text(result.transactions, options)
    result.total_amount = sum(
        (abs(t.amount) if options.amount_sign_mode == "abs" else t.amount)
        for t in result.transactions
    )
    return result


def build_preprocessor_options(
        date_format_input: str = "",
        balance_column_mode: str = "auto",
        allow_compact_mmdd: bool = False,
) -> ProcessOptions:
    """根据主界面上的设置，构造预处理参数。

    这里顺带做了一个自动推断：如果用户在"日期示例"里填的是用点号分隔的
    写法（例如 4.22），就自动打开预处理的"点号也算日期分隔符"，省得用户
    在两个地方各设置一次。
    """
    sample = str(date_format_input or "")
    uses_dot = "." in sample and not re.search(r"\d\.\d{2}(?!\d)", sample)

    return ProcessOptions(
        allow_dot_separator=uses_dot,
        allow_compact_mmdd=bool(allow_compact_mmdd),
        balance_column_mode=balance_column_mode,
        # 输出零填充的 MM/DD，正好是本程序默认日期规则能识别的格式。
        date_output_format="MM/DD",
        output_style="two_line",
    )


def run_statement_preprocessor(raw_text: str, options: Optional[ProcessOptions] = None):
    """整理原始账单文本，返回 (整理后的文本, 结果对象)。"""
    result = process_text(raw_text, options or ProcessOptions())
    return result.output_text, result


def excel_col_letter(col_num: int) -> str:
    result = ""
    while col_num > 0:
        col_num, rem = divmod(col_num - 1, 26)
        result = chr(65 + rem) + result
    return result


# ================= Category Rules (XLSX) =================

def category_rules_folder() -> Path:
    folder = script_dir() / "Category Rules"
    folder.mkdir(parents=True, exist_ok=True)
    return folder


def category_rules_path() -> Path:
    return category_rules_folder() / "category_rules.xlsx"


def normalize_merchant_for_category(name: str) -> str:
    s = str(name or "").upper().strip()
    s = re.sub(r'[^A-Z0-9]+', ' ', s)

    stop_words = {
        "STORE", "STORES", "MARKET", "ONLINE", "PAYMENT", "PURCHASE",
        "DEBIT", "CREDIT", "CHECKCARD", "CHECK", "POS", "AUTH", "CARD",
        "WITHDRAWAL", "DBT", "ACH", "VISA", "MASTERCARD", "MC"
    }
    parts = [p for p in s.split() if p and p not in stop_words]
    parts = [p for p in parts if not p.isdigit()]

    return " ".join(parts).strip()


def create_default_category_rules_xlsx(path: Path):
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Category Rules"

    ws.cell(row=1, column=1, value="merchant")
    ws.cell(row=1, column=2, value="category")

    sample_rows = [
        ("TARGET", "Office Expense"),
        ("WHOLEFOODS", "Grocery"),
        ("WHOLE FOODS", "Grocery"),
        ("COSTCO", "Grocery"),
        ("UBER", "Travel"),
        ("SHELL", "Gasoline"),
    ]

    row_idx = 2
    for merchant, category in sample_rows:
        ws.cell(row=row_idx, column=1, value=merchant)
        ws.cell(row=row_idx, column=2, value=category)
        row_idx += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    wb.save(path)


def load_category_rules() -> List[Tuple[str, str]]:
    path = category_rules_path()

    if not path.exists():
        create_default_category_rules_xlsx(path)

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    rules = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        merchant = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        category = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

        if merchant and category:
            rules.append((normalize_merchant_for_category(merchant), category))

    return rules


def get_category_for_merchant(merchant: str, rules: List[Tuple[str, str]]) -> str:
    raw = str(merchant or "").strip()
    norm = normalize_merchant_for_category(raw)

    if not norm:
        return ""

    for rule_merchant, category in rules:
        if norm == rule_merchant:
            return category

    for rule_merchant, category in rules:
        if rule_merchant and rule_merchant in norm:
            return category

    for rule_merchant, category in rules:
        if norm and norm in rule_merchant:
            return category

    return ""


# ================= Smart Category Learning =================

CATEGORY_REVIEW_LABEL = "REVIEW"
CATEGORY_FUZZY_THRESHOLD = 0.86
CATEGORY_FUZZY_MARGIN = 0.03


def collect_category_history_from_workbook(wb) -> Dict[str, str]:
    """Collect normalized Merchant -> Category from existing workbook sheets.

    Only non-empty, human-confirmed categories are learned. REVIEW / UNCLASSIFIED
    are deliberately ignored so uncertain guesses never become training data.
    If the same normalized merchant has conflicting categories, the most frequent
    category wins; ties keep the first category encountered.
    """
    counts: Dict[str, OrderedDict] = {}

    for ws in wb.worksheets:
        # Category learning follows the same Debit-layout detection used by the UI.
        # This also supports sheets whose A1 contains a bank name instead of the
        # literal word Merchant.
        layout = find_debit_layout(ws)
        if layout is None:
            continue
        header_row, merchant_col, _, _, category_col = layout
        if category_col is None:
            continue

        for row in range(header_row + 1, ws.max_row + 1):
            merchant = normalized_cell_text(ws.cell(row=row, column=merchant_col).value)
            category = normalized_cell_text(ws.cell(row=row, column=category_col).value)

            if not merchant or merchant.upper() in {"TOTAL", "GRAND TOTAL"}:
                continue
            if not category or category.upper() in {"REVIEW", "UNCLASSIFIED", "UNKNOWN"}:
                continue

            norm = normalize_merchant_for_category(merchant)
            if not norm:
                continue

            bucket = counts.setdefault(norm, OrderedDict())
            bucket[category] = bucket.get(category, 0) + 1

    history: Dict[str, str] = {}
    for norm, bucket in counts.items():
        best_category = max(bucket.items(), key=lambda item: item[1])[0]
        history[norm] = best_category
    return history


# 相似度剪枝下限。低于这个分数的候选，无论真实分数是多少，都不可能影响
# 最终判定，因此可以直接跳过昂贵的字符串比对。
#
# 推导：自动归类要求最高分 >= CATEGORY_FUZZY_THRESHOLD，且
# (最高分 - 第二名 >= CATEGORY_FUZZY_MARGIN) 或 最高分 >= 0.94。
# 第二名只在"离最高分很近"时才起作用，而最高分至少是 0.86，所以只有分数
# 高于 0.86 - 0.03 = 0.83 的候选才可能改变结论。低于它的候选，报 0 还是报
# 真实分数，判定结果完全一样。
CATEGORY_SIMILARITY_PRUNE_FLOOR = CATEGORY_FUZZY_THRESHOLD - CATEGORY_FUZZY_MARGIN


def _merchant_similarity_normalized(a: str, b: str) -> float:
    """相似度计算本体，要求传入的两个名字【已经归一化】。

    拆出这一层是为了性能：调用方（历史模糊匹配）手里的两个字符串本来就
    已经归一化过，没必要在每次比较里重新跑一遍正则和分词。归一化是幂等的，
    所以结果与先前完全一致。
    """
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0

    # 包含关系会直接把分数抬到 0.94，与两者长度差多少无关，所以必须先判断，
    # 不能被下面的长度剪枝挡掉。
    contained = min(len(a), len(b)) >= 4 and (a in b or b in a)

    if not contained:
        # SequenceMatcher 的 ratio 上界是 2*较短长度/(总长度)：匹配字符数
        # 最多就是较短的那个字符串的长度。据此可以在不做任何字符串比对的
        # 情况下算出这一对的分数上限，上限低于剪枝下限就直接放弃。
        length_bound = 2 * min(len(a), len(b)) / (len(a) + len(b))
        if max(length_bound, 0.70 * length_bound + 0.30) < CATEGORY_SIMILARITY_PRUNE_FLOOR:
            return 0.0

    seq = SequenceMatcher(None, a, b).ratio()
    ta, tb = set(a.split()), set(b.split())
    union = ta | tb
    jaccard = (len(ta & tb) / len(union)) if union else 0.0
    score = max(seq, 0.70 * seq + 0.30 * jaccard)

    # Strong brand-prefix/containment signal, but only for meaningful names.
    if contained:
        score = max(score, 0.94)
    return score


def merchant_similarity(a: str, b: str) -> float:
    """Conservative similarity score for two normalized merchant names."""
    return _merchant_similarity_normalized(
        normalize_merchant_for_category(a),
        normalize_merchant_for_category(b),
    )


def get_smart_category_for_merchant(
        merchant: str,
        rules: List[Tuple[str, str]],
        history: Optional[Dict[str, str]] = None,
        fallback: str = CATEGORY_REVIEW_LABEL,
) -> Tuple[str, str, float]:
    """Return (category, source, confidence).

    Priority:
      1) exact normalized match in workbook history
      2) category_rules.xlsx exact/keyword match
      3) conservative fuzzy match against workbook history
      4) REVIEW
    """
    norm = normalize_merchant_for_category(merchant)
    history = history or {}
    if not norm:
        return fallback, "review", 0.0

    if norm in history:
        return history[norm], "history-exact", 1.0

    rule_category = get_category_for_merchant(merchant, rules)
    if rule_category:
        return rule_category, "rule", 1.0

    best_norm = ""
    best_category = ""
    best_score = 0.0
    second_score = 0.0

    # 这里两个名字都已经归一化（norm 来自本函数开头，hist_norm 是历史表的
    # 键），直接走免归一化的版本，省掉每次比较重复跑一遍正则分词。
    for hist_norm, category in history.items():
        score = _merchant_similarity_normalized(norm, hist_norm)
        if score > best_score:
            second_score = best_score
            best_score = score
            best_norm = hist_norm
            best_category = category
        elif score > second_score:
            second_score = score

    # Do not auto-classify an ambiguous fuzzy match.
    if (
            best_category
            and best_score >= CATEGORY_FUZZY_THRESHOLD
            and (best_score - second_score >= CATEGORY_FUZZY_MARGIN or best_score >= 0.94)
    ):
        return best_category, f"history-fuzzy:{best_norm}", best_score

    return fallback, "review", best_score


# ================= 预处理删除内容 =================

def parse_remove_items(raw: str) -> List[str]:
    if not raw:
        return []

    parts = re.split(r'[,\n;，；]+', raw)
    cleaned = []
    seen = set()

    for p in parts:
        item = p.strip()
        if not item:
            continue
        low = item.lower()
        if low not in seen:
            seen.add(low)
            cleaned.append(item)

    return cleaned


def preprocess_statement_text(text: str, remove_items: List[str]) -> str:
    if not text or not remove_items:
        return text

    new_text = text

    for item in remove_items:
        if not item.strip():
            continue
        pattern = re.compile(re.escape(item), re.IGNORECASE)
        new_text = pattern.sub("", new_text)

    cleaned_lines = []
    for line in new_text.splitlines():
        line = re.sub(r'[ \t]+', ' ', line).strip()
        cleaned_lines.append(line)

    return "\n".join(cleaned_lines)


# ================= 把整串文本拆成“伪多行” =================

def build_amount_at_end_pattern() -> str:
    return (
        r'('
        r'(?:-\s*\$?\s*[\d,]+\.\d{2})'
        r'|'
        r'(?:\$?\s*-\s*[\d,]+\.\d{2})'
        r'|'
        r'(?:\(\s*\$?\s*[\d,]+\.\d{2}\s*\))'
        r'|'
        r'(?:\$?\s*[\d,]+\.\d{2})'
        r')\s*$'
    )


def expand_compact_transactions(text: str) -> List[str]:
    text = ' '.join(text.split())
    if not text:
        return []

    matches = list(DATE_IN_TEXT_RE.finditer(text))
    if not matches:
        return [text]

    amount_at_end_re = re.compile(build_amount_at_end_pattern())
    out_lines = []

    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        chunk = text[start:end].strip()

        dm = DATE_IN_TEXT_RE.match(chunk)
        if not dm:
            out_lines.append(chunk)
            continue

        date_str = dm.group(0)
        rest = chunk[dm.end():].strip()

        amt_match = amount_at_end_re.search(rest)
        if not amt_match:
            out_lines.append(chunk)
            continue

        amount_str = amt_match.group(1).strip()
        merchant_str = rest[:amt_match.start()].strip()

        if date_str and merchant_str:
            out_lines.append(f"{date_str} {merchant_str}")
        elif date_str:
            out_lines.append(date_str)

        if amount_str:
            out_lines.append(amount_str)

    return out_lines


def normalize_lines(text: str) -> List[str]:
    text = text.replace('\r\n', '\n').replace('\r', '\n').strip()
    if not text:
        return []

    lines = [ln.strip() for ln in text.split('\n') if ln.strip()]

    if len(lines) > 1:
        rebuilt = []
        for ln in lines:
            expanded = expand_compact_transactions(ln)
            rebuilt.extend(expanded)
        return rebuilt

    return expand_compact_transactions(text)


# ================= 解析器 =================

@dataclass
class ParseResult:
    span: Tuple[int, int]
    merchant: str
    amount: str
    who: str


class BaseExtractor:
    name = "base"

    def extract(self, text: str) -> List['ParseResult']:
        raise NotImplementedError


class SimpleDateStoreAmountExtractor(BaseExtractor):
    name = "simple_date_merchant_then_amount"

    def extract(self, text: str) -> List['ParseResult']:
        lines = normalize_lines(text)
        results: List[ParseResult] = []

        offsets = []
        off = 0
        for ln in lines:
            offsets.append(off)
            off += len(ln) + 1

        i = 0
        n = len(lines)

        while i < n:
            line = lines[i].strip()

            if not line:
                i += 1
                continue

            date_str, merchant_raw = extract_date_and_merchant(line)

            if date_str and merchant_raw:
                start_idx = i
                merchant = clean_merchant(merchant_raw)

                i += 1
                amount = None
                amount_idx = None

                while i < n:
                    cur = lines[i].strip()

                    if not cur:
                        i += 1
                        continue

                    next_date, _ = extract_date_and_merchant(cur)
                    if next_date:
                        break

                    if is_amount(cur):
                        amount = clean_amount(cur)
                        amount_idx = i
                        i += 1
                        break

                    i += 1

                if amount is not None and amount_idx is not None:
                    span = (offsets[start_idx], offsets[amount_idx] + len(lines[amount_idx]))
                    results.append(ParseResult(span, merchant, amount, self.name))

                continue

            i += 1

        return results


EXTRACTORS = {
    "a": SimpleDateStoreAmountExtractor(),
}


def merge_non_overlapping(results: List[ParseResult]) -> List[ParseResult]:
    results = sorted(results, key=lambda r: r.span[0])
    merged: List[ParseResult] = []
    last_end = -1
    for r in results:
        if r.span[0] >= last_end:
            merged.append(r)
            last_end = r.span[1]
    return merged


def parse_with_extractors(text: str, keys: List[str]) -> List[ParseResult]:
    hits: List[ParseResult] = []
    for k in keys:
        ext = EXTRACTORS.get(k)
        if not ext:
            continue
        hits.extend(ext.extract(text))
    return merge_non_overlapping(hits)


def parse_auto(text: str) -> List[ParseResult]:
    return parse_with_extractors(text, ["a"])


# ================= 数据合并 =================

def merge_same_merchants(rows):
    merged = OrderedDict()

    for merchant, amount, who in rows:
        merchant = merchant.strip()
        amount = normalize_amount_string(amount)

        if merchant not in merged:
            merged[merchant] = {
                "amounts": [],
                "who_list": []
            }

        merged[merchant]["amounts"].append(amount)
        merged[merchant]["who_list"].append(who)

    return merged


# ================= Credit 月总表 =================

def normalize_credit_header_name(name: str) -> str:
    s = " ".join(str(name).strip().split())
    return s.upper()


def style_credit_sheet(ws, last_col):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(1, 16):
        for col in range(1, last_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=2).font = Font(bold=True)
    ws.cell(row=1, column=2).alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, last_col + 1):
        ws.cell(row=2, column=col).font = Font(bold=True)

    for row in range(3, 16):
        for col in range(2, last_col + 1):
            ws.cell(row=row, column=col).number_format = "0.00"

    ws.column_dimensions["A"].width = 16
    for col in range(2, last_col + 1):
        ws.column_dimensions[excel_col_letter(col)].width = 22

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24


# ================= Debit Summary Sheet =================

def style_debit_summary_sheet(ws, last_row: int):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_col = 15  # A~O

    for row in range(1, last_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, max_col + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    ws.column_dimensions["A"].width = 42
    for col in range(2, 14):
        ws.column_dimensions[excel_col_letter(col)].width = 12
    ws.column_dimensions["N"].width = 14
    ws.column_dimensions["O"].width = 18

    for row in range(2, last_row + 1):
        for col in range(2, 15):
            ws.cell(row=row, column=col).number_format = "0.00"

    ws.auto_filter.ref = f"A1:O{last_row}"


def sync_credit_debit_from_debit_sheet(wb):
    if "Credit Summary" not in wb.sheetnames:
        return

    if DEBIT_SHEET_NAME not in wb.sheetnames:
        return

    ws_credit = wb["Credit Summary"]
    ws_debit = wb[DEBIT_SHEET_NAME]

    debit_total_row = ws_debit.max_row

    header_col_map = {}
    for col in range(2, ws_credit.max_column + 1):
        val = ws_credit.cell(row=2, column=col).value
        if val:
            header_col_map[str(val).strip()] = col

    if CREDIT_DEBIT_HEADER not in header_col_map:
        return

    debit_col_credit = header_col_map[CREDIT_DEBIT_HEADER]

    for idx, month in enumerate(CREDIT_MONTHS, start=2):
        credit_row = CREDIT_MONTH_ROWS[month]
        debit_month_col_letter = excel_col_letter(idx)
        ws_credit.cell(
            row=credit_row,
            column=debit_col_credit,
            value=f"='{DEBIT_SHEET_NAME}'!{debit_month_col_letter}{debit_total_row}"
        )
        ws_credit.cell(row=credit_row, column=debit_col_credit).number_format = "0.00"

    ws_credit.cell(
        row=15,
        column=debit_col_credit,
        value=f"='{DEBIT_SHEET_NAME}'!N{debit_total_row}"
    )
    ws_credit.cell(row=15, column=debit_col_credit).number_format = "0.00"


# ================= 已有公司报表模板导入 =================

FIXED_CREDIT_HEADERS = {
    CREDIT_BEGIN_HEADER,
    CREDIT_TOTAL_HEADER,
    CREDIT_DEBIT_HEADER,
    CREDIT_ENDING_HEADER,
}


def normalized_cell_text(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def normalized_header_key(value) -> str:
    return normalized_cell_text(value).upper()


def normalize_merchant_key(name: str) -> str:
    """Normalize a merchant/income header name for de-duplication and matching.

    Used to decide whether a newly parsed merchant already has a matching
    column header on a Credit sheet (Row 2) so amounts are added to the same
    column instead of creating duplicate columns for the same merchant due to
    case or whitespace differences.
    """
    return normalized_header_key(name)


def find_header_row(ws, required_headers: List[str], max_scan_rows: int = 80) -> Optional[int]:
    required = {h.upper() for h in required_headers}
    limit = min(ws.max_row, max_scan_rows)

    for row in range(1, limit + 1):
        values = {
            normalized_header_key(ws.cell(row=row, column=col).value)
            for col in range(1, ws.max_column + 1)
        }
        if required.issubset(values):
            return row
    return None


def find_column_by_header(ws, header_row: int, aliases: List[str]) -> Optional[int]:
    alias_set = {a.upper() for a in aliases}
    for col in range(1, ws.max_column + 1):
        if normalized_header_key(ws.cell(row=header_row, column=col).value) in alias_set:
            return col
    return None


def extract_expense_template_from_sheet(ws) -> List[Tuple[str, str]]:
    """Extract Merchant/Category from any recognized Debit layout."""
    layout = find_debit_layout(ws)
    if layout is None:
        return []

    header_row, merchant_col, _, _, category_col = layout
    if category_col is None:
        return []

    rows: List[Tuple[str, str]] = []
    seen = set()

    for row in range(header_row + 1, ws.max_row + 1):
        merchant = normalized_cell_text(ws.cell(row=row, column=merchant_col).value)
        category = normalized_cell_text(ws.cell(row=row, column=category_col).value)

        if not merchant:
            continue
        if merchant.upper() in {"TOTAL", "GRAND TOTAL", "BEGIN", "ADD", "LESS", "ENDING", "ENDING`"}:
            if merchant.upper() in {"TOTAL", "GRAND TOTAL"}:
                break
            continue

        key = merchant.casefold()
        if key in seen:
            continue
        seen.add(key)
        rows.append((merchant, category))

    return rows


def extract_income_headers_from_credit_layout(ws) -> List[str]:
    """Find dynamic income columns from a Credit Summary-style report."""
    limit = min(ws.max_row, 80)
    for row in range(1, limit + 1):
        values = [
            normalized_header_key(ws.cell(row=row, column=col).value)
            for col in range(1, ws.max_column + 1)
        ]
        value_set = set(values)

        # Strong signal for the report layout already used by this program.
        if CREDIT_TOTAL_HEADER in value_set and CREDIT_ENDING_HEADER in value_set:
            headers: List[str] = []
            for value in values:
                if not value or value in FIXED_CREDIT_HEADERS:
                    continue
                if value in {"MONTH", "MERCHANT", "CATEGORY", "TOTAL", "AMOUNT"}:
                    continue
                if value not in headers:
                    headers.append(value)
            return headers
    return []


def extract_income_headers_from_merchant_layout(ws) -> List[str]:
    """Support an Income sheet arranged like Merchant | Jan..Dec | Total | Category."""
    header_row = find_header_row(ws, ["Merchant"])
    if header_row is None:
        return []

    merchant_col = find_column_by_header(ws, header_row, ["Merchant", "Income", "Source", "Description"])
    if merchant_col is None:
        return []

    headers: List[str] = []
    seen = set()
    for row in range(header_row + 1, ws.max_row + 1):
        name = normalized_cell_text(ws.cell(row=row, column=merchant_col).value)
        if not name or name.upper() in {"TOTAL", "GRAND TOTAL"}:
            continue
        key = name.casefold()
        if key in seen:
            continue
        seen.add(key)
        headers.append(normalize_credit_header_name(name))
    return headers


DIRECT_SHEET_REF_RE = re.compile(
    r"^\s*=\s*\+?\s*(?:'((?:[^']|'')+)'|([^'!]+))!\s*(\$?[A-Z]{1,3}\$?\d+)\s*$",
    re.IGNORECASE,
)

DIRECT_LOCAL_REF_RE = re.compile(
    r"^\s*=\s*\+?\s*(\$?[A-Z]{1,3}\$?\d+)\s*$",
    re.IGNORECASE,
)


def shift_local_row_references(ws, from_row: int, offset: int,
                               min_row: Optional[int] = None) -> int:
    """按 Excel 的插入/删除语义，平移同一工作表内公式里的行号。

    Excel 在第 P 行插入 N 行时，工作簿中所有指向第 P 行及以下的引用都会
    自动 +N；删除行时同理 -N。openpyxl 只会把单元格搬走，完全不改公式
    文字，所以这件事必须由我们自己做。

    Debit 模板常见的 Total 下方汇总块（BEGIN / ADD / LESS / END）就依赖
    这个行为，它里面有两类引用，缺一不可：

      * 指向 Total 行本身，例如 ADD 行的 ``=B22``；
      * 块内互相引用，例如 END 行的 ``=B24+B25-B26``（分别指向 BEGIN /
        ADD / LESS 三行）。

    只修第一类是不够的：新增商户会把整个汇总块往下推，块内互相引用如果
    不跟着平移，就会掉头指向上面的商户行，算出一个看起来正常、其实完全
    错误的数字。按行号统一平移可以一次覆盖这两类。

    ``from_row``  从这一行开始（含）的引用需要平移，也就是插入/删除发生
                  的位置；它上面的行没有移动，引用必须原样保留。
    ``offset``    平移量，插入为正、删除为负。
    ``min_row``   限制只重写这一行及以下的公式，调用方必须传。数据区里
                  的商户 Total 公式和 Total 行自己的 SUM 都是程序刚刚
                  按当前行号重新写好的，绝不能再被平移一次。
    """
    if not offset or from_row is None:
        return 0

    start_row = max(1, int(min_row)) if min_row else 1
    if start_row > ws.max_row:
        return 0

    # (?<!!) 排除紧跟在 "!" 后面的部分，那是跨表引用（='Sheet'!B14）的尾巴，
    # 由 shift_cross_sheet_row_references 单独处理。
    ref_re = re.compile(r"(?<!!)(\$?[A-Z]{1,3}\$?)(\d+)\b")

    def repl(match):
        column, row_text = match.group(1), match.group(2)
        row_number = int(row_text)
        if row_number < from_row:
            return match.group(0)
        return f"{column}{row_number + offset}"

    updated = 0
    for row_cells in ws.iter_rows(min_row=start_row):
        for cell in row_cells:
            value = cell.value
            if not isinstance(value, str) or not value.lstrip().startswith("="):
                continue
            new_value = ref_re.sub(repl, value)
            if new_value != value:
                cell.value = new_value
                updated += 1
    return updated


def shift_cross_sheet_row_references(wb, sheet_name: str, from_row: int, offset: int) -> int:
    """同样的行号平移，但作用于【其他工作表】指向本表的跨表引用。

    例如 Credit 表的 DEBIT 列写着 ='CHASE #8565'!G22 指向 Debit 表的 Total
    行。Debit 表插入新商户后 Total 下移，这个引用必须跟着走，否则会一直
    读到一个已经不是 Total 的旧行。除 Total 外，模板也可能引用汇总块里的
    END 行等，所以这里同样按行号统一平移，而不是只认某一个行号。
    """
    if not offset or from_row is None:
        return 0

    escaped_name = re.escape(sheet_name.replace("'", "''"))
    needs_quotes = bool(re.search(r"[^A-Za-z0-9_]", sheet_name)) or sheet_name[:1].isdigit()
    if needs_quotes:
        sheet_pattern = rf"'{escaped_name}'"
    else:
        sheet_pattern = rf"(?:'{escaped_name}'|{re.escape(sheet_name)})"

    ref_re = re.compile(rf"({sheet_pattern}!\s*\$?[A-Z]{{1,3}}\$?)(\d+)\b")

    def repl(match):
        prefix, row_text = match.group(1), match.group(2)
        row_number = int(row_text)
        if row_number < from_row:
            return match.group(0)
        return f"{prefix}{row_number + offset}"

    updated = 0
    for other_ws in wb.worksheets:
        for row_cells in other_ws.iter_rows():
            for cell in row_cells:
                value = cell.value
                if not isinstance(value, str) or not value.lstrip().startswith("="):
                    continue
                new_value = ref_re.sub(repl, value)
                if new_value != value:
                    cell.value = new_value
                    updated += 1
    return updated


def _resolve_direct_reference_cell(cell, visited=None):
    """Resolve a simple Excel reference formula to the cell it points to.

    Supported examples:
      ='CHASE #3444'!A3
      =Income!$B$4
      =A3

    Only direct cell references are followed. Arithmetic/functions are deliberately
    left untouched so the program never tries to become a full Excel formula engine.
    References are followed recursively with cycle protection.
    """
    value = cell.value
    if not isinstance(value, str) or not value.lstrip().startswith("="):
        return cell

    visited = set(visited or ())
    try:
        key = (cell.parent.title, cell.coordinate)
    except Exception:
        return cell
    if key in visited:
        return cell
    visited.add(key)

    formula = value.strip()
    workbook = getattr(cell.parent, "parent", None)
    if workbook is None:
        return cell

    match = DIRECT_SHEET_REF_RE.fullmatch(formula)
    if match:
        quoted_name, plain_name, coordinate = match.groups()
        sheet_name = (quoted_name if quoted_name is not None else plain_name).strip()
        sheet_name = sheet_name.replace("''", "'")
        coordinate = coordinate.replace("$", "")
        if sheet_name not in workbook.sheetnames:
            return cell
        target = workbook[sheet_name][coordinate]
        return _resolve_direct_reference_cell(target, visited)

    match = DIRECT_LOCAL_REF_RE.fullmatch(formula)
    if match:
        coordinate = match.group(1).replace("$", "")
        target = cell.parent[coordinate]
        return _resolve_direct_reference_cell(target, visited)

    return cell


def display_cell_text(cell) -> str:
    """Return the actual readable label represented by an Excel cell.

    If the cell is a direct formula reference to another sheet (for example
    ``='CHASE #3444'!A3``), follow the reference first and display the source
    cell's real value. This is especially important for Month headers used by
    the UI. The current sheet/row/column remains the write target; only the
    display label is resolved from the referenced cell.
    """
    resolved_cell = _resolve_direct_reference_cell(cell)
    value = resolved_cell.value
    if value is None:
        return ""

    # Excel date/datetime cells should be displayed using their own number format
    # where possible, instead of Python's ``2026-01-01 00:00:00`` representation.
    try:
        if getattr(resolved_cell, "is_date", False) and hasattr(value, "strftime"):
            fmt = str(resolved_cell.number_format or "").lower()

            # Preserve common month-header formats such as Apr-24 / April-2024
            # instead of converting them to a full numeric date.
            month_token = None
            if "mmmm" in fmt:
                month_token = "%B"
            elif "mmm" in fmt:
                month_token = "%b"

            year_token = "%Y" if "yyyy" in fmt else ("%y" if "yy" in fmt else None)
            if month_token and year_token:
                sep = "-" if "-" in fmt else ("/" if "/" in fmt else " ")
                return value.strftime(month_token) + sep + value.strftime(year_token)
            if month_token:
                return value.strftime(month_token)

            if "yyyy" in fmt:
                return value.strftime("%m/%d/%Y")
            if "yy" in fmt:
                return value.strftime("%m/%d/%y")
            return value.strftime("%m/%d")
    except Exception:
        pass

    return normalized_cell_text(value)


def looks_like_period_header(cell) -> bool:
    """Return True when a header cell looks like a month/date/period label.

    This includes normal month names (AUG), date-like labels (31-Jul, 6-Aug,
    08/06/2026), and real Excel date cells. It allows valid period columns
    to appear after Total/Category instead of requiring one contiguous block.
    """
    value = cell.value
    if value is None:
        return False
    try:
        resolved_cell = _resolve_direct_reference_cell(cell)
        if getattr(resolved_cell, "is_date", False):
            return True
    except Exception:
        pass

    text = display_cell_text(cell).strip()
    if not text:
        return False
    key = text.upper().replace(".", "")
    if key in {
        "JAN", "JANUARY", "FEB", "FEBRUARY", "MAR", "MARCH",
        "APR", "APRIL", "MAY", "JUN", "JUNE", "JUL", "JULY",
        "AUG", "AUGUST", "SEP", "SEPT", "SEPTEMBER",
        "OCT", "OCTOBER", "NOV", "NOVEMBER", "DEC", "DECEMBER"
    }:
        return True

    month_word = r"(?:JAN(?:UARY)?|FEB(?:RUARY)?|MAR(?:CH)?|APR(?:IL)?|MAY|JUN(?:E)?|JUL(?:Y)?|AUG(?:UST)?|SEP(?:T(?:EMBER)?)?|OCT(?:OBER)?|NOV(?:EMBER)?|DEC(?:EMBER)?)"
    patterns = [
        # Month + year, e.g. Apr-24 / April 2024.
        rf"^{month_word}[-/\s]\d{{2,4}}$",
        rf"^\d{{2,4}}[-/\s]{month_word}$",
        # Day + month or month + day, optionally with year.
        rf"^\d{{1,2}}[-/\s]{month_word}(?:[-/\s]\d{{2,4}})?$",
        rf"^{month_word}[-/\s]\d{{1,2}}(?:[-/\s]\d{{2,4}})?$",
        r"^\d{1,2}[-/]\d{1,2}(?:[-/]\d{2,4})?$",
        r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$",
    ]
    return any(re.fullmatch(p, text, re.IGNORECASE) for p in patterns)


def build_sum_formula_for_columns(ws, row: int, columns: List[int]) -> str:
    """Build a SUM formula for possibly non-contiguous period columns."""
    refs = [ws.cell(row=row, column=c).coordinate for c in columns]
    return f"=SUM({','.join(refs)})" if refs else ""


def extract_month_labels_from_sheet(ws) -> List[str]:
    """Read horizontal Debit period/date labels from the detected Debit layout.

    This deliberately does not require a literal ``Merchant`` header.  It uses
    ``find_debit_layout`` so UI Month detection, import, clearing and writing all
    share the exact same Debit-layout rules.
    """
    layout = find_debit_layout(ws)
    if layout is None:
        return []

    header_row, _, month_cols, _, _ = layout
    labels: List[str] = []
    for col in month_cols:
        cell = ws.cell(row=header_row, column=col)
        if not looks_like_period_header(cell):
            continue
        text = display_cell_text(cell).strip()
        if text:
            labels.append(text)
    return labels


def find_credit_period_rows(ws, header_row: int, begin_col: int) -> Tuple[List[int], Optional[int]]:
    """Return all labeled period rows and the TOTAL row for a Credit layout."""
    if begin_col <= 1:
        return [], None

    label_col = begin_col - 1
    period_rows: List[int] = []
    total_row: Optional[int] = None

    for row in range(header_row + 1, ws.max_row + 1):
        label = display_cell_text(ws.cell(row=row, column=label_col)).strip()
        key = label.upper()
        if key in {"TOTAL", "GRAND TOTAL"}:
            total_row = row
            break
        if label:
            period_rows.append(row)

    # Some templates may not already contain a TOTAL row.
    if total_row is None and period_rows:
        total_row = period_rows[-1] + 1
    return period_rows, total_row


def extract_month_labels_from_income_sheet(ws) -> List[str]:
    """Read every date/period label from an Income/Credit page."""
    limit = min(ws.max_row, 100)
    begin_row = None
    begin_col = None
    for row in range(1, limit + 1):
        for col in range(1, ws.max_column + 1):
            if normalized_header_key(ws.cell(row=row, column=col).value) == CREDIT_BEGIN_HEADER:
                begin_row, begin_col = row, col
                break
        if begin_row is not None:
            break

    if begin_row is not None and begin_col is not None and begin_col > 1:
        period_rows, _ = find_credit_period_rows(ws, begin_row, begin_col)
        labels = [display_cell_text(ws.cell(row=r, column=begin_col - 1)) for r in period_rows]
        if labels:
            return labels

    return extract_month_labels_from_sheet(ws)


def discover_source_report_structure(source_path: Path):
    """Discover income names and expense Merchant/Category rows.

    Amount cells from the source workbook are deliberately never imported.
    """
    wb = load_workbook(source_path, data_only=False)

    income_headers: List[str] = []
    expense_rows: List[Tuple[str, str]] = []
    month_labels: List[str] = []

    # Income is scanned first because its 12 date/period labels are the
    # authoritative labels for both generated sheets and for the UI dropdown.
    ordered_income_sheets = sorted(
        wb.worksheets,
        key=lambda ws: (0 if any(k in ws.title.lower() for k in ("income", "credit")) else 1)
    )
    for ws in ordered_income_sheets:
        found = extract_income_headers_from_credit_layout(ws)
        if not found and any(k in ws.title.lower() for k in ("income", "credit")):
            found = extract_income_headers_from_merchant_layout(ws)

        detected_labels = extract_month_labels_from_income_sheet(ws)
        if found:
            income_headers = found
            if detected_labels:
                month_labels = detected_labels
            break

    # Expense contributes only Merchant and Category. Its own date headers are
    # deliberately ignored because the Income page controls the date labels.
    ordered_expense_sheets = sorted(
        wb.worksheets,
        key=lambda ws: (0 if any(k in ws.title.lower() for k in ("expense", "debit")) else 1)
    )
    for ws in ordered_expense_sheets:
        found = extract_expense_template_from_sheet(ws)
        if found:
            expense_rows = found
            break

    wb.close()

    # Always keep the program's standard deposit column.
    final_income_headers = [CREDIT_DEFAULT_HEADER]
    for header in income_headers:
        normalized = normalize_credit_header_name(header)
        if normalized and normalized not in FIXED_CREDIT_HEADERS and normalized not in final_income_headers:
            final_income_headers.append(normalized)

    return final_income_headers, expense_rows, normalize_month_labels(month_labels)


def create_credit_template_sheet(wb, income_headers: List[str], bank_name: str, month_labels: List[str]):
    if "Credit Summary" in wb.sheetnames:
        wb.remove(wb["Credit Summary"])

    ws = wb.create_sheet("Credit Summary", 0)
    ws.cell(row=1, column=2, value=(bank_name or DEFAULT_BANK_NAME).strip() or DEFAULT_BANK_NAME)

    dynamic_headers = []
    for header in income_headers:
        normalized = normalize_credit_header_name(header)
        if normalized and normalized not in FIXED_CREDIT_HEADERS and normalized not in dynamic_headers:
            dynamic_headers.append(normalized)
    if CREDIT_DEFAULT_HEADER not in dynamic_headers:
        dynamic_headers.insert(0, CREDIT_DEFAULT_HEADER)

    headers = [CREDIT_BEGIN_HEADER] + dynamic_headers + [
        CREDIT_TOTAL_HEADER,
        CREDIT_DEBIT_HEADER,
        CREDIT_ENDING_HEADER,
    ]

    for col, header in enumerate(headers, start=2):
        ws.cell(row=2, column=col, value=header)

    month_labels = normalize_month_labels(month_labels)
    for index, month in enumerate(CREDIT_MONTHS):
        ws.cell(row=CREDIT_MONTH_ROWS[month], column=1, value=month_labels[index])
    ws.cell(row=15, column=1, value="TOTAL")

    header_col_map = {
        normalized_cell_text(ws.cell(row=2, column=col).value): col
        for col in range(2, len(headers) + 2)
    }
    begin_col = header_col_map[CREDIT_BEGIN_HEADER]
    total_credit_col = header_col_map[CREDIT_TOTAL_HEADER]
    debit_col = header_col_map[CREDIT_DEBIT_HEADER]
    ending_col = header_col_map[CREDIT_ENDING_HEADER]
    dynamic_start_col = header_col_map[dynamic_headers[0]]
    dynamic_end_col = header_col_map[dynamic_headers[-1]]

    # All imported historical amounts are intentionally blank.
    ws.cell(row=CREDIT_MONTH_ROWS["JAN"], column=begin_col, value=0)
    ws.cell(row=CREDIT_MONTH_ROWS["JAN"], column=begin_col).number_format = "0.00"

    for index, month in enumerate(CREDIT_MONTHS):
        row = CREDIT_MONTH_ROWS[month]
        if index > 0:
            previous_row = CREDIT_MONTH_ROWS[CREDIT_MONTHS[index - 1]]
            previous_ending = ws.cell(row=previous_row, column=ending_col).coordinate
            ws.cell(row=row, column=begin_col, value=f"={previous_ending}")

        start_ref = ws.cell(row=row, column=dynamic_start_col).coordinate
        end_ref = ws.cell(row=row, column=dynamic_end_col).coordinate
        ws.cell(row=row, column=total_credit_col, value=f"=SUM({start_ref}:{end_ref})")

        begin_ref = ws.cell(row=row, column=begin_col).coordinate
        total_credit_ref = ws.cell(row=row, column=total_credit_col).coordinate
        debit_ref = ws.cell(row=row, column=debit_col).coordinate
        ws.cell(row=row, column=ending_col, value=f"={begin_ref}+{total_credit_ref}-{debit_ref}")

    # Annual formulas remain formulas, not pasted values.
    for header in dynamic_headers:
        col = header_col_map[header]
        start_ref = ws.cell(row=CREDIT_MONTH_ROWS["JAN"], column=col).coordinate
        end_ref = ws.cell(row=CREDIT_MONTH_ROWS["DEC"], column=col).coordinate
        ws.cell(row=15, column=col, value=f"=SUM({start_ref}:{end_ref})")

    start_ref = ws.cell(row=CREDIT_MONTH_ROWS["JAN"], column=total_credit_col).coordinate
    end_ref = ws.cell(row=CREDIT_MONTH_ROWS["DEC"], column=total_credit_col).coordinate
    ws.cell(row=15, column=total_credit_col, value=f"=SUM({start_ref}:{end_ref})")

    start_ref = ws.cell(row=CREDIT_MONTH_ROWS["JAN"], column=debit_col).coordinate
    end_ref = ws.cell(row=CREDIT_MONTH_ROWS["DEC"], column=debit_col).coordinate
    ws.cell(row=15, column=debit_col, value=f"=SUM({start_ref}:{end_ref})")

    # TOTAL row intentionally leaves BOTH Begin and Ending untouched -- only
    # the columns strictly between them get a value here. The user fills
    # Ending in manually, so this must never write to it.

    last_col = 1 + len(headers)
    style_credit_sheet(ws, last_col)
    ws.auto_filter.ref = f"A2:{excel_col_letter(last_col)}15"


def create_debit_template_sheet(wb, expense_rows: List[Tuple[str, str]], month_labels: List[str]):
    if DEBIT_SHEET_NAME in wb.sheetnames:
        wb.remove(wb[DEBIT_SHEET_NAME])

    ws = wb.create_sheet(DEBIT_SHEET_NAME)
    headers = ["Merchant"] + month_labels + ["Total", "Category"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    row = 2
    for merchant, category in expense_rows:
        ws.cell(row=row, column=1, value=merchant)
        # Jan-Dec remain blank. Total is still a live formula.
        ws.cell(row=row, column=14, value=f"=SUM(B{row}:M{row})")
        ws.cell(row=row, column=14).number_format = "0.00"
        ws.cell(row=row, column=15, value=category)
        row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value="Total")
    for col in range(2, 14):
        letter = excel_col_letter(col)
        if total_row > 2:
            ws.cell(row=total_row, column=col, value=f"=SUM({letter}2:{letter}{total_row - 1})")
            ws.cell(row=total_row, column=col).number_format = "0.00"
    if total_row > 2:
        ws.cell(row=total_row, column=14, value=f"=SUM(B{total_row}:M{total_row})")
        ws.cell(row=total_row, column=14).number_format = "0.00"

    style_debit_summary_sheet(ws, total_row)


def create_default_monthly_workbook(xlsx_path: Path, bank_name: str = DEFAULT_BANK_NAME):
    """Create a brand-new default workbook with exactly two sheets.

    Sheet 1: Credit Summary
    Sheet 2: debit summary
    Both use the standard Jan-Dec period labels and the existing formula/layout
    conventions used by this program.
    """
    xlsx_path = Path(xlsx_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Remove openpyxl's automatic blank sheet so the workbook contains exactly
    # the two program sheets requested by the user.
    default_ws = wb.active
    wb.remove(default_ws)

    month_labels = list(DEFAULT_UI_MONTH_LABELS)
    create_credit_template_sheet(
        wb,
        [CREDIT_DEFAULT_HEADER],
        bank_name=(bank_name or DEFAULT_BANK_NAME).strip() or DEFAULT_BANK_NAME,
        month_labels=month_labels,
    )
    create_debit_template_sheet(wb, [], month_labels)

    # Link Credit Summary's DEBIT column to the monthly totals on debit summary.
    # This is done only after both sheets exist.  The same sync helper is also
    # called after future Debit writes so references follow the Debit Total row
    # when new Merchant rows are inserted.
    sync_credit_debit_from_debit_sheet(wb)

    # Explicitly enforce the requested order even if the helper implementations
    # are changed later.
    wb._sheets = [wb["Credit Summary"], wb[DEBIT_SHEET_NAME]]
    wb.save(xlsx_path)
    wb.close()
    return xlsx_path


def clear_credit_amounts_keep_layout(ws) -> int:
    """Clear Credit amounts after the SINGLE Column-A type rule classifies the sheet."""
    if not is_credit_sheet_by_column_a(ws):
        return 0

    # Layout lookup is positioning only; it never decides the sheet type.
    header_row, header_map = locate_credit_columns(ws)
    if header_row is None:
        return 0

    begin_col = header_map[CREDIT_BEGIN_HEADER]
    total_credit_col = header_map[CREDIT_TOTAL_HEADER]
    debit_col = header_map[CREDIT_DEBIT_HEADER]
    ending_col = header_map[CREDIT_ENDING_HEADER]
    period_rows, total_row = find_credit_period_rows(ws, header_row, begin_col)
    if not period_rows:
        return 0
    first_data_row = period_rows[0]
    last_data_row = period_rows[-1]
    if total_row is None:
        total_row = last_data_row + 1

    dynamic_cols = list(range(begin_col + 1, total_credit_col))
    for row in period_rows:
        for col in dynamic_cols:
            ws.cell(row=row, column=col).value = None
        # The DEBIT column is different from the dynamic income columns above:
        # those hold plain historical amounts and should always be reset, but
        # DEBIT very often holds a FORMULA -- most commonly a cross-sheet link
        # to the paired Debit sheet's Total row (e.g. ='Some Debit Sheet'!G14),
        # imported as-is from the original company report. Blindly clearing it
        # to None destroys that link entirely, so future Debit entries would
        # stop showing up here at all. Only clear a literal stored amount;
        # leave any formula (the structural link) untouched.
        debit_cell = ws.cell(row=row, column=debit_col)
        debit_value = debit_cell.value
        if not (isinstance(debit_value, str) and debit_value.lstrip().startswith("=")):
            debit_cell.value = None

    # Rebuild the controlled formulas without renaming or recreating the sheet.
    if first_data_row <= ws.max_row:
        ws.cell(row=first_data_row, column=begin_col, value=0)
        ws.cell(row=first_data_row, column=begin_col).number_format = "0.00"

    for row_index, row in enumerate(period_rows):
        if row_index > 0:
            prev_end = ws.cell(row=period_rows[row_index - 1], column=ending_col).coordinate
            ws.cell(row=row, column=begin_col, value=f"={prev_end}")
        if dynamic_cols:
            start_ref = ws.cell(row=row, column=dynamic_cols[0]).coordinate
            end_ref = ws.cell(row=row, column=dynamic_cols[-1]).coordinate
            ws.cell(row=row, column=total_credit_col, value=f"=SUM({start_ref}:{end_ref})")
        else:
            ws.cell(row=row, column=total_credit_col, value=0)
        begin_ref = ws.cell(row=row, column=begin_col).coordinate
        credit_ref = ws.cell(row=row, column=total_credit_col).coordinate
        debit_ref = ws.cell(row=row, column=debit_col).coordinate
        ws.cell(row=row, column=ending_col, value=f"={begin_ref}+{credit_ref}-{debit_ref}")

    for col in dynamic_cols + [total_credit_col, debit_col]:
        start_ref = ws.cell(row=first_data_row, column=col).coordinate
        end_ref = ws.cell(row=last_data_row, column=col).coordinate
        ws.cell(row=total_row, column=col, value=f"=SUM({start_ref}:{end_ref})")
    # TOTAL row intentionally leaves BOTH Begin and Ending untouched -- only
    # the columns strictly between them get a value here. The user fills
    # Ending in manually, so this must never write to it.
    return len(dynamic_cols)


def clear_debit_amounts_keep_layout(ws) -> int:
    """Clear historical Debit amounts in-place; keep Merchant/Category and sheet name."""
    layout = find_debit_layout(ws)
    if layout is None:
        return 0
    header_row, merchant_col, month_cols, total_col, category_col = layout

    total_row = None
    merchant_count = 0
    for row in range(header_row + 1, ws.max_row + 1):
        merchant = normalized_cell_text(ws.cell(row=row, column=merchant_col).value)
        if not merchant:
            continue
        if merchant.upper() in {"TOTAL", "GRAND TOTAL"}:
            total_row = row
            break
        merchant_count += 1
        for col in month_cols:
            ws.cell(row=row, column=col).value = None
        ws.cell(row=row, column=total_col, value=build_sum_formula_for_columns(ws, row, month_cols))
        ws.cell(row=row, column=total_col).number_format = "0.00"

    if total_row is None:
        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=merchant_col, value="Total")

    first_data_row = header_row + 1
    last_data_row = total_row - 1
    if last_data_row >= first_data_row:
        for col in month_cols:
            letter = excel_col_letter(col)
            ws.cell(row=total_row, column=col,
                    value=f"=SUM({letter}{first_data_row}:{letter}{last_data_row})")
            ws.cell(row=total_row, column=col).number_format = "0.00"
        ws.cell(row=total_row, column=total_col,
                value=build_sum_formula_for_columns(ws, total_row, month_cols))
        ws.cell(row=total_row, column=total_col).number_format = "0.00"
    return merchant_count


def import_existing_report_template(
        source_path: Path,
        target_path: Path,
        bank_name: str = DEFAULT_BANK_NAME,
        clear_existing_data: bool = True,
):
    """完整复制源工作簿的所有工作表，并保留原始名称和顺序。

    clear_existing_data=True：模板模式，清空可识别 Credit/Debit 页面中的旧月份金额，
    但保留工作表名称、顺序、格式、公式框架、Merchant 与 Category。

    clear_existing_data=False：完整保留模式，源 Excel 中所有原有数据、公式、格式、
    工作表名称和顺序均保持不变，之后可在 UI 中继续追加新数据。
    """
    import shutil
    import tempfile

    source_path = Path(source_path).expanduser().resolve()
    target_path = Path(target_path).expanduser().resolve()

    if not source_path.exists():
        raise FileNotFoundError(f"找不到源 Excel：{source_path}")
    if source_path.suffix.lower() != ".xlsx":
        raise ValueError("目前只支持 .xlsx 文件。")
    if source_path == target_path:
        raise ValueError("源报表和目标 Excel 不能是同一个文件。")

    # 先读取源工作簿并记录所有工作表名称。这里包含隐藏工作表。
    source_wb = load_workbook(source_path, read_only=True, data_only=False)
    try:
        source_sheet_names = list(source_wb.sheetnames)
    finally:
        source_wb.close()

    if not source_sheet_names:
        raise ValueError("源 Excel 中没有可读取的Excel Sheet。")

    target_path.parent.mkdir(parents=True, exist_ok=True)

    # 使用同目录临时文件，避免覆盖过程中出现只剩部分工作表的损坏文件。
    fd, temp_name = tempfile.mkstemp(
        prefix=target_path.stem + "_import_",
        suffix=".xlsx",
        dir=str(target_path.parent),
    )
    import os
    os.close(fd)
    temp_path = Path(temp_name)

    try:
        shutil.copyfile(source_path, temp_path)

        # 第一次验证：整本复制后，所有 sheet 名称和顺序必须完全一致。
        check_wb = load_workbook(temp_path, read_only=True, data_only=False)
        try:
            copied_sheet_names = list(check_wb.sheetnames)
        finally:
            check_wb.close()

        if copied_sheet_names != source_sheet_names:
            raise RuntimeError(
                "Excel Sheet复制验证失败。\n"
                f"源文件Excel Sheet({len(source_sheet_names)}): {source_sheet_names}\n"
                f"复制后Excel Sheet({len(copied_sheet_names)}): {copied_sheet_names}"
            )

        wb = load_workbook(temp_path, data_only=False)
        try:
            credit_item_count = 0
            debit_merchant_count = 0
            month_labels: List[str] = []

            # Month/period labels must come from a real recognized layout.
            # Prefer Credit/Income because its period labels are authoritative;
            # only fall back to Debit/Expense if no Credit sheet is available.
            # Month/type detection uses one rule everywhere:
            # Column A has a month/date period => Credit; all other sheets => Debit.
            for candidate_ws in list(wb.worksheets):
                credit_options = get_credit_period_options_from_column_a(candidate_ws)
                # Insurance rule: only >5 recognizable Month/date values in Column A
                # makes this a Credit sheet. 1~5 accidental dates must not steal
                # the workbook's authoritative Month labels.
                if len(credit_options) > 5:
                    month_labels = [label for label, _ in credit_options]
                    break
            if not month_labels:
                for candidate_ws in list(wb.worksheets):
                    debit_layout = find_debit_layout(candidate_ws)
                    if debit_layout is not None:
                        labels = extract_month_labels_from_sheet(candidate_ws)
                        if labels:
                            month_labels = list(labels)
                            break

            # 注意：这里只读取/修改工作表内容，绝不 remove/create/rename worksheet。
            for ws in list(wb.worksheets):
                if is_credit_sheet_by_column_a(ws):
                    if clear_existing_data:
                        credit_item_count += clear_credit_amounts_keep_layout(ws)
                    else:
                        # 类型只由 Column A 决定；下面仅统计可定位到的 Credit 控制/收入列。
                        layout = locate_credit_columns(ws)
                        if layout[0] is not None:
                            credit_item_count += max(0, len(layout[1]))
                    continue

                # All non-Credit sheets are classified as Debit.  Only recognizable
                # Debit layouts are modified/counted; Cover/Notes sheets remain untouched.
                debit_layout = find_debit_layout(ws)
                if debit_layout is not None:
                    if clear_existing_data:
                        debit_merchant_count += clear_debit_amounts_keep_layout(ws)
                    else:
                        header_row_d, merchant_col, _, _, _ = debit_layout
                        for row in range(header_row_d + 1, ws.max_row + 1):
                            merchant = normalized_cell_text(ws.cell(row=row, column=merchant_col).value)
                            if not merchant or merchant.upper() in {"TOTAL", "GRAND TOTAL"}:
                                continue
                            debit_merchant_count += 1

            # 保存前再次确认代码没有意外删除、增加或改名。
            before_save_names = list(wb.sheetnames)
            if before_save_names != source_sheet_names:
                raise RuntimeError(
                    "处理过程中Excel Sheet名称或数量发生变化。\n"
                    f"源文件: {source_sheet_names}\n处理后: {before_save_names}"
                )
            wb.save(temp_path)
        finally:
            wb.close()

        # 第二次验证：保存并重新打开后，仍然必须保留全部工作表。
        final_check_wb = load_workbook(temp_path, read_only=True, data_only=False)
        try:
            final_sheet_names = list(final_check_wb.sheetnames)
        finally:
            final_check_wb.close()

        if final_sheet_names != source_sheet_names:
            raise RuntimeError(
                "保存后的Excel Sheet验证失败。\n"
                f"源文件Excel Sheet({len(source_sheet_names)}): {source_sheet_names}\n"
                f"保存后Excel Sheet({len(final_sheet_names)}): {final_sheet_names}"
            )

        # 所有验证通过后再替换正式目标文件。
        os.replace(temp_path, target_path)

        return (
            credit_item_count,
            debit_merchant_count,
            (month_labels if month_labels else list(DEFAULT_UI_MONTH_LABELS)),
            final_sheet_names,
        )
    finally:
        if temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


# ================= 多工作表读取与原位追加 =================

def list_workbook_sheet_names(path: Path) -> List[str]:
    path = Path(path)
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def get_credit_period_options_from_column_a(ws) -> List[Tuple[str, int]]:
    """Return Credit Month options found specifically in worksheet Column A.

    Program rule: a worksheet is Credit only when Column A contains more than
    5 recognizable month/date/period values.  The returned position is the real row
    number so the UI Month stays directly linked to the Credit target row.

    The scan is intentionally bounded to CREDIT_COLUMN_A_SCAN_MAX_ROWS (the real
    Credit layout is at most ~15 rows: 12 months + 1 TOTAL row). Without this
    bound, stray date-like text left far below the real data on a large/old
    worksheet (for example leftover formatting or unrelated text many rows
    down) could be miscounted as extra Credit periods and cause a normal
    Debit sheet to be misclassified as Credit -- which is what produced the
    transposed/garbled layout (merchant names turned into column headers).

    Direct-reference formulas are supported through ``display_cell_text`` /
    ``looks_like_period_header``; for example ``='CHASE #3444'!A3`` can resolve
    to ``Apr-24`` while the writable target remains the current worksheet row.
    """
    options: List[Tuple[str, int]] = []
    limit = min(ws.max_row, CREDIT_COLUMN_A_SCAN_MAX_ROWS)
    for row in range(1, limit + 1):
        cell = ws.cell(row=row, column=1)
        if not looks_like_period_header(cell):
            continue
        label = display_cell_text(cell).strip()
        if label:
            options.append((label, row))
    return options


def is_credit_sheet_by_column_a(ws) -> bool:
    """Credit only when Column A contains more than 5 recognizable periods."""
    return len(get_credit_period_options_from_column_a(ws)) > 5


def classify_sheet_type(ws) -> str:
    """Column A has >5 Month/date periods => Credit; otherwise Debit."""
    return "credit" if is_credit_sheet_by_column_a(ws) else "debit"


def get_period_options_from_ws(ws) -> Tuple[str, List[Tuple[str, int]]]:
    """Classify by Column A and return Month UI options with real positions.

    Credit:
      - The ONLY type-detection rule is whether Column A contains more than 5 month/date values.
      - Month options come directly from Column A.
      - Position = real Excel row number.

    Debit:
      - Every sheet that fails the Column-A Credit rule is classified as Debit.
      - Merchant is fixed at Column A.
      - Month/date cells are read ONLY from Row 1, horizontally from Column B onward.
      - Position = real Excel column number; no Debit header keywords are used.
    """
    credit_options = get_credit_period_options_from_column_a(ws)
    # IMPORTANT: keep the exact same classification rule used everywhere else.
    # Column A must contain MORE THAN 5 recognizable Month/date values.
    if len(credit_options) > 5:
        return "credit", credit_options

    # Per the new rule, every non-Credit sheet is Debit.  It may still have no
    # writable Debit layout (e.g. a Cover sheet), in which case options stay empty.
    debit_layout = find_debit_layout(ws)
    if debit_layout is None:
        return "debit", []

    header_row, _, month_cols, _, category_col = debit_layout
    all_cols = list(month_cols) + find_extra_debit_period_columns(ws, category_col)
    options: List[Tuple[str, int]] = []
    for col in all_cols:
        label = display_cell_text(ws.cell(row=header_row, column=col)).strip()
        if label:
            options.append((label, col))
    return "debit", options


def read_period_options_from_selected_sheet(path: Path, sheet_name: str) -> Tuple[str, List[Tuple[str, int]]]:
    """Read the saved workbook, classify once, then build the UI Month mapping."""
    path = Path(path)
    if not path.exists() or not sheet_name:
        return "", []
    # IMPORTANT: do not open with read_only=True here. openpyxl's read-only mode
    # is optimized for sequential streaming; ws.cell(row=, column=) random access
    # (used throughout classification and layout detection) forces it to
    # internally re-scan from the top of the sheet on every call, which turns an
    # otherwise-fast lookup into an O(n^2) operation on large worksheets and is
    # the direct cause of the UI freezing when switching sheets.
    wb = load_workbook(path, read_only=False, data_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            return "", []
        ws = wb[sheet_name]

        # ONE classification decision for this UI read, reused via cache so
        # repeated sheet switches don't keep re-scanning Column A.
        sheet_type = get_cached_sheet_type(path, sheet_name)
        if sheet_type is None:
            sheet_type = classify_sheet_type(ws)
            cache_sheet_type(path, sheet_name, sheet_type)

        if sheet_type == "credit":
            return "credit", get_credit_period_options_from_column_a(ws)

        # Debit: A is Merchant; ALL UI Month mapping comes from Row 1, B1 onward.
        layout = get_cached_debit_layout(path, sheet_name)
        if layout is None:
            layout = find_debit_layout(ws)
            if layout is not None:
                cache_debit_layout(path, sheet_name, layout)
        if layout is None:
            return "debit", []

        header_row, _, month_cols, _, category_col = layout
        all_cols = list(month_cols) + find_extra_debit_period_columns(ws, category_col)
        options = []
        for col in all_cols:
            label = display_cell_text(ws.cell(row=header_row, column=col)).strip()
            if label:
                options.append((label, col))
        return "debit", options
    finally:
        wb.close()


def read_sheet_period_status(path: Path, sheet_name: str):
    """一次读取工作表，同时返回类型、Month 选项、以及每个 Month 是否已有数据。

    "哪些月份已经写完"这件事的唯一可靠来源是 Excel 文件本身，而不是程序
    运行期间的操作记录 —— 后者一关程序就没了，换台电脑也看不到。所以这里
    直接去表里数：某个 Month 的数据区只要有一个非空单元格，就算已写入。

    判断只看【数据区】，不看 Total 行：Total 行永远挂着 SUM 公式，哪怕一分
    钱都没写也是非空的，拿它判断会全部显示成已完成。

    返回 (sheet_type, [(标签, 位置), ...], {位置: 是否已有数据}, {位置: 金额合计})。
    金额合计是把该月份数据区的单元格逐个求值相加得到的 —— openpyxl 不会计算
    公式，所以不能直接读 Total 行的 =SUM(...)，只能自己把 "=100.00+200.00"
    这类累加式拆开求和（safe_float already handles that form）。这个数字就是
    Excel 里 Total 行会显示的值，可以直接和银行账单核对。

    与 read_period_options_from_selected_sheet 共用同一次文件打开，避免
    切换 Sheet 时把工作簿读两遍。
    """
    path = Path(path)
    if not path.exists() or not sheet_name:
        return "", [], {}, {}

    # 与 read_period_options_from_selected_sheet 一样，这里不能用 read_only：
    # 随机访问单元格在流式模式下会退化成每次从头扫描。
    wb = load_workbook(path, read_only=False, data_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            return "", [], {}, {}
        ws = wb[sheet_name]

        sheet_type = get_cached_sheet_type(path, sheet_name)
        if sheet_type is None:
            sheet_type = classify_sheet_type(ws)
            cache_sheet_type(path, sheet_name, sheet_type)

        def scan(cells):
            """返回 (是否有非空单元格, 金额合计)。"""
            found = False
            total = 0.0
            for value in cells:
                if value is None or str(value).strip() == "":
                    continue
                found = True
                total += safe_float(value)
            return found, total

        if sheet_type == "credit":
            options = get_credit_period_options_from_column_a(ws)
            header_row, header_map = locate_credit_columns(ws)
            written, totals = {}, {}
            if header_row is not None and header_map:
                begin_col = header_map[CREDIT_BEGIN_HEADER]
                total_credit_col = header_map[CREDIT_TOTAL_HEADER]
                # Credit 的数据区是 Begin 与 Total Credit 之间的收入列。
                income_cols = list(range(begin_col + 1, total_credit_col))
                for _, row in options:
                    found, total = scan(
                        ws.cell(row=row, column=col).value for col in income_cols
                    )
                    written[row] = found
                    totals[row] = total
            return "credit", options, written, totals

        layout = get_cached_debit_layout(path, sheet_name)
        if layout is None:
            layout = find_debit_layout(ws)
            if layout is not None:
                cache_debit_layout(path, sheet_name, layout)
        if layout is None:
            return "debit", [], {}, {}

        header_row, merchant_col, month_cols, _, category_col = layout
        all_cols = list(month_cols) + find_extra_debit_period_columns(ws, category_col)

        # 数据区到 Total 行为止。
        last_data_row = ws.max_row
        for row in range(header_row + 1, ws.max_row + 1):
            label = normalized_cell_text(ws.cell(row=row, column=merchant_col).value)
            if label.upper() in {"TOTAL", "GRAND TOTAL"}:
                last_data_row = row - 1
                break

        options, written, totals = [], {}, {}
        for col in all_cols:
            text = display_cell_text(ws.cell(row=header_row, column=col)).strip()
            if not text:
                continue
            options.append((text, col))
            found, total = scan(
                ws.cell(row=row, column=col).value
                for row in range(header_row + 1, last_data_row + 1)
            )
            written[col] = found
            totals[col] = total
        return "debit", options, written, totals
    finally:
        wb.close()


def format_amount_for_display(value: float) -> str:
    """千分位 + 两位小数，方便和账单逐位核对。"""
    try:
        return f"{float(value):,.2f}"
    except (TypeError, ValueError):
        return "-"


def format_period_progress_lines(sheet_name, options, written, totals,
                                 per_line: int = 3, max_lines: int = 4):
    """把"已写入的月份 + 各自金额"排成若干行紧凑文本。

    只列【已写入】的月份，未写入的只给一个数量。用户关心的就是"哪些月份
    做完了"，把 12 个未写入的圆圈也铺出来只会把真正有用的信息稀释掉，还会
    把状态栏撑得很宽。
    """
    if not options:
        return [f"{sheet_name}  (无可识别月份)"]

    done = [(label, totals.get(position, 0.0))
            for label, position in options if written.get(position)]
    pending = len(options) - len(done)

    head = f"{sheet_name}  ·  已写入 {len(done)}/{len(options)}"
    if pending:
        head += f"  ·  待写入 {pending}"
    lines = [head]

    if not done:
        lines.append("(暂无已写入月份)")
        return lines

    rows = [done[i:i + per_line] for i in range(0, len(done), per_line)]
    shown, hidden = rows[:max_lines], rows[max_lines:]
    for chunk in shown:
        lines.append("   ".join(
            f"{label:<9}{format_amount_for_display(amount):>13}"
            for label, amount in chunk
        ))
    if hidden:
        lines.append(f"…另有 {sum(len(c) for c in hidden)} 个月份已写入")
    return lines


def find_first_writable_sheet_name(path: Path, preferred_names: Optional[List[str]] = None) -> str:
    """Return the first sheet that has a recognized Credit or Debit period layout."""
    path = Path(path)
    if not path.exists():
        return ""
    # See read_period_options_from_selected_sheet: read_only=True + random cell
    # access is what caused the UI to freeze, so this also opens non-read-only
    # and reuses the sheet-type/layout caches.
    wb = load_workbook(path, read_only=False, data_only=False)
    try:
        ordered_names = list(preferred_names or []) + [n for n in wb.sheetnames if n not in (preferred_names or [])]
        for name in ordered_names:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]

            sheet_type = get_cached_sheet_type(path, name)
            if sheet_type is None:
                sheet_type = classify_sheet_type(ws)
                cache_sheet_type(path, name, sheet_type)

            if sheet_type == "credit":
                options = get_credit_period_options_from_column_a(ws)
            else:
                layout = get_cached_debit_layout(path, name)
                if layout is None:
                    layout = find_debit_layout(ws)
                    if layout is not None:
                        cache_debit_layout(path, name, layout)
                options = []
                if layout is not None:
                    header_row, _, month_cols, _, category_col = layout
                    all_cols = list(month_cols) + find_extra_debit_period_columns(ws, category_col)
                    options = [
                        c for c in all_cols
                        if display_cell_text(ws.cell(row=header_row, column=c)).strip()
                    ]
            if options:
                return name
        return ""
    finally:
        wb.close()


def resolve_period_position(
        ws, sheet_type: str, selected_label: str, selected_position: int,
        xlsx_path: Optional[Path] = None, sheet_name: str = ""
) -> int:
    """Resolve UI target. Debit uses cached numeric columns; Credit stays dynamic."""
    if sheet_type == "debit":
        layout = (
            get_cached_debit_layout(xlsx_path, sheet_name)
            if xlsx_path is not None and sheet_name else None
        )
        if layout is None:
            layout = find_debit_layout(ws)
            if layout is not None and xlsx_path is not None and sheet_name:
                cache_debit_layout(xlsx_path, sheet_name, layout)
        if layout is None:
            raise ValueError(f"Excel Sheet“{ws.title}”没有可用的 Debit 固定列映射。")
        header_row, _, month_cols, _, category_col = layout
        valid_cols = set(month_cols) | set(find_extra_debit_period_columns(ws, category_col))
        if selected_position not in valid_cols:
            raise ValueError("当前 Month 对应的 Debit 固定列不存在，请重新选择 Sheet。")
        # Column number is authoritative. Label is UI display only.
        return selected_position

    # Credit columns can move as merchants are inserted, but Month rows in A are
    # stable. Re-read Column A and recover by label if necessary.
    options = get_credit_period_options_from_column_a(ws)
    if len(options) <= 5:
        raise ValueError(f"Excel Sheet“{ws.title}”已不符合 Credit 的 Column A 规则。")
    for label, position in options:
        if position == selected_position and label == selected_label:
            return position
    matches = [position for label, position in options if label == selected_label]
    if len(matches) == 1:
        return matches[0]
    raise ValueError(
        f"日期/期间“{selected_label}”与当前 Excel 位置不一致。请重新选择 Month。"
    )


def append_amount_to_cell(cell, amounts: List[str]):
    parts = split_formula_parts(cell.value)
    parts.extend(normalize_amount_string(a) for a in amounts)
    formula = build_plus_formula(parts)
    cell.value = formula if formula else None
    cell.number_format = "0.00"


def copy_row_style(ws, source_row: int, target_row: int, max_col: int):
    """把一整行的格式复制到另一行。

    只复制 ``_style`` 一个属性就够了。它是 openpyxl 内部的样式索引数组，
    字体、边框、填充、对齐、保护、数字格式全都由它索引，复制它等于复制
    了全部格式（已实测验证）。

    早期实现在复制 ``_style`` 之后，又逐个复制了 alignment / border /
    fill / font / protection / number_format。那 5 次额外复制完全是多余的，
    却是最贵的部分：每次都要把样式代理对象解析成真实对象、深拷贝一遍，
    再重新登记进工作簿的样式索引表。实测中它占掉整个写入耗时的七成以上。
    """
    copy_row_style_to_rows(ws, source_row, [target_row], max_col)


def copy_row_style_to_rows(ws, source_row: int, target_rows, max_col: int):
    """把同一行的格式批量复制到多行。

    新增多个商户时，所有新行的格式都来自同一行模板，因此源行只需要读取
    一次，之后复用。这样避免了"每插入一行就把源行重新扫一遍"的重复开销。
    """
    from copy import copy

    target_rows = [r for r in (target_rows or []) if r and r >= 1]
    if source_row < 1 or not target_rows:
        return

    # 源行样式只读一次。没有样式的列记 None，跳过即可，避免给新行凭空
    # 创建默认样式对象。
    source_styles = []
    for col in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        source_styles.append(src._style if src.has_style else None)

    source_height = ws.row_dimensions[source_row].height

    for target_row in target_rows:
        for col, style in enumerate(source_styles, start=1):
            if style is None:
                continue
            ws.cell(row=target_row, column=col)._style = copy(style)
        ws.row_dimensions[target_row].height = source_height


def copy_column_style(ws, source_col: int, target_col: int, max_row: int):
    from copy import copy
    if source_col < 1:
        return
    source_letter = excel_col_letter(source_col)
    target_letter = excel_col_letter(target_col)
    ws.column_dimensions[target_letter].width = ws.column_dimensions[source_letter].width
    for row in range(1, max_row + 1):
        src = ws.cell(row=row, column=source_col)
        dst = ws.cell(row=row, column=target_col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.alignment = copy(src.alignment)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.font = copy(src.font)
        dst.protection = copy(src.protection)
        dst.number_format = src.number_format


def locate_credit_columns(ws):
    """Return Credit control-column positions using numeric structure only.

    This function NEVER decides sheet type. Credit/Debit type is already fixed by
    the single Column-A rule.  For Credit, positions follow one stable structure:

        A = Month
        B = Beginning Balance (fixed)
        C.. = dynamic merchant/income columns
        rightmost 3 non-empty header cells = Total Credit, Debit, Ending Balance

    Merchant columns may grow, so only the three right-edge control columns are
    recalculated. No old header-text validation is used here.
    """
    credit_options = get_credit_period_options_from_column_a(ws)
    if len(credit_options) <= 5:
        return None, {}

    first_period_row = min(row for _, row in credit_options)
    header_row = first_period_row - 1
    if header_row < 1:
        return None, {}

    begin_col = CREDIT_BEGIN_COL
    col_limit = min(ws.max_column, 500)

    # Dynamic Credit additions happen before the three control columns.  The
    # rightmost three non-empty cells on the header row therefore remain the
    # numeric anchors for Total Credit / Debit / Ending Balance.
    used_header_cols = []
    for col in range(begin_col, col_limit + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is not None and str(value).strip() != "":
            used_header_cols.append(col)

    if len(used_header_cols) < 4:
        return None, {}

    total_credit_col, debit_col, ending_col = used_header_cols[-3:]
    if not (begin_col < total_credit_col < debit_col < ending_col):
        return None, {}

    # Keep raw merchant names for existing-column lookup, then overlay canonical
    # control keys at their numeric positions.
    header_map = {}
    for col in used_header_cols:
        key = normalized_header_key(ws.cell(row=header_row, column=col).value)
        if key:
            header_map[key] = col

    header_map[CREDIT_BEGIN_HEADER] = begin_col
    header_map[CREDIT_TOTAL_HEADER] = total_credit_col
    header_map[CREDIT_DEBIT_HEADER] = debit_col
    header_map[CREDIT_ENDING_HEADER] = ending_col
    return header_row, header_map


def find_debit_layout(ws):
    """Return the fixed Debit numeric layout using ONLY Row 1.

    Strict template contract:
      - Column A is always Merchant.
      - Debit period headers are always on Row 1.
      - Read Row 1 only, starting from Column B.
      - Recognizable Month/date cells map directly to their Excel column numbers.
      - No Merchant/Total/Category keyword search and no multi-row scanning.
    """
    header_row = 1
    merchant_col = DEBIT_MERCHANT_COL
    col_limit = min(ws.max_column, DEBIT_SCAN_MAX_COLS)
    if col_limit < DEBIT_FIRST_MONTH_COL:
        return None

    month_cols = []
    for col in range(DEBIT_FIRST_MONTH_COL, col_limit + 1):
        if looks_like_period_header(ws.cell(row=1, column=col)):
            month_cols.append(col)

    if len(month_cols) < DEBIT_MIN_PERIOD_HEADERS:
        return None

    period_set = set(month_cols)
    total_col = None
    category_col = None
    seen_periods = 0
    col = DEBIT_FIRST_MONTH_COL
    while col <= col_limit:
        if col in period_set:
            seen_periods += 1
            col += 1
            continue
        if seen_periods >= DEBIT_MIN_PERIOD_HEADERS and col + 1 <= col_limit:
            if col not in period_set and (col + 1) not in period_set:
                total_col = col
                category_col = col + 1
                break
        col += 1

    if total_col is None or category_col is None:
        return None

    # Only the date/period columns BEFORE Total/Category count as real Debit
    # data columns for AGGREGATION purposes. Some templates have a stray
    # recognizable date-like column AFTER Category (e.g. a pre-created
    # "Apr-26" column for a future month) -- looks_like_period_header()
    # matches it, so the initial scan above picks it up too. Without this
    # filter, that trailing column silently gets pulled into every SUM/Total
    # formula (per-merchant Total, the bottom TOTAL row, and the
    # merge-duplicate-merchants feature). Trimming month_cols here fixes that
    # everywhere at once, since every aggregation caller (write's Total
    # formula, merge, clear-amounts, category learning) goes through this
    # single function. A column after Category should still be selectable in
    # the UI as a write target -- that is handled separately by
    # find_extra_debit_period_columns() below, used only by the UI-facing
    # month-option builders, never by anything that builds a SUM formula.
    month_cols = [c for c in month_cols if c < total_col]
    if len(month_cols) < DEBIT_MIN_PERIOD_HEADERS:
        return None

    return header_row, merchant_col, month_cols, total_col, category_col


def find_extra_debit_period_columns(ws, category_col: int, col_limit: Optional[int] = None) -> List[int]:
    """Find additional recognizable period columns positioned AFTER Category.

    Some templates append a new month column after Total/Category (e.g. a
    freshly added "Apr-26" column for next year) instead of inserting it
    within the original month block, so Total/Category stay in place. Such a
    column should still be selectable in the UI as a write target, but per
    program rule it must NEVER be pulled into any Total/SUM aggregation --
    Total only ever sums the block strictly before Total/Category (see
    find_debit_layout). This is intentionally decided purely by column
    position, never by header wording, since header text varies across
    templates.
    """
    if col_limit is None:
        col_limit = min(ws.max_column, DEBIT_SCAN_MAX_COLS)
    extra_cols = []
    for col in range(category_col + 1, col_limit + 1):
        if looks_like_period_header(ws.cell(row=1, column=col)):
            extra_cols.append(col)
    return extra_cols


def safe_set_merged_value(ws, row: int, col: int, value):
    """Write a value to a cell, redirecting to the top-left cell if merged.

    openpyxl raises ``AttributeError`` when writing to a non-top-left cell of a
    merged range. Bank-name/header cells on imported templates are sometimes
    merged, so this resolves the correct writable cell before assigning.
    """
    target_coord = ws.cell(row=row, column=col).coordinate
    for merged_range in ws.merged_cells.ranges:
        if target_coord in merged_range:
            top_left_coord = merged_range.coord.split(":")[0]
            ws[top_left_coord] = value
            return
    ws.cell(row=row, column=col, value=value)


def rebuild_credit_total_row_formulas(ws):
    """Rebuild every Credit TOTAL formula after dynamic column insertion.

    Credit has one deliberately simple structural contract:
      - Column A contains the writable Month/date rows.
      - The TOTAL label is also in Column A.
      - Column B is BEGIN BALANCE and is not changed here.
      - Row immediately above the first Month/date row is the header row.
      - TOTAL CREDIT is always exactly 2 columns before ENDING BALANCE (the
        rightmost populated header), with DEBIT directly in between. Each
        period row's own Total Credit cell is rebuilt to SUM every dynamic
        column for that row (column 3 through the column right before Total
        Credit) -- this must be redone every time rather than trusted from
        whatever was written before, since inserting a new merchant column
        only shifts existing cells positionally and never rewrites another
        cell's formula text, so a stale Total Credit formula would otherwise
        keep summing only its original narrow range forever.
      - Every populated header column strictly between Begin (column B) and
        Ending (the rightmost populated header) also receives a normal
        vertical SUM formula in the bottom TOTAL row, over the full
        Month/date row range. Both Begin and Ending themselves are left
        completely untouched in the TOTAL row -- the user fills these in
        manually. All boundaries are found purely by position, never by
        matching header text such as "Begin"/"Ending"/"Total Credit", since
        real templates word these differently.

    ``openpyxl.insert_cols`` moves cells but does not reliably translate every
    existing formula reference. Recreating this row after all Credit writes
    makes the result independent of how many new merchant columns were inserted.
    """
    credit_periods = get_credit_period_options_from_column_a(ws)
    period_rows = sorted({row for _, row in credit_periods})
    if not period_rows:
        raise ValueError("Credit Sheet 的 Column A 中找不到可用于 TOTAL 的 Month/日期行。")

    first_period_row = period_rows[0]
    last_period_row = period_rows[-1]
    header_row = first_period_row - 1
    if header_row < 1:
        raise ValueError("Credit Sheet 的 Month/日期行上方找不到标题行。")

    # TOTAL is located only from Column A. Prefer a TOTAL below the final
    # Month/date row, while still accepting an existing one elsewhere.
    total_row = None
    fallback_total_row = None
    for row in range(1, ws.max_row + 1):
        label = normalized_cell_text(ws.cell(row=row, column=1).value).upper()
        if label in {"TOTAL", "GRAND TOTAL"}:
            if fallback_total_row is None:
                fallback_total_row = row
            if row > last_period_row:
                total_row = row
                break

    if total_row is None:
        total_row = fallback_total_row
    if total_row is None:
        total_row = last_period_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL")

    # Do not trust ws.max_column: imported templates may contain formatting far
    # beyond the actual table. The last non-empty header is the real last column.
    last_header_col = None
    for col in range(ws.max_column, 2, -1):
        value = ws.cell(row=header_row, column=col).value
        if value is not None and str(value).strip() != "":
            last_header_col = col
            break

    if last_header_col is None:
        return

    # Position-based rule, never text-based (header wording for Begin/Ending
    # varies by template -- e.g. "Begin" vs "BEGIN BALANCE"): Column B is
    # always BEGIN and Ending (the rightmost populated header, whatever it is
    # named) are BOTH left untouched by the TOTAL row -- the user fills these
    # in manually. Only every column strictly between Begin and Ending (the
    # dynamic income columns, TOTAL CREDIT, DEBIT) gets a normal vertical SUM.
    ending_col = last_header_col

    # Total Credit is always exactly 2 columns before Ending (Debit sits
    # directly in between: ... TOTAL CREDIT, DEBIT, ENDING BALANCE), the same
    # fixed header order this program always writes -- so this holds
    # regardless of header wording.
    total_credit_col = ending_col - 2

    # Each period row's OWN "Total Credit" formula (e.g. row 3's =SUM(C3:E3))
    # must also be rebuilt every time, not left as whatever static text
    # happened to be written once. Inserting a new merchant column
    # (ws.insert_cols) only shifts cells positionally -- it never rewrites
    # formula TEXT elsewhere -- so a per-row Total Credit formula silently
    # keeps summing only its original narrow range forever, excluding every
    # merchant column added afterward. Rebuilding it here from the current
    # first-dynamic-column through the column right before Total Credit fixes
    # that for every period row, every time.
    if total_credit_col > 3:
        first_dynamic_col = 3
        last_dynamic_col = total_credit_col - 1
        for row in period_rows:
            start_ref = ws.cell(row=row, column=first_dynamic_col).coordinate
            end_ref = ws.cell(row=row, column=last_dynamic_col).coordinate
            total_credit_cell = ws.cell(row=row, column=total_credit_col)
            total_credit_cell.value = f"=SUM({start_ref}:{end_ref})"
            total_credit_cell.number_format = "0.00"

    for col in range(3, ending_col):
        col_letter = excel_col_letter(col)
        total_cell = ws.cell(row=total_row, column=col)
        total_cell.value = (
            f"=SUM({col_letter}{first_period_row}:"
            f"{col_letter}{last_period_row})"
        )
        total_cell.number_format = "0.00"

    # Ending is intentionally left untouched here -- the user fills the TOTAL
    # row's Ending cell in manually.


def update_credit_sheet_in_place(ws, rows, selected_period_row: int, bank_name: str = DEFAULT_BANK_NAME):
    """Very simple Credit writer.

    Fixed Credit contract:
      - Column A contains Month values vertically.
      - selected_period_row is the row already mapped from Column A by the UI.
      - Column B is BEGIN BALANCE and stays fixed.
      - For every new Credit merchant, insert a new column BETWEEN B and current C
        (that is, insert at Excel column 3), then write the merchant header and amount.
      - No Credit header/layout validation is used.
    """
    month_col = 1  # A
    begin_col = 2  # B
    insert_col = 3  # always insert between B and C

    if not isinstance(selected_period_row, int) or selected_period_row < 1 or selected_period_row > ws.max_row:
        raise ValueError("Credit Month 对应的 Row 无效。")

    # The row came from Column-A Month mapping. Only verify that A[row] still has
    # a recognizable Month/date; do not inspect any Credit control headers.
    month_cell = ws.cell(row=selected_period_row, column=month_col)
    if not looks_like_period_header(month_cell):
        raise ValueError(
            f"Credit Month Row {selected_period_row} 的 Column A 已不是可识别的 Month/日期，请重新选择 Month。"
        )

    # Optional bank name: only write to the top-left writable cell of Row 1,
    # Column B (the same convention used everywhere else in this program,
    # e.g. create_credit_template_sheet).
    if bank_name:
        try:
            safe_set_merged_value(ws, 1, begin_col, bank_name)
        except Exception:
            pass

    # Aggregate duplicate merchants from this import first.
    merchant_amounts = {}
    merchant_order = []
    # Parsed transaction rows in this program are normally 3-tuples:
    # (merchant, amount, who/source). Credit only needs the first two values.
    for item in rows:
        if not item or len(item) < 2:
            continue
        merchant, amount = item[0], item[1]
        merchant = str(merchant or "").strip()
        if not merchant:
            continue
        try:
            amount = float(amount)
        except Exception:
            continue
        key = normalize_merchant_key(merchant)
        if not key:
            continue
        if key not in merchant_amounts:
            merchant_order.append((key, merchant))
            merchant_amounts[key] = 0.0
        merchant_amounts[key] += amount

    # Existing merchant headers are on Row 2. Reuse them when possible.
    existing = {}
    for col in range(3, ws.max_column + 1):
        value = ws.cell(row=2, column=col).value
        key = normalize_merchant_key(str(value or ""))
        if key:
            existing.setdefault(key, col)

    for key, merchant in merchant_order:
        amount = merchant_amounts[key]

        if key in existing:
            target_col = existing[key]
        else:
            # The user's fixed rule: every new merchant is inserted at C,
            # directly between BEGIN BALANCE (B) and the old Column C.
            ws.insert_cols(insert_col, 1)
            target_col = insert_col
            ws.cell(row=2, column=target_col, value=merchant)

            # Existing column numbers shifted right by one after insertion.
            existing = {k: (c + 1 if c >= insert_col else c) for k, c in existing.items()}
            existing[key] = target_col

        cell = ws.cell(row=selected_period_row, column=target_col)
        current = cell.value
        if isinstance(current, (int, float)):
            cell.value = float(current) + amount
        elif current in (None, ""):
            cell.value = amount
        else:
            # Do not try to interpret complex formulas/text in a merchant amount cell.
            # Replace it with this imported amount for the selected Month row.
            cell.value = amount

    # openpyxl moves existing cells during insert_cols, but it does not update
    # all affected formulas. Always rebuild the entire Credit TOTAL row once,
    # after every merchant/amount write has finished.
    rebuild_credit_total_row_formulas(ws)

    # Keep the fixed B-column balance chain simple. No header lookup.
    return


def update_debit_sheet_in_place(ws, rows, selected_period_col: int, layout=None):
    # UI/writer passes the cached numeric Debit layout.  Only fall back to one
    # scan when this function is called independently.
    if layout is None:
        layout = find_debit_layout(ws)
    if layout is None:
        raise ValueError(f"Excel Sheet“{ws.title}”不是可识别的 Debit 格式。")
    header_row, merchant_col, month_cols, total_col, category_col = layout
    extra_cols = find_extra_debit_period_columns(ws, category_col)
    if selected_period_col not in month_cols and selected_period_col not in extra_cols:
        raise ValueError(
            f"Excel Sheet“{ws.title}”中的目标日期列已变化，请重新选择 Month 后再试。"
        )
    selected_col = selected_period_col
    rules = load_category_rules()
    category_history = collect_category_history_from_workbook(ws.parent)
    merged = merge_same_merchants(rows)
    total_row = None
    merchant_rows = {}
    for r in range(header_row + 1, ws.max_row + 1):
        name = normalized_cell_text(ws.cell(r, merchant_col).value)
        if not name:
            continue
        if name.upper() in {"TOTAL", "GRAND TOTAL"}:
            total_row = r
            break
        merchant_rows[name.casefold()] = r

        # Fill only blank categories; never overwrite a user's existing choice.
        if category_col is not None:
            category_cell = ws.cell(r, category_col)
            existing_category = normalized_cell_text(category_cell.value)
            if not existing_category:
                learned_category, _, _ = get_smart_category_for_merchant(
                    name, rules, category_history
                )
                category_cell.value = learned_category
                if learned_category.upper() != CATEGORY_REVIEW_LABEL:
                    category_history[normalize_merchant_for_category(name)] = learned_category
    if total_row is None:
        # No existing Total row to preserve a reference to -- nothing else in
        # the workbook could already be pointing at it, so no repair is needed.
        old_total_row = None
        total_row = ws.max_row + 1
        ws.cell(total_row, merchant_col, "Total")
    else:
        old_total_row = total_row
    # 先把这一批里"表中还没有的商户"全部找出来，再一次性插入。
    #
    # 早期实现是发现一个新商户就 insert_rows 一行。openpyxl 的每次插入都要
    # 把插入点以下的所有单元格整体搬动一遍，插 N 个商户就搬 N 次；而且样式
    # 模板行也被反复重新扫描。改成"算好数量、插一次"之后，搬动只发生一次，
    # 样式模板也只读一次。
    #
    # 这里必须【在批内也按不区分大小写去重】。merge_same_merchants 是按原始
    # 字符串分组的，所以同一批里 "AMAZON" / "Amazon" / "amazon" 会是三个条目。
    # 逐个插入的老写法天然不会出问题：第一个插完就登记进 merchant_rows，后两
    # 个查表即可命中同一行。但批量写法是在插入【之前】一次性统计的，那时三个
    # 变体都还没登记，如果不去重就会各占一行，前两行还是没有金额的空行。
    new_merchants = []
    seen_new_keys = set()
    for merchant in merged:
        key = merchant.casefold()
        if merchant_rows.get(key) is not None or key in seen_new_keys:
            continue
        seen_new_keys.add(key)
        new_merchants.append(merchant)

    if new_merchants:
        # 样式模板行必须在插入之前确定：插入发生在 total_row 处，
        # total_row-1 这一行不会移动，插入前后指向同一行。
        style_source_row = max(header_row + 1, total_row - 1)
        style_max_col = ws.max_column

        ws.insert_rows(total_row, len(new_merchants))
        inserted_rows = list(range(total_row, total_row + len(new_merchants)))
        copy_row_style_to_rows(ws, style_source_row, inserted_rows, style_max_col)

        for merchant, r in zip(new_merchants, inserted_rows):
            ws.cell(r, merchant_col, merchant)
            for c in month_cols:
                ws.cell(r, c, None)
            if category_col is not None:
                learned_category, _, _ = get_smart_category_for_merchant(
                    merchant, rules, category_history
                )
                ws.cell(r, category_col, learned_category)
                if learned_category.upper() != CATEGORY_REVIEW_LABEL:
                    category_history[normalize_merchant_for_category(merchant)] = learned_category
            merchant_rows[merchant.casefold()] = r

        total_row += len(new_merchants)

    for merchant, data in merged.items():
        r = merchant_rows[merchant.casefold()]
        append_amount_to_cell(ws.cell(r, selected_col), data["amounts"])
        ws.cell(r, total_col, build_sum_formula_for_columns(ws, r, month_cols))
    ws.cell(total_row, merchant_col, "Total")
    first_data_row, last_data_row = header_row + 1, total_row - 1
    if last_data_row >= first_data_row:
        for c in month_cols:
            letter = excel_col_letter(c)
            ws.cell(total_row, c, f"=SUM({letter}{first_data_row}:{letter}{last_data_row})")
        ws.cell(total_row, total_col,
                build_sum_formula_for_columns(ws, total_row, month_cols))

    # Inserting a brand-new merchant row pushes the Total row down. Repair any
    # cross-sheet formula elsewhere in the workbook (e.g. a paired Credit
    # sheet's DEBIT column imported from an existing report) that still
    # references the old Total row position.
    if old_total_row is not None and old_total_row != total_row:
        # 新增商户是在旧 Total 行的位置插入的，因此"旧 Total 行及其下方"的
        # 所有内容整体下移了 offset 行，指向它们的公式必须同步平移。
        offset = total_row - old_total_row
        shift_cross_sheet_row_references(ws.parent, ws.title, old_total_row, offset)
        # Some Debit templates also have a summary row directly below Total
        # -- commonly labeled "Add" / "ADD" (or Begin/Less/End) -- whose
        # formula reads Total's own cells directly (e.g. a plain "=B14").
        # That is a SAME-SHEET reference, so it must be repaired separately
        # from the cross-sheet fix above; otherwise it keeps pointing at the
        # row Total used to occupy before this new merchant row pushed it
        # down, and silently shows stale or blank numbers.
        #
        # min_row 必不可少：只能重写 Total 行【下方】的汇总块。上方的商户行
        # 和 Total 行自己的公式，程序刚刚按当前行号重新写好，再平移一次就
        # 全错了。
        shift_local_row_references(
            ws, old_total_row, offset, min_row=total_row + 1
        )


def append_rows_to_selected_sheet(
        xlsx_path: Path, sheet_name: str, rows,
        selected_period_label: str, selected_period_position: int,
        expected_sheet_type: str = "", bank_name: str = DEFAULT_BANK_NAME
):
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"找不到 Excel 文件：{xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Excel 中找不到Excel Sheet：{sheet_name}")
        ws = wb[sheet_name]

        # Safety check: never blindly trust the UI-cached type. Re-classify this
        # exact worksheet right before writing. If the UI's remembered type no
        # longer matches, stop and raise instead of silently writing a Debit
        # sheet through the Credit path (or vice versa) -- that mismatch is what
        # produced the transposed/garbled Credit-shaped output on a Debit sheet.
        fresh_type = classify_sheet_type(ws)
        if expected_sheet_type in ("credit", "debit") and expected_sheet_type != fresh_type:
            raise ValueError(
                f"Excel Sheet“{sheet_name}”的识别结果发生变化"
                f"(界面记录为 {expected_sheet_type}，重新检测为 {fresh_type})。\n"
                "请重新在下拉框中选择该 Sheet 后再点击 Start，避免写错位置。"
            )
        detected_type = fresh_type
        cache_sheet_type(xlsx_path, sheet_name, detected_type)

        debit_layout = None
        if detected_type == "debit":
            debit_layout = get_cached_debit_layout(xlsx_path, sheet_name)
            if debit_layout is None:
                debit_layout = find_debit_layout(ws)
                if debit_layout is None:
                    raise ValueError(f"Excel Sheet“{sheet_name}”没有可用的 Debit 固定列映射。")
                cache_debit_layout(xlsx_path, sheet_name, debit_layout)

        real_position = resolve_period_position(
            ws, detected_type, selected_period_label, selected_period_position,
            xlsx_path=xlsx_path, sheet_name=sheet_name
        )

        if detected_type == "credit":
            update_credit_sheet_in_place(ws, rows, real_position, bank_name)
        else:
            update_debit_sheet_in_place(ws, rows, real_position, layout=debit_layout)
            # Adding a new Debit Merchant can move the Debit Total row downward.
            # Rebuild Credit Summary's DEBIT references so Jan-Dec always point
            # to the current monthly Total cells rather than a stale row number.
            sync_credit_debit_from_debit_sheet(wb)
        wb.save(xlsx_path)
        return detected_type
    finally:
        wb.close()


# ================= 输出路径 =================

def get_summary_output_path(base_path_str: str) -> Path:
    """Return the exact selected .xlsx path, or the default file in a folder."""
    raw = (base_path_str or "").strip()

    if raw:
        p = Path(raw).expanduser()
        if p.suffix.lower() == ".xlsx":
            p.parent.mkdir(parents=True, exist_ok=True)
            return p
        folder = p
    else:
        folder = script_dir()

    folder.mkdir(parents=True, exist_ok=True)
    return folder / "credit_monthly_summary.xlsx"


# ================= 当前工作表重复 Merchant 合并 =================

# ================= Debit 商户名清洗 =================

# 交易码/流水号在账单里常见的两种形态：紧跟在品牌后面的长串数字
# （Shell247592917239），以及星号后面的一段随机码（Amazon.Com*Nx6）。
# 这些内容对记账毫无意义，却会让同一个商户被拆成很多行。

# 连续 3 位及以上的数字才算流水号。阈值定在 3 是有意的：账单里像
# "No1 Sushi 88"、"Chase #156" 这种 1~2 位数字是商户名的一部分，
# 一旦把它们也删掉，正常名字就被破坏了。
MERCHANT_LONG_DIGITS_RE = re.compile(r"\d{3,}")

# 域名后缀只在"后面不再跟字母"时才删，例如 amazon.com -> amazon。
# 而 Amazon.Comseattlewa 这种 .Com 后面直接粘着字母的，说明它已经和
# 后面的词连在一起了，贸然删除会得到 Amazonseattlewa，反而更难认，
# 因此保持原样。
MERCHANT_DOMAIN_SUFFIX_RE = re.compile(
    r"(?:\.|/)(?:com|net|org|info|biz)(?![A-Za-z0-9])", re.IGNORECASE
)

# 独立的日期/编号碎片，例如 Expida143242//4/24/com 里的 4/24。
MERCHANT_DATE_FRAGMENT_RE = re.compile(
    r"(?<![A-Za-z0-9])\d{1,4}[/\-.]\d{1,4}(?:[/\-.]\d{1,4})?(?![A-Za-z0-9])"
)

# 字母数字混排的参考码，例如 Bill Pay 后面的 6Bu1T6Kr、4Bq1P6Kr。
#
# 判定看的是"字母段和数字段交替了多少次"，而不是简单的"含字母又含数字"。
# 这一点很关键，两类字符串必须区分开：
#
#   Shell247592917239  字母段 + 数字段，只有 2 段 —— 这是品牌加流水号，
#                      应该删掉数字保留 Shell；
#   6Bu1T6Kr           6|Bu|1|T|6|Kr 交替 6 段 —— 这才是随机参考码，整词丢掉。
#
# 门槛定在 4 段，同时要求长度不少于 6、至少含 2 个数字，这样 7Eleven（2 段）、
# WD40（2 段）、3M（太短）这些真实品牌名都不会被误伤。
MERCHANT_CODE_MIN_LENGTH = 6
MERCHANT_CODE_MIN_DIGITS = 2
MERCHANT_CODE_MIN_RUNS = 4


def _is_reference_code_token(token: str) -> bool:
    """判断一个词是不是字母数字随机交替的参考码/交易码。"""
    if len(token) < MERCHANT_CODE_MIN_LENGTH or not token.isalnum():
        return False
    if sum(ch.isdigit() for ch in token) < MERCHANT_CODE_MIN_DIGITS:
        return False

    runs = 1
    for previous, current in zip(token, token[1:]):
        if previous.isdigit() != current.isdigit():
            runs += 1
    return runs >= MERCHANT_CODE_MIN_RUNS


# 纯分隔符（斜杠、星号、井号等）没有含义，可以整词丢掉。
# 注意这里刻意不包含 &：Orange & Rockland 里的 & 是名字的一部分。
MERCHANT_SEPARATOR_TOKEN_RE = re.compile(r"^[\s\-–—/\\*#|,;:.·]+$")


def _is_pure_number_token(token: str) -> bool:
    """判断一个词是不是"没有字母、且数字够长"的流水号。

    这样 200045、0964、1-518-457-5434 会被丢掉，而 No1 Sushi 里的 88
    因为只有两位数字，会被保留下来 —— 它是店名的一部分。
    """
    if any(ch.isalpha() for ch in token):
        return False
    return sum(ch.isdigit() for ch in token) >= 3


# 这些词在 Debit 表里有结构含义：Total 标记明细区的结尾，BEGIN/ADD/LESS/
# END 是 Total 下方的汇总行。清洗后的商户名一旦变成它们中的任何一个，
# 整张表的结构就被破坏了 —— 例如 "Total 12345" 被清成 "Total" 后，程序会
# 把这一行当成汇总行，后续写入的商户会插到它上面，真正的商户数据反而被
# 挤到假 Total 行下面，Total 公式也只统计到假行为止。
# 遇到这种情况一律放弃清洗、保留原名：名字难看远好过表结构损坏。
MERCHANT_STRUCTURAL_LABELS = {
    "TOTAL", "GRAND TOTAL", "SUBTOTAL", "SUB TOTAL",
    "BEGIN", "BEGINNING", "ADD", "LESS", "END", "ENDING",
}


# ---------- 银行转账 / 付款类商户：整条跳过清洗 ----------
#
# 这类商户名里的数字往往正是判断分类所需的关键信息，例如
# "Online Payment To Chase4568" 里的 4568 是对方账户尾号 —— 一旦被当成
# 流水号删掉，就再也分不清这笔钱是还哪张卡、转到哪个账户了。
#
# 因此只要名字里出现银行名或转账/付款类关键词，就完整保留、不做任何清洗。
# 这里刻意把范围放宽：误保护的代价只是名字没被清干净，而漏保护会直接
# 导致无法分类，两者代价不对等。
#
# 银行简称必须用词边界匹配。boa 如果按子串匹配会命中 boat、boarding，
# td 会命中 ltd、std，那样就会把无关商户也保护起来。
MERCHANT_BANK_KEYWORDS = [
    # 具体银行 / 金融机构
    r"chase", r"jp\s*morgan", r"jpmorgan",
    r"boa", r"bofa", r"bank\s*of\s*america",
    r"wells\s*fargo", r"wellsfargo",
    r"citi", r"citibank", r"citicard",
    r"capital\s*one", r"capitalone",
    r"amex", r"american\s*express",
    r"discover", r"us\s*bank", r"usbank",
    r"pnc", r"td\s*bank", r"tdbank",
    r"truist", r"suntrust", r"bb&t",
    r"hsbc", r"santander", r"ally", r"barclays",
    r"schwab", r"fidelity", r"vanguard",
    r"navy\s*federal", r"navyfederal", r"usaa",
    r"sofi", r"huntington", r"regions",
    r"key\s*bank", r"keybank", r"m&t", r"citizens",
    r"fifth\s*third", r"bmo", r"synchrony",
    r"credit\s*one", r"creditone", r"first\s*republic",
    r"goldman", r"marcus", r"chime", r"varo", r"axos",
    r"venmo", r"paypal", r"zelle", r"cash\s*app", r"cashapp",
    r"western\s*union", r"moneygram",
    # 通用金融机构词
    r"bank", r"banco", r"credit\s*union", r"fcu", r"bancorp",
]

MERCHANT_TRANSFER_KEYWORDS = [
    r"transfers?", r"trf", r"xfer", r"wire",
    r"online\s*payments?", r"onlinepayments?", r"online\s*pmt",
    r"payments?\s*to", r"pmt\s*to", r"epayments?", r"e-payments?",
    r"card\s*payments?", r"cardpayments?", r"cc\s*pmt",
    r"autopay", r"auto\s*pay", r"ach",
    r"direct\s*dep", r"deposits?", r"withdrawals?", r"withdraw",
    r"atm", r"overdraft", r"mobile\s*payments?", r"web\s*pmt",
]

# 边界处理是这条规则的关键，两头要求不同：
#
#   前面不能是字母 —— 否则 Ltd 里的 td、Standard 里的片段都会误命中；
#   后面同样不能是字母 —— 否则 Boat / Boarding 里的 boa、Chaser 里的
#   chase 会被误判成银行。
#
# 但后面【可以是数字】，这一点必须留出来：真实账单里银行名常常和账号
# 尾号直接粘在一起，例如 "Chase4568"。如果按普通词边界 \b 来匹配，
# Chase 和 4568 之间没有边界，整条就保护不住，4568 会被当成流水号删掉 ——
# 而那正是判断分类最需要的信息。
MERCHANT_PROTECTED_RE = re.compile(
    r"(?<![A-Za-z])(?:"
    + "|".join(MERCHANT_BANK_KEYWORDS + MERCHANT_TRANSFER_KEYWORDS)
    + r")(?![A-Za-z])",
    re.IGNORECASE,
)


def is_bank_transfer_merchant(name: str) -> bool:
    """判断商户名是否属于银行 / 转账 / 付款类，需要完整保留。"""
    return bool(MERCHANT_PROTECTED_RE.search(str(name or "")))


# ================= Merchant Alias Rules (XLSX) =================
#
# 有些商户名怎么整理，本质上是人的判断，规则推不出来。例如同样是
# "城市 + 州"结尾：
#
#     POS Mac Amazon Seattle WA   希望变成  Amazon
#     Kindercare Portland OR      希望保持  Kindercare Portland OR
#
# 一个要删一个要留，正则无从判断。类似的还有 Veolia Water New Y 想留成
# Veolia Water、Amazon Prime 要和 Amazon 分开统计。
#
# 因此把这类判断交给一张用户可以自己编辑的对照表：命中关键字就整条替换成
# 指定的标准名。表格放在程序目录的 "Merchant Rules" 文件夹里，第一次运行
# 会自动生成一份示例，之后随时可以增删改，不需要动代码。

MERCHANT_ALIAS_CACHE = {"path": None, "mtime": None, "rules": []}


def merchant_rules_folder() -> Path:
    folder = script_dir() / "Merchant Rules"
    folder.mkdir(parents=True, exist_ok=True)
    return folder


def merchant_rules_path() -> Path:
    return merchant_rules_folder() / "merchant_rules.xlsx"


def create_default_merchant_rules_xlsx(path: Path):
    """生成示例别名表。

    预置的几条来自真实账单里最常见的情况，可以直接看出用法：左边填一段
    出现在原始商户名里的关键字（不区分大小写），右边填希望显示的标准名。
    """
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Merchant Rules"
    ws.cell(row=1, column=1, value="keyword")
    ws.cell(row=1, column=2, value="merchant")

    sample_rows = [
        # 关键字越具体的越优先，所以 Amazon Prime 不会被 Amazon 抢走。
        ("amazon prime", "Amazon Prime"),
        ("amazon", "Amazon"),
        ("veolia", "Veolia Water"),
        ("no1 sushi", "No1 Sushi 88"),
        ("withdrawal branch", "Withdrawal Branch"),
        ("amex epayment", "Amex Epayment"),
    ]
    for index, (keyword, merchant) in enumerate(sample_rows, start=2):
        ws.cell(row=index, column=1, value=keyword)
        ws.cell(row=index, column=2, value=merchant)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 28
    wb.save(path)


def load_merchant_alias_rules() -> List[Tuple[str, str]]:
    """读取别名表，按关键字长度倒序返回，保证更具体的规则先匹配。

    结果会按文件修改时间缓存，避免每次清洗都重新读盘。
    """
    path = merchant_rules_path()
    if not path.exists():
        create_default_merchant_rules_xlsx(path)

    try:
        mtime = path.stat().st_mtime
    except OSError:
        mtime = None

    if (MERCHANT_ALIAS_CACHE["path"] == str(path)
            and MERCHANT_ALIAS_CACHE["mtime"] == mtime):
        return MERCHANT_ALIAS_CACHE["rules"]

    rules: List[Tuple[str, str]] = []
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            keyword = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
            merchant = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            if not keyword or not merchant:
                continue

            # 单字符关键字会命中几乎所有商户，把整列合并成一行，几乎肯定
            # 是手滑写错的，直接忽略。
            if len(keyword) < 2:
                continue

            # 标准名不能是表结构里的关键字。别名表的优先级高于所有自动
            # 规则，如果放任 "Total" 这样的值进来，被映射的商户行就会被
            # 当成汇总行，Total 公式只统计到那一行为止，整张表就废了。
            # 这里在加载时就把非法规则挡掉，比在每个调用点各判断一次可靠。
            if merchant.upper() in MERCHANT_STRUCTURAL_LABELS:
                continue

            rules.append((keyword.casefold(), merchant))
        wb.close()
    except Exception:
        rules = []

    # 关键字长的排前面："amazon prime" 必须比 "amazon" 先匹配，
    # 否则 Amazon Prime 会被并进 Amazon。
    rules.sort(key=lambda item: len(item[0]), reverse=True)

    MERCHANT_ALIAS_CACHE.update({"path": str(path), "mtime": mtime, "rules": rules})
    return rules


def apply_merchant_alias(name: str, rules: Optional[List[Tuple[str, str]]] = None) -> str:
    """命中别名表则返回标准名，否则返回空字符串。"""
    text = " ".join(str(name or "").split()).casefold()
    if not text:
        return ""
    for keyword, merchant in (rules if rules is not None else []):
        if keyword in text:
            return merchant
    return ""


# ---------- 交易渠道前缀 ----------
#
# 这些词描述的是"这笔钱怎么付出去的"，不是商户本身，去掉之后剩下的才是
# 真正的商户名：Bill Pay:Orange & Rockland -> Orange & Rockland、
# POS Mac Amazon Seattle WA -> Amazon Seattle WA。
#
# 只匹配开头，且只列出确定属于渠道标记的词，避免误伤以这些字眼开头的
# 真实商户名。
MERCHANT_CHANNEL_PREFIX_RE = re.compile(
    r"^(?:"
    r"bill\s*pay(?:ment)?\s*[:\-]?\s*"
    r"|billpay\s*[:\-]?\s*"
    r"|pos\s+mac\s+"
    r"|pos\s+debit\s+"
    r"|pos\s+"
    r"|checkcard\s+"
    r"|check\s*card\s+"
    r")+",
    re.IGNORECASE,
)


def strip_merchant_channel_prefix(name: str) -> str:
    """去掉 Bill Pay: / POS Mac 这类交易渠道前缀。

    去完之后如果一个字母都不剩（例如 "POS 12345" 只剩 "12345"），说明这个
    名字里本来就没有真正的商户信息，去前缀反而让它更难辨认，此时退回原名。
    """
    original = str(name or "").strip()
    result = MERCHANT_CHANNEL_PREFIX_RE.sub("", original).strip()
    if not result or not any(ch.isalpha() for ch in result):
        return original
    return result


def clean_merchant_display_name(name: str, alias_rules: Optional[List[Tuple[str, str]]] = None) -> str:
    """清理商户名里无意义的流水号、交易码和域名后缀。

    典型效果：
        amazon.com91591732948173          -> amazon
        Shell247592917239                 -> Shell
        Expida143242//4/24/com            -> Expida
        POS Mac Amazon.Com*Nx6 Seattle WA -> POS Mac Amazon Seattle WA
        Bill Pay:Veolia Water New Y 200045 6Bu1T6Kr -> Bill Pay:Veolia Water New Y

    设计上刻意保守，只删除"确定无意义"的部分，绝不把商户名截断成第一个
    单词。像 Bill Pay:Veolia Water New Y、No1 Sushi 88 Corkbbo ACH、
    Orange & Rockland、Internet Trf To Client-Added Transfer Account 这些
    名字里本来就带数字或标点，必须原样保留，否则清洗反而把数据毁了。

    万一某个名字被清洗后什么都不剩，一律退回原始名字，宁可不清洗也不能
    产生空白商户。
    """
    original = " ".join(str(name or "").split())
    if not original:
        return ""

    # 第一优先：用户自己写在别名表里的规则。这是人的明确判断，
    # 优先级高于下面所有自动规则，包括银行保护。
    alias = apply_merchant_alias(original, alias_rules)
    if alias:
        return alias

    # 去掉 Bill Pay: / POS Mac 这类渠道前缀，剩下的才是真正的商户。
    # 这一步放在银行判断之前：像 "Bill Pay:Chase4568" 既要去掉前缀，
    # 又要保住 4568，两件事不冲突。
    working = strip_merchant_channel_prefix(original)

    # 银行 / 转账 / 付款类的名字到此为止：里面的数字通常是账户尾号或
    # 卡号后四位，正是判断分类要用的信息，删掉就没法分辨了。
    if is_bank_transfer_merchant(working):
        return working

    cleaned_tokens = []
    for token in working.split(" "):
        original_token = token

        # 整词就是流水号或纯分隔符的，直接丢弃。
        if _is_pure_number_token(token) or MERCHANT_SEPARATOR_TOKEN_RE.match(token):
            continue

        # 星号后面是交易码，从星号处截断（Amazon.Com*Nx6 -> Amazon.Com）。
        star = token.find("*")
        if star != -1:
            token = token[:star]

        if _is_reference_code_token(token):
            continue

        token = MERCHANT_LONG_DIGITS_RE.sub("", token)
        token = MERCHANT_DOMAIN_SUFFIX_RE.sub("", token)
        token = MERCHANT_DATE_FRAGMENT_RE.sub("", token)

        # 上面几步会留下连续的标点，例如 Expida// ，压平后再去掉首尾标点。
        token = re.sub(r"[\\/*#|]{2,}", " ", token)
        token = token.strip(" \t-–—/\\*#|,;:.")

        if not token or MERCHANT_SEPARATOR_TOKEN_RE.match(token):
            continue

        # 删掉流水号后只剩一个字母的，多半是编号的前缀而不是名字的一部分，
        # 例如门店号 "T-2109" 会剩下一个孤零零的 "T"。这种残渣一并丢掉，
        # 得到干净的 "Target" 而不是 "Target T"。
        # 判断条件里要求原词确实含有被删掉的长数字，这样单独成词的 "A"、
        # "&" 之类不会受影响。
        if len(token) == 1 and MERCHANT_LONG_DIGITS_RE.search(original_token):
            continue

        cleaned_tokens.append(token)

    result = " ".join(" ".join(cleaned_tokens).split())
    if not result:
        return working

    # 清洗结果不能变成表结构里的关键字，否则会被误认成汇总行。
    if result.upper() in MERCHANT_STRUCTURAL_LABELS:
        return working

    return result


def clean_debit_merchant_names_in_sheet(xlsx_path: Path, sheet_name: str):
    """把指定 Debit 工作表 Column A 的商户名整体清洗一遍。

    只处理表头行与 Total 行之间的商户行；Total 行本身、以及它下面的
    BEGIN / ADD / LESS / END 汇总行都不会被碰到。Credit 页面同样不处理。

    清洗后若出现同名行（例如多行 Amazon.Com*XXX 被统一成同一个名字），
    会直接复用既有的合并功能把它们并成一行 —— 这样金额累加、Category
    后面的额外月份列、Total 公式、以及 ADD/END 汇总块的引用修复全都沿用
    同一套已验证过的逻辑，不另写一遍。

    返回 (改名列表, 合并结果)；改名列表元素为 (原名, 新名)。
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists() or not sheet_name:
        return [], None

    renamed = []
    wb = load_workbook(xlsx_path, data_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            return [], None
        ws = wb[sheet_name]

        # 类型判断沿用全局唯一的那条规则，Credit 页面直接跳过。
        if classify_sheet_type(ws) != "debit":
            return [], None

        layout = find_debit_layout(ws)
        if layout is None:
            return [], None
        header_row, merchant_col, _, _, _ = layout

        # 别名表只读一次，整列共用。
        alias_rules = load_merchant_alias_rules()

        cleaned_names = []
        for row in range(header_row + 1, ws.max_row + 1):
            raw_name = normalized_cell_text(ws.cell(row=row, column=merchant_col).value)
            if not raw_name:
                continue
            if raw_name.upper() in {"TOTAL", "GRAND TOTAL"}:
                break

            new_name = clean_merchant_display_name(raw_name, alias_rules)
            if new_name and new_name != raw_name:
                ws.cell(row=row, column=merchant_col, value=new_name)
                renamed.append((raw_name, new_name))
            cleaned_names.append((new_name or raw_name).casefold())

        if not renamed:
            return [], None

        wb.save(xlsx_path)
    finally:
        wb.close()

    # 清洗之后可能出现重名行，交给既有的合并功能处理。
    merge_result = None
    if len(cleaned_names) != len(set(cleaned_names)):
        merge_result = merge_duplicate_merchants_in_selected_sheet(xlsx_path, sheet_name)

    clear_debit_layout_cache(xlsx_path)
    return renamed, merge_result


def merge_duplicate_merchants_in_selected_sheet(xlsx_path: Path, sheet_name: str):
    """在指定工作表中原地合并重复 Merchant。

    先完整读取并汇总所有重复行的数据，再删除重复行。这样即使月份/日期列位于
    Total 或 Category 后面（例如 6-Aug），其中的数据也不会因为先删行而丢失。
    其他工作表不会被修改。
    """
    if not xlsx_path.exists():
        raise FileNotFoundError(f"找不到 Excel 文件：{xlsx_path}")
    if not sheet_name:
        raise ValueError("请先在 UI 中选择Excel Sheet。")

    wb = load_workbook(xlsx_path, data_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Excel 中找不到Excel Sheet：{sheet_name}")

        ws = wb[sheet_name]
        layout = find_debit_layout(ws)
        if layout is None:
            raise ValueError("当前Excel Sheet中找不到可识别的 Debit 月份表格。")

        header_row, merchant_col, period_cols, total_col, category_col = layout
        period_cols = sorted(set(period_cols))
        if not period_cols:
            raise ValueError("当前Excel Sheet中找不到可合并的月份或日期列。")

        # Some templates have an extra recognizable period/date column placed
        # AFTER Total/Category (e.g. a freshly added "Apr-26" column for next
        # year -- see find_extra_debit_period_columns). Total/SUM formulas
        # intentionally never include these columns (that rule is unchanged
        # below), but the actual per-row transaction VALUES stored in them
        # still belong to that merchant and must be combined the same way as
        # any other period column when two rows for the same Merchant get
        # merged into one. period_cols alone (used for Total) does NOT include
        # these trailing columns, so relying on it for the merge step was the
        # bug: a duplicate row's amount in a trailing column like this was
        # silently discarded the moment the duplicate row got deleted below.
        extra_period_cols = (
            find_extra_debit_period_columns(ws, category_col)
            if category_col is not None else []
        )
        merge_cols = sorted(set(period_cols) | set(extra_period_cols))

        # 明细区域截止到 Merchant 列中的 Total 行之前。
        summary_row = None
        for row in range(header_row + 1, ws.max_row + 1):
            key = normalized_header_key(ws.cell(row=row, column=merchant_col).value)
            if key in {"TOTAL", "GRAND TOTAL"}:
                summary_row = row
                break
        data_end_row = summary_row - 1 if summary_row else ws.max_row

        # 先做完整快照。不能一边读取一边删除，否则 Category 后面的期间数据
        # 可能跟随重复行一起被删除。快照范围是 merge_cols（真正的期间列 +
        # Category 后面的额外期间列），而不只是参与 Total 求和的 period_cols。
        groups = OrderedDict()
        for row in range(header_row + 1, data_end_row + 1):
            merchant_value = ws.cell(row=row, column=merchant_col).value
            merchant_text = normalized_cell_text(merchant_value)
            if not merchant_text:
                continue

            key = merchant_text.casefold()
            if key not in groups:
                groups[key] = {
                    "keeper": row,
                    "rows": [],
                    "period_values": {col: [] for col in merge_cols},
                }

            groups[key]["rows"].append(row)
            for col in merge_cols:
                # 保存原始值/公式；此时尚未删除任何行。
                groups[key]["period_values"][col].append(
                    ws.cell(row=row, column=col).value
                )

        duplicate_rows = []
        merged_groups = 0

        # 先把每一组的所有期间数据写入保留行，包含 Category 后面的额外期间列。
        for data in groups.values():
            rows = data["rows"]
            if len(rows) <= 1:
                continue

            merged_groups += 1
            keep_row = data["keeper"]
            duplicate_rows.extend(rows[1:])

            for col in merge_cols:
                all_parts = []
                for value in data["period_values"][col]:
                    all_parts.extend(split_formula_parts(value))
                ws.cell(row=keep_row, column=col).value = (
                        build_plus_formula(all_parts) or None
                )

            # Total 只计算真正参与聚合的期间列 (period_cols)；Category 后面
            # 的额外期间列按既定规则永远不计入 Total/SUM，这一点保持不变。
            ws.cell(row=keep_row, column=total_col).value = build_sum_formula_for_columns(
                ws, keep_row, period_cols
            )

        # 所有 Category 后面的月份数据已经汇总写入后，才从底部删除重复行。
        for row in sorted(duplicate_rows, reverse=True):
            ws.delete_rows(row, 1)

        # 删除后重新定位 Total 行，并重建明细 Total 与底部汇总公式。
        new_summary_row = None
        for row in range(header_row + 1, ws.max_row + 1):
            key = normalized_header_key(ws.cell(row=row, column=merchant_col).value)
            if key in {"TOTAL", "GRAND TOTAL"}:
                new_summary_row = row
                break

        if new_summary_row is not None:
            first_data_row = header_row + 1
            last_data_row = new_summary_row - 1
            if last_data_row >= first_data_row:
                for row in range(first_data_row, last_data_row + 1):
                    ws.cell(row=row, column=total_col).value = build_sum_formula_for_columns(
                        ws, row, period_cols
                    )

                for col in period_cols:
                    letter = excel_col_letter(col)
                    ws.cell(row=new_summary_row, column=col).value = (
                        f"=SUM({letter}{first_data_row}:{letter}{last_data_row})"
                    )

                ws.cell(row=new_summary_row, column=total_col).value = (
                    build_sum_formula_for_columns(ws, new_summary_row, period_cols)
                )

        # Merging can delete rows and move the Total row upward. Any formula
        # on another sheet (most commonly the paired Credit sheet's DEBIT
        # column, imported from an existing company report) that references
        # this sheet's OLD Total row must be repointed at the new one, or it
        # will keep reading a stale/empty row after the merge.
        if summary_row is not None and new_summary_row is not None:
            merge_offset = new_summary_row - summary_row
            shift_cross_sheet_row_references(wb, sheet_name, summary_row, merge_offset)
            # Some templates also have extra summary rows below Total (BEGIN /
            # ADD / LESS / END) whose formulas pull directly from Total's own
            # cells (e.g. a plain =B14 in the "ADD" row). Those same-sheet
            # references need the same repair.
            shift_local_row_references(
                ws, summary_row, merge_offset, min_row=new_summary_row + 1
            )

        wb.save(xlsx_path)
        return {
            "merged_groups": merged_groups,
            "removed_rows": len(duplicate_rows),
            "period_columns": len(period_cols),
        }
    finally:
        wb.close()


# ================= GUI =================

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

BG = "#1e1e1e"
FG = "#ffffff"
BTN = "#2d2d2d"


def mk_label(parent, text="", **kw):
    params = {"bg": BG, "fg": FG}
    params.update(kw)
    return tk.Label(parent, text=text, **params)


def mk_button(parent, text, cmd):
    return tk.Button(
        parent,
        text=text,
        command=cmd,
        bg=BTN,
        fg=FG,
        activebackground="#3a3a3a",
        relief="raised",
        padx=10,
        pady=6,
        bd=1,
        highlightthickness=0
    )


def run_parser_ui():
    root = tk.Tk()
    root.title("BSDP - Bank Statement Data Processing")
    root.geometry("1280x900")
    root.minsize(1080, 760)
    root.configure(bg=BG)

    mk_label(
        root,
        "BSDP - Bank Statement Data Processing",
        font=("Helvetica", 17, "bold")
    ).pack(anchor="w", padx=14, pady=(12, 10))

    default_monthly_summary = script_dir() / "credit_monthly_summary.xlsx"

    summary_var = tk.StringVar(value=str(default_monthly_summary.resolve()))
    remove_var = tk.StringVar(value="")
    date_format_var = tk.StringVar(value="")
    month_var = tk.StringVar(value="")
    bank_name_var = tk.StringVar(value=DEFAULT_BANK_NAME)
    sheet_var = tk.StringVar(value="")
    # 下面三个变量对应的输入控件已从界面移除，但变量本身保留：写入流程仍会
    # 照常读取它们，只是现在恒为默认值 —— Bank Name 为空（即不往 Credit 表
    # 第一行写银行名）、余额列走自动判断、无分隔符日期识别关闭。这样处理
    # 逻辑一行都不用改，将来想把某个选项放回界面也只需重新加控件。
    balance_mode_var = tk.StringVar(value="自动判断")
    compact_date_var = tk.BooleanVar(value=False)

    sheet_names: List[str] = []
    months: List[str] = []
    period_positions: List[int] = []
    current_sheet_type = [""]

    # ---------- 顶部文件和解析设置 ----------
    filebar = tk.Frame(root, bg=BG)
    filebar.pack(fill="x", padx=14, pady=(0, 8))
    filebar.columnconfigure(1, weight=1)

    mk_label(filebar, "月份总表 Excel 保存位置(Monthly summary location)：").grid(
        row=0, column=0, sticky="w", pady=3
    )
    ent_summary = tk.Entry(
        filebar, textvariable=summary_var, bg="#111", fg=FG,
        insertbackground=FG, relief="flat"
    )
    ent_summary.grid(row=0, column=1, padx=(8, 8), sticky="we", pady=3)

    mk_label(filebar, "预处理删除内容(Preprocess remove text)：").grid(
        row=1, column=0, sticky="w", pady=3
    )
    ent_remove = tk.Entry(
        filebar, textvariable=remove_var, bg="#111", fg=FG,
        insertbackground=FG, relief="flat"
    )
    ent_remove.grid(row=1, column=1, padx=(8, 8), sticky="we", pady=3)
    mk_label(
        filebar,
        "例如: target, Mcdonald || 不同的关键字请使用逗号隔开",
        fg="#9cdcfe"
    ).grid(row=2, column=1, sticky="w", padx=8, pady=(0, 5))

    mk_label(filebar, "日期示例或格式(Date example / format)：").grid(
        row=3, column=0, sticky="w", pady=3
    )
    ent_date_format = tk.Entry(
        filebar, textvariable=date_format_var, bg="#111", fg=FG,
        insertbackground=FG, relief="flat"
    )
    ent_date_format.grid(row=3, column=1, padx=(8, 8), sticky="we", pady=3)
    mk_label(
        filebar,
        "直接输入账单里的日期示例，程序会自动转换格式：\n"
        "4-22 → M-DD    May 17 → M DD    17 May → DD M    2025-4-22 → YYYY-M-DD",
        fg="#9cdcfe", justify="left", anchor="w"
    ).grid(row=4, column=1, columnspan=2, sticky="w", padx=8, pady=(0, 7))

    # ---------- 文本框及工作表/日期选择 ----------
    body = tk.Frame(root, bg=BG)
    body.pack(fill="both", expand=True, padx=10, pady=(4, 4))

    control_row = tk.Frame(body, bg=BG)
    control_row.pack(fill="x", pady=(0, 6))

    mk_label(control_row, "Excel sheet:", font=("Helvetica", 12, "bold")).pack(
        side="left", padx=(0, 5)
    )
    sheet_box = ttk.Combobox(
        control_row, textvariable=sheet_var, values=sheet_names,
        width=22, state="readonly"
    )
    sheet_box.pack(side="left", padx=(0, 14))

    mk_label(control_row, "Month:", font=("Helvetica", 12, "bold")).pack(
        side="left", padx=(0, 5)
    )
    month_box = ttk.Combobox(
        control_row, textvariable=month_var, values=months,
        width=12, state="readonly"
    )
    month_box.pack(side="left", padx=(0, 14))

    mk_label(
        control_row,
        "在下方文本框输入账单(Please paste your bank statement data below.)",
        fg="#9cdcfe"
    ).pack(side="left", padx=(4, 10))

    txt_frame = tk.Frame(body, bg=BG)
    txt_frame.pack(fill="both", expand=True)

    txt_input = tk.Text(
        txt_frame, wrap="word", bg="#111", fg=FG,
        insertbackground=FG, relief="flat", undo=True
    )
    txt_input.pack(side="left", fill="both", expand=True)

    sb = tk.Scrollbar(txt_frame, command=txt_input.yview)
    sb.pack(side="right", fill="y")
    txt_input.configure(yscrollcommand=sb.set)

    def clear_textbox():
        txt_input.delete("1.0", tk.END)
        txt_input.edit_reset()
        txt_input.focus_set()
        log_status("文本框已清空 (Text box cleared)")

    # 右上角操作按钮区域：Start 和清空文本框会在函数定义完成后加入
    top_action_buttons = tk.Frame(control_row, bg=BG)
    top_action_buttons.pack(side="right", padx=(10, 0))

    # 状态栏的进度渲染函数定义在后面，但 refresh_sheet_and_month_options 需要
    # 在读完工作表后顺手更新进度。用一个 hook 占位解耦：启动阶段第一次调用时
    # 它还是空的（那时状态栏尚未建好），之后由状态栏区域填入。
    status_hooks = {"apply_progress": None}

    def refresh_sheet_and_month_options(preferred_sheet: str = "", preferred_month: str = ""):
        path = get_summary_output_path(summary_var.get())
        names = list_workbook_sheet_names(path)
        sheet_names[:] = names
        sheet_box["values"] = sheet_names

        previous_sheet = sheet_var.get().strip()
        previous_month = preferred_month or month_var.get().strip()

        selected = preferred_sheet if preferred_sheet in sheet_names else ""
        if not selected and previous_sheet in sheet_names:
            selected = previous_sheet
        if not selected and sheet_names:
            # On first load/import, skip cover/notes sheets and select the first
            # actually writable Credit/Debit sheet.
            selected = find_first_writable_sheet_name(path, sheet_names) or sheet_names[0]
        sheet_var.set(selected)

        # 一次读取拿全：Month 选项、各月是否已写入、各月金额合计。
        # 之前这里和状态栏各开一次工作簿，同样的数据读了两遍 —— 切换 Sheet 时
        # 等于把成本翻倍，商户多的表能明显感觉到卡顿。
        sheet_type, options, written, totals = (
            read_sheet_period_status(path, selected)
            if selected else ("", [], {}, {})
        )
        current_sheet_type[0] = sheet_type
        months[:] = [label for label, _ in options]
        period_positions[:] = [position for _, position in options]
        month_box["values"] = months

        if previous_month in months:
            month_var.set(previous_month)
            month_box.current(months.index(previous_month))
        elif months:
            month_var.set(months[0])
            month_box.current(0)
        else:
            month_var.set("")
            month_box.set("")

        # 复用刚才那次读取的结果刷新进度，不再另开一次工作簿。
        apply_progress = status_hooks.get("apply_progress")
        if apply_progress is not None:
            apply_progress(selected, options, written, totals)

    def choose_summary():
        p = filedialog.askopenfilename(
            title="选择需要继续写入的 Excel（支持多个Excel Sheet）",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if p:
            summary_var.set(p)
            # A newly chosen workbook should select its first writable sheet,
            # not reuse a same-named sheet/month from the previous workbook.
            sheet_var.set("")
            month_var.set("")
            refresh_sheet_and_month_options()
            log_status(f"已选择 Excel: {Path(p).name}")

    browse_btn = mk_button(filebar, "选择 Excel(Browse)", choose_summary)
    browse_btn.grid(row=0, column=2, sticky="e", pady=3)

    def on_month_selected(event=None):
        detected = current_sheet_type[0].title() if current_sheet_type[0] else "Unknown"
        log_status(
            f"已选择Excel Sheet: {sheet_var.get()} | 日期: {month_var.get()} | "
            f"自动识别: {detected}"
        )

    def on_sheet_selected(event=None):
        # Changing sheets refreshes periods from that exact sheet only.
        # refresh_sheet_and_month_options 内部已经顺带刷新过进度了，
        # 这里不再重复读取工作簿。
        refresh_sheet_and_month_options(sheet_var.get(), "")
        on_month_selected()

    sheet_box.bind("<<ComboboxSelected>>", on_sheet_selected)
    month_box.bind("<<ComboboxSelected>>", on_month_selected)

    refresh_sheet_and_month_options()

    def keep_cursor_visible(event=None):
        txt_input.see("insert")
        return None

    txt_input.bind("<KeyRelease>", keep_cursor_visible)
    txt_input.bind("<ButtonRelease-1>", keep_cursor_visible)
    txt_input.bind("<MouseWheel>", keep_cursor_visible)
    txt_input.bind("<Return>", keep_cursor_visible)
    txt_input.bind("<<Paste>>", keep_cursor_visible)

    # ---------- 状态显示：上次写入 / 本表月份进度 / 最近操作 ----------
    status_frame = tk.Frame(root, bg=BG)
    status_frame.pack(fill="x", padx=12, pady=(5, 2))
    mk_label(status_frame, "状态(Status)：").pack(anchor="w")

    status_text = tk.Text(
        status_frame,
        height=4,
        wrap="none",
        bg=BG,
        fg="#9cdcfe",
        insertbackground="#9cdcfe",
        relief="flat",
        bd=0,
        highlightthickness=0,
        takefocus=0,
        state="disabled",
        # 等宽字体：月份和金额需要按列对齐，人工核对时才好一眼扫过去。
        font=("Courier New", 9),
    )
    status_text.pack(fill="x", anchor="w")

    # 三部分内容各自独立维护，互不覆盖：
    #   last_write     最近一次成功写入的 Month / 笔数 / 金额 —— 用户最关心
    #   progress_lines 当前 Sheet 各月份完成情况（直接读 Excel，重启也准）
    #   recent         最近一次操作或错误提示
    status_state = {
        "last_write": "尚未写入",
        "progress_lines": [],
        "recent": "尚未开始 (Not started)",
    }

    LABEL_WIDTH = 10  # "上次写入"等前缀的对齐宽度

    def render_status():
        pad = " " * LABEL_WIDTH
        lines = [f"{'上次写入':<6}  {status_state['last_write']}"]

        progress = status_state["progress_lines"] or ["(尚未选择 Sheet)"]
        lines.append(f"{'本表进度':<6}  {progress[0]}")
        for extra in progress[1:]:
            lines.append(pad + extra)

        lines.append(f"{'最近操作':<6}  {status_state['recent']}")

        # 高度随内容变化：只写了一两个月份时不占地方，写满一年也能完整显示。
        status_text.configure(state="normal", height=max(3, min(len(lines), 9)))
        status_text.delete("1.0", tk.END)
        status_text.insert("1.0", "\n".join(lines))
        status_text.configure(state="disabled")

    def log_status(message: str):
        """记录一条最近操作（带时间戳），不影响另外两部分。"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        status_state["recent"] = f"[{timestamp}] {message}"
        render_status()

    def set_last_write(sheet_name: str, month_label: str, count: int, amount=None):
        timestamp = datetime.now().strftime("%H:%M:%S")
        parts = [f"【{month_label}】", f"{count} 笔"]
        if amount is not None:
            parts.append(f"合计 {format_amount_for_display(amount)}")
        status_state["last_write"] = (
            "  ".join(parts) + f"   →  {sheet_name}   ({timestamp})"
        )
        render_status()

    def apply_progress(sheet_name, options, written, totals):
        """用已经读好的数据渲染进度，不再重复打开工作簿。"""
        if not sheet_name:
            status_state["progress_lines"] = []
        else:
            status_state["progress_lines"] = format_period_progress_lines(
                sheet_name, options, written, totals
            )
        render_status()

    status_hooks["apply_progress"] = apply_progress

    def refresh_progress(sheet_name: str = "", month_position=None):
        """重新统计当前 Sheet 的月份完成情况，并返回指定月份的金额合计。

        直接读 Excel 而不是靠会话内的记录：换台电脑、重开程序，甚至别人先
        写过一部分，进度都能正确显示出来。读不到就留空，不打扰用户。
        """
        target = (sheet_name or sheet_var.get()).strip()
        if not target:
            status_state["progress_lines"] = []
            render_status()
            return None
        try:
            path = get_summary_output_path(summary_var.get())
            _, options, written, totals = read_sheet_period_status(path, target)
            apply_progress(target, options, written, totals)
            return totals.get(month_position) if month_position is not None else None
        except Exception:
            status_state["progress_lines"] = []
            render_status()
            return None

    render_status()
    refresh_progress()
    refresh_progress()

    def import_existing_report(clear_existing_data: bool):
        mode_name = "清空旧金额（模板模式）" if clear_existing_data else "保留全部原有数据"
        source = filedialog.askopenfilename(
            title=f"选择已有公司报表 Excel - {mode_name}",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not source:
            return

        try:
            target = get_summary_output_path(summary_var.get())
            action_text = (
                "可识别页面中的旧月份金额会被清空，但格式、公式框架、Merchant、Category、"
                "Excel Sheet名称和顺序会保留。"
                if clear_existing_data
                else
                "源 Excel 中的所有原有数据、公式、格式、Excel Sheet名称和顺序都会完整保留。"
            )

            if target.exists():
                overwrite = messagebox.askyesno(
                    "目标 Excel 已存在",
                    f"目标文件已经存在：\n{target}\n\n"
                    f"导入模式：{mode_name}\n{action_text}\n\n"
                    "目标文件将被这次导入结果覆盖，是否继续？"
                )
                if not overwrite:
                    return
            else:
                proceed = messagebox.askyesno(
                    "确认导入模式",
                    f"导入模式：{mode_name}\n\n{action_text}\n\n是否继续？"
                )
                if not proceed:
                    return

            log_status(f"正在导入已有报表：{mode_name}...")
            income_count, expense_count, imported_month_labels, imported_sheet_names = import_existing_report_template(
                Path(source), target,
                bank_name=bank_name_var.get().strip() or DEFAULT_BANK_NAME,
                clear_existing_data=clear_existing_data,
            )
            # STRICT ORDER: generation/save is complete before UI reads the generated workbook.
            summary_var.set(str(target))
            clear_debit_layout_cache(target)
            sheet_var.set("")
            month_var.set("")
            first_writable_sheet = find_first_writable_sheet_name(target, imported_sheet_names)
            refresh_sheet_and_month_options(first_writable_sheet)
            log_status(
                f"导入完成：{target.name} | 模式: {mode_name} | "
                f"共 {len(imported_sheet_names)} 个Excel Sheet | 当前: {sheet_var.get()}"
            )

            result_note = (
                "可识别页面中的历史月份金额已清空；其他内容已保留。"
                if clear_existing_data
                else
                "所有原有数据、公式和格式均已完整保留，可直接继续追加。"
            )
            messagebox.showinfo(
                "导入完成",
                f"目标 Excel：\n{target}\n\n"
                f"导入模式：{mode_name}\n"
                f"导入Excel Sheet：{len(imported_sheet_names)} 个\n"
                f"Excel Sheet名称：{', '.join(imported_sheet_names)}\n"
                f"Credit 项目：{income_count}\n"
                f"Debit Merchant：{expense_count}\n\n"
                f"{result_note}\n"
                f"UI 日期顺序：{', '.join(months)}"
            )
        except Exception as e:
            log_status(f"模板导入失败：{type(e).__name__}: {e}")
            messagebox.showerror("导入失败", f"{type(e).__name__}: {e}")

    def start():
        try:
            content = txt_input.get("1.0", "end-1c")
            custom_date_input = date_format_var.get().strip()
            resolved_date_format = resolve_date_input_to_format(custom_date_input)
            configure_date_format(resolved_date_format)

            selected_month_display = month_var.get().strip()
            selected_month_index = month_box.current()
            if (
                    selected_month_index < 0
                    or selected_month_index >= len(months)
                    or selected_month_index >= len(period_positions)
            ):
                messagebox.showerror("错误", "请选择有效日期/月份")
                log_status("未开始：没有选择有效的日期/月份")
                return
            selected_period_position = period_positions[selected_month_index]

            selected_sheet = sheet_var.get().strip()
            if not selected_sheet:
                messagebox.showerror("错误", "请选择要修改的Excel Sheet/Page")
                log_status("未开始：没有选择 Excel Sheet")
                return

            selected_sheet_type = current_sheet_type[0]
            if selected_sheet_type not in ("credit", "debit"):
                messagebox.showerror(
                    "错误",
                    "当前 Excel Sheet 无法自动识别为 Credit 或 Debit。\n"
                    "请选择包含可识别月份/日期结构的 Sheet。"
                )
                log_status(f"未开始：Sheet「{selected_sheet}」无法识别为 Credit 或 Debit")
                return

            bank_name = bank_name_var.get().strip() or DEFAULT_BANK_NAME
            remove_items = parse_remove_items(remove_var.get())
            log_status(
                f"正在处理... Excel Sheet: {selected_sheet} | 日期: {selected_month_display} | "
                f"自动识别: {selected_sheet_type.title()}"
            )

            # 第一步：按原有方式删除用户指定的关键词。
            preprocessed_text = preprocess_statement_text(content, remove_items)

            # 第二步：直接解析一次（这就是接入自动整理之前的老流程），
            # 结果只作为对照和兜底，保证新增的自动整理永远不会让识别结果变差。
            configure_date_format(resolved_date_format)
            direct_hits = parse_auto(preprocessed_text)

            # 第三步：自动整理后再解析一次。
            #
            # 注意这里必须把日期规则切回默认：整理后的文本日期已经被统一成
            # 零填充的 MM/DD，如果继续套用用户填写的自定义日期格式
            # （例如 M-DD），反而会因为格式对不上而解析失败。
            balance_mode = {
                "自动判断": "auto",
                "没有余额列": "none",
                "最后一个是余额": "last_is_balance",
            }.get(balance_mode_var.get(), "auto")

            pre_options = build_preprocessor_options(
                date_format_input=custom_date_input,
                balance_column_mode=balance_mode,
                allow_compact_mmdd=bool(compact_date_var.get()),
            )
            cleaned_text, pre_result = run_statement_preprocessor(
                preprocessed_text, pre_options
            )
            configure_date_format("")
            auto_hits = parse_auto(cleaned_text)

            # 第四步：取识别更多的那一套。自动整理只有在确实更好时才会被采用，
            # 否则自动退回老流程，绝不会因为多了这一步而少写入交易。
            if len(auto_hits) >= len(direct_hits):
                hits = auto_hits
                mode_note = "自动整理"
            else:
                hits = direct_hits
                mode_note = "直接解析（自动整理结果更少，已自动回退）"
                # 回退时把日期规则恢复成用户填写的设置，保持与老流程一致。
                configure_date_format(resolved_date_format)

            # 第五步：整理过程中如果有内容没能识别，先告诉用户再写入，
            # 避免这些内容被静默丢掉。
            if mode_note == "自动整理" and pre_result is not None and pre_result.issues:
                preview = "\n".join(
                    f"  第 {','.join(str(n) for n in issue.line_nos)} 行：{issue.reason} - {issue.content[:50]}"
                    for issue in pre_result.issues[:6]
                )
                more = (f"\n  ...另有 {len(pre_result.issues) - 6} 条"
                        if len(pre_result.issues) > 6 else "")
                hint_text = ("\n\n" + "\n".join(pre_result.hints)) if pre_result.hints else ""
                proceed = messagebox.askyesno(
                    "有内容未能识别",
                    f"自动整理时有 {len(pre_result.issues)} 处内容无法识别，"
                    f"这些内容不会被写入 Excel：\n\n{preview}{more}{hint_text}\n\n"
                    f"将写入 {len(hits)} 笔交易，是否继续？"
                )
                if not proceed:
                    log_status("已取消写入（存在未识别内容）")
                    return

            rows = [(h.merchant, h.amount, h.who) for h in hits]
            summary_xlsx = get_summary_output_path(summary_var.get())

            detected_type = append_rows_to_selected_sheet(
                summary_xlsx, selected_sheet, rows,
                selected_month_display, selected_period_position,
                expected_sheet_type=selected_sheet_type,
                bank_name=bank_name,
            )

            # 写入成功后，对 Debit 页面的商户名做一次清洗：去掉流水号、
            # 交易码和域名后缀，并把因此变成同名的行合并起来。
            #
            # 放在写入之后而不是写入之前，是因为清洗要覆盖整列 —— 既包括
            # 这次刚写进去的商户，也包括表里原本就存在的旧商户，让同一个
            # 商户不会因为一个带流水号、一个不带而长期占着两行。
            # Credit 页面不做处理（那里的商户是列标题，结构完全不同）。
            renamed_merchants = []
            clean_merge_result = None
            if detected_type == "debit":
                renamed_merchants, clean_merge_result = clean_debit_merchant_names_in_sheet(
                    summary_xlsx, selected_sheet
                )

            # 写入成功后自动清空输入框，方便直接粘贴下一批账单。
            #
            # 位置很关键：这行必须放在 append_rows_to_selected_sheet 成功返回
            # 之后。前面任何一步抛异常、或者用户在"有内容未能识别"的确认框里
            # 选择了取消，都会在到达这里之前就退出，原始数据因此得以保留 ——
            # 否则用户就得回银行网站重新复制一遍。
            #
            # 这里刻意不调用 edit_reset()：文本框创建时开启了 undo，保留撤销
            # 记录意味着万一写错了 Sheet 或月份，按 Ctrl+Z 就能把刚才粘贴的
            # 内容找回来，不至于因为自动清空而丢失。
            txt_input.delete("1.0", tk.END)
            txt_input.focus_set()

            clean_note = ""
            if renamed_merchants:
                clean_note = f" | 已清洗 {len(renamed_merchants)} 个商户名"
                if clean_merge_result and clean_merge_result.get("merged_groups"):
                    clean_note += f"，合并 {clean_merge_result['merged_groups']} 组重复"

            # 用户最关心的是"哪个月写完了"，所以单独用一行突出显示，
            # 不再和处理方式、文件名等细节挤在一起。
            # 先刷新进度（顺带取回该月份写完后的金额合计），再更新"上次写入"，
            # 这样金额和进度里显示的是同一次读取的结果，不会出现两处对不上。
            month_total = refresh_progress(selected_sheet, selected_period_position)
            set_last_write(selected_sheet, selected_month_display, len(rows), month_total)
            log_status(
                f"完成 {detected_type.title()} 写入（{mode_note}）{clean_note}，文本框已清空"
            )
            date_format_display = (
                resolved_date_format if resolved_date_format
                else "默认格式 MM/DD、MM/DD/YY、MM/DD/YYYY"
            )
            extra_note = ""
            if pre_result is not None and mode_note == "自动整理":
                extra_note = f"未识别内容: {len(pre_result.issues)} 处\n"
                if pre_result.balance_column_detected:
                    extra_note += "已自动排除行尾余额列\n"
            if renamed_merchants:
                extra_note += f"商户名清洗: {len(renamed_merchants)} 个\n"
                if clean_merge_result and clean_merge_result.get("merged_groups"):
                    extra_note += (
                        f"清洗后合并重复: {clean_merge_result['merged_groups']} 组"
                        f"（减少 {clean_merge_result['removed_rows']} 行）\n"
                    )
            messagebox.showinfo(
                "Completed",
                f"提取成功: {len(rows)} 笔交易\n"
                f"处理方式: {mode_note}\n"
                f"{extra_note}"
                f"当前Excel Sheet: {selected_sheet}\n"
                f"当前日期: {selected_month_display}\n"
                f"自动识别类型: {detected_type.title()}\n"
                "写入方式: 保留原有数据并继续追加\n"
                f"日期格式: {date_format_display}\n"
                f"总表文件: {summary_xlsx.name}\n"
                f"分类规则文件: {category_rules_path()}\n"
                "\n文本框已自动清空，可直接粘贴下一批账单。\n"
                "如需找回刚才的内容，在文本框内按 Ctrl+Z 即可撤销。"
            )
        except Exception as e:
            messagebox.showerror("异常 / Error", f"{type(e).__name__}: {e}")
            log_status(f"解析失败 / Failed: {type(e).__name__}: {e}")

    def merge_current_sheet():
        try:
            selected_sheet = sheet_var.get().strip()
            if not selected_sheet:
                messagebox.showerror("错误", "请先选择要合并的Excel Sheet")
                return

            summary_xlsx = get_summary_output_path(summary_var.get())
            confirm = messagebox.askyesno(
                "确认合并",
                f"将在当前 Excel 中直接处理Excel Sheet：\n{selected_sheet}\n\n"
                "执行前请关闭 Excel 文件，是否继续？"
            )
            if not confirm:
                return

            log_status(f"正在合并当前Excel Sheet：{selected_sheet}...")
            result = merge_duplicate_merchants_in_selected_sheet(
                summary_xlsx, selected_sheet
            )
            refresh_sheet_and_month_options(selected_sheet)

            log_status(
                f"合并完成 | Excel Sheet: {selected_sheet} | "
            )
            messagebox.showinfo(
                "合并完成",
                f"Excel Sheet：{selected_sheet} 合并完成\n"

            )
        except PermissionError:
            log_status("合并失败：Excel 文件可能正在打开")
            messagebox.showerror(
                "无法保存",
                "Excel 文件可能正在打开。请关闭 Excel 后再执行。"
            )
        except Exception as e:
            log_status(f"合并失败：{type(e).__name__}: {e}")
            messagebox.showerror("合并失败", f"{type(e).__name__}: {e}")

    def create_default_excel():
        try:
            target = get_summary_output_path(summary_var.get())

            if target.exists():
                overwrite = messagebox.askyesno(
                    "目标 Excel 已存在",
                    f"目标文件已经存在：\n{target}\n\n"
                    "生成默认表格会覆盖这个文件。是否继续？"
                )
                if not overwrite:
                    return

            log_status("正在生成默认 Excel（Credit + Debit，Jan-Dec）...")
            create_default_monthly_workbook(
                target,
                bank_name=bank_name_var.get().strip() or DEFAULT_BANK_NAME,
            )
            summary_var.set(str(target))
            clear_debit_layout_cache(target)
            refresh_sheet_and_month_options("Credit Summary")
            log_status(
                f"默认 Excel 已生成 | {target.name} | "
                "2 个 Sheet | Jan-Dec"
            )
            messagebox.showinfo(
                "生成完成",
                f"默认 Excel 已生成：\n{target}\n\n"
                "Sheet 1: Credit Summary\n"
                "Sheet 2: debit summary\n"
                "月份: Jan - Dec"
            )
        except PermissionError:
            log_status("生成失败：Excel 文件可能正在打开")
            messagebox.showerror(
                "无法保存",
                "目标 Excel 文件可能正在打开。请关闭 Excel 后再执行。"
            )
        except Exception as e:
            log_status(f"生成默认 Excel 失败：{type(e).__name__}: {e}")
            messagebox.showerror("生成失败", f"{type(e).__name__}: {e}")

    # ---------- 右上角按钮：按照界面布局放置 Start 和清空文本框 ----------
    mk_button(top_action_buttons, "Start", start).pack(side="left", padx=(0, 10))
    mk_button(top_action_buttons, "清空文本框", clear_textbox).pack(side="left")

    # ---------- 底部按钮：左侧导入、真正居中的默认表格、右侧合并 ----------
    # 左右按钮组使用 pack；中间按钮使用 place(relx=0.5)，这样它的位置
    # 永远以整个 footer 的几何中心为准，不会因为左侧按钮更宽而偏移。
    footer = tk.Frame(root, bg=BG, height=44)
    footer.pack(fill="x", padx=10, pady=(6, 12))
    footer.pack_propagate(False)

    import_buttons = tk.Frame(footer, bg=BG)
    import_buttons.pack(side="left", anchor="w")
    mk_button(import_buttons, "导入并清空旧金额", lambda: import_existing_report(True)).pack(
        side="left", padx=(0, 8)
    )
    mk_button(import_buttons, "导入并保留全部数据", lambda: import_existing_report(False)).pack(
        side="left", padx=(0, 8)
    )

    merge_buttons = tk.Frame(footer, bg=BG)
    merge_buttons.pack(side="right", anchor="e")
    mk_button(merge_buttons, "合并当前 Sheet", merge_current_sheet).pack(side="right")

    default_buttons = tk.Frame(footer, bg=BG)
    default_buttons.place(relx=0.5, rely=0.5, anchor="center")
    mk_button(default_buttons, "生成默认表格", create_default_excel).pack()

    root.mainloop()


if __name__ == "__main__":
    try:
        run_parser_ui()
    except Exception as e:
        print(f"程序异常：{e}", file=sys.stderr)
        raise
